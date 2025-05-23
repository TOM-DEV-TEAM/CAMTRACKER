from django.contrib import messages
import pandas as pd
from admin_datta.utils import JsonResponse
from django.db import transaction
from django.http import HttpResponse, HttpResponseBadRequest
from admin_datta.forms import LoginForm
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import  authenticate, login
#from numpy import error_message
from openpyxl.workbook import Workbook
from .forms import MouvementForm, ChauffeurForm, CamionForm, TransitaireForm, ClientForm, VehiculeForm
def index(request,id):
    user=Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index.html',
                  { 'util' :user})
def tables(request):
  context = {
    'segment': 'tables'
  }
  return render(request, "pages/dynamic-tables.html", context)
def detail_mvt(request):
  context = {
    'segment': 'index',
  }
  return render(request, "pages/detail.html", context)
def tout_mvt(request):
  tout_mvt=Mouvement1.objects.all()
  ids_camion = set(tout_mvt.values_list('camion_id', flat=True))
  ids_point = set(tout_mvt.values_list('camion_id', flat=True))
  data = []
  for doc in tout_mvt:
    camion=Camion.objects.get(id_cam=doc.camion_id)
    doc.id_cam=camion.id_cam
    doc.immat=camion.immatriculation
    doc.ttrans=camion.transporteur
    ptr_etr=Utilisateurs.objects.get(id_user=doc.pointeur_entree)
    doc.ptr_etr= ptr_etr.nom + ptr_etr.prenom
    if doc.pointeur_sortie :
      ptr_srt = Utilisateurs.objects.get(id_user=doc.pointeur_entree)
      doc.ptr_srt = ptr_srt.nom + ptr_srt.prenom
    else:
      doc.ptr_srt = 'Auncun'
  context = {
    'segment': 'index',
    'mvt' : tout_mvt
  }
  return render(request, "pages/detail.html", context)
def mvt_termine(request):
  tout_mvt_crs=Mouvement1.objects.filter(date_sortie__isnull=False)
  ids_camion = set(tout_mvt.values_list('camion_id', flat=True))
  ids_point = set(tout_mvt.values_list('camion_id', flat=True))
  ids_camion = set(tout_mvt.values_list('camion_id', flat=True))
  ids_point = set(tout_mvt.values_list('camion_id', flat=True))
  # Ajouter dynamiquement l'attribut 'acces=0' aux documents restreints dans listedoc
  data = []
  moins_de_30_minutes = []
  plus_de_30_minutes = []
  for doc in tout_mvt:
    camion=Camion.objects.get(id_cam=doc.camion_id)
    doc.id_cam=camion.id_cam
    doc.immat=camion.immatriculation
    doc.ttrans=camion.transporteur
    ptr_etr=Utilisateurs.objects.get(id_user=doc.pointeur_entree)
    doc.ptr_etr= ptr_etr.nom + ptr_etr.prenom
    ptr_srt = Utilisateurs.objects.get(id_user=doc.pointeur_entree)
    doc.ptr_srt = ptr_srt.nom + ptr_srt.prenom
    duree= (doc.date_sortie - doc.date_entree)/60
    if duree <= 30:
      moins_de_30_minutes.append(doc)
    else:
      plus_de_30_minutes.append(doc)
  context = {
    'segment': 'index',
    'mvt' : tout_mvt,
    'moins': moins_de_30_minutes,
    'plus': plus_de_30_minutes
    # 'products' : Product.objects.all()
  }
  return render(request, "pages/detail.html", context)
def mvt_termine(request):
  tout_mvt_crs=Mouvement1.objects.filter(date_sortie__isnull=True)
  ids_camion = set(tout_mvt.values_list('camion_id', flat=True))
  ids_point = set(tout_mvt.values_list('camion_id', flat=True))
  ids_camion = set(tout_mvt.values_list('camion_id', flat=True))
  ids_point = set(tout_mvt.values_list('camion_id', flat=True))
  # Ajouter dynamiquement l'attribut 'acces=0' aux documents restreints dans listedoc
  data = []
  moins_de_30_minutes = []
  plus_de_30_minutes = []
  for doc in tout_mvt:
    camion=Camion.objects.get(id_cam=doc.camion_id)
    doc.id_cam=camion.id_cam
    doc.immat=camion.immatriculation
    doc.ttrans=camion.transporteur
    ptr_etr=Utilisateurs.objects.get(id_user=doc.pointeur_entree)
    doc.ptr_etr= ptr_etr.nom + ptr_etr.prenom
    ptr_srt = Utilisateurs.objects.get(id_user=doc.pointeur_entree)
    doc.ptr_srt = ptr_srt.nom + ptr_srt.prenom
    now = timezone.now()  # Obtenir la date et l'heure actuelles
    difference = now - doc.date_entree
    duree=difference.total_seconds() / 60
    if duree <= 30:
      moins_de_30_minutes.append(doc)
    else:
      plus_de_30_minutes.append(doc)
  context = {
    'segment': 'index',
    'mvt' : tout_mvt,
    'moins':moins_de_30_minutes,
    'plus' :plus_de_30_minutes
    # 'products' : Product.objects.all()
  }
  return render(request, "pages/detail.html", context)
def fetch_mvt_termine(request):
  tout_mvt_crs=Mouvement1.objects.filter(date_sortie__isnull=False)
  ids_camion = set(tout_mvt.values_list('camion_id', flat=True))
  ids_point = set(tout_mvt.values_list('camion_id', flat=True))
  ids_camion = set(tout_mvt.values_list('camion_id', flat=True))
  ids_point = set(tout_mvt.values_list('camion_id', flat=True))
  # Ajouter dynamiquement l'attribut 'acces=0' aux documents restreints dans listedoc
  data = []
  moins_de_30_minutes = []
  plus_de_30_minutes = []
  for doc in tout_mvt:
    camion=Camion.objects.get(id_cam=doc.camion_id)
    doc.id_cam=camion.id_cam
    doc.immat=camion.immatriculation
    doc.ttrans=camion.transporteur
    ptr_etr=Utilisateurs.objects.get(id_user=doc.pointeur_entree)
    doc.ptr_etr= ptr_etr.nom + ptr_etr.prenom
    ptr_srt = Utilisateurs.objects.get(id_user=doc.pointeur_entree)
    doc.ptr_srt = ptr_srt.nom + ptr_srt.prenom
    duree= (doc.date_sortie - doc.date_entree)/60
    if duree <= 30:
      moins_de_30_minutes.append(doc)
    else:
      plus_de_30_minutes.append(doc)

  context = {
    'segment': 'index',
    'mvt' : tout_mvt,
    'moins': moins_de_30_minutes,
    'plus': plus_de_30_minutes
    # 'products' : Product.objects.all()
  }
  return render(request, "pages/detail.html", context)



###################### Index Controller ##################################
################### RENDER INDEX CONTROLLER ##############################
def index_controlleur(request, id):
    # Récupérer l'utilisateur en fonction de l'ID
    user = get_object_or_404(Utilisateurs, id_user=id)
    # Rendre le template avec les données
    return render(request, 'pages/index_controlleur.html', { 'util': user})
################### RENDER CAMIONS ZUD ##############################
def camion_zud(request, id):
    # Récupérer l'utilisateur en fonction de l'ID
    user = get_object_or_404(Utilisateurs, id_user=id)
    # Rendre le template avec les données
    return render(request, 'pages/camion_zud.html', { 'util': user})
################### LOGIQUE INDEX CONTROLLER ##############################
def liste_observations(request, id):
    # Récupérer l'utilisateur en fonction de l'ID
    user = get_object_or_404(Utilisateurs, id_user=id)
    try:
        observations = Observation.objects.select_related('user', 'camion')
        data = []
        for observation in observations:
            if observation.bd == 1:
                observation.bd ='Oui'
            else:
                observation.bd = 'Non'
            if observation.dd == 1:
                observation.dd ='Oui'
            else:
                observation.dd = 'Non'
            if observation.enpanne == 1:
                observation.enpanne ='Oui'
            else:
                observation.enpanne = 'Non'
            observation_data = {
                'id_observation': observation.id_observation,
                'date': observation.date.strftime("%Y-%m-%d %H:%M:%S"),
                'bd': observation.bd,
                'dd': observation.dd,
                'enpanne': observation.enpanne,
                'motif_stationnement': observation.motif_stationnement,
                'camion': {
                    'id_cam': observation.camion.id_cam,
                    'immatriculation': observation.camion.immatriculation,
                    'transporteur': observation.camion.transporteur,
                    'type': observation.camion.type
                },
                'user': {
                    'id_user': observation.user.id_user,
                    'fullname': observation.user.fullname,
                },
            }
            data.append(observation_data)
    except Exception as ex:
        return redirect (f'/index_controller/{user.id_user}?error={ex}')
    # Rendre le template avec les données
    return render(request, 'pages/liste_observations.html', { 'util': user, 'data': data})
#################################### LOGIQUE RENDER OBSERVATIONS ADMIN ######################
def liste_observationsadmin(request, id):
    # Récupérer l'utilisateur en fonction de l'ID
    user = get_object_or_404(Utilisateurs, id_user=id)
    # Rendre le template avec les données
    return render(request, 'pages/liste_observationsadmin.html', { 'util': user})
from django.utils.timezone import now
def fetch_controlleur(request):
    try:
        mouvements = Mouvement0.objects.filter(
            destination='zud',
            date_sortie__isnull=True,
            camion__type__in=['VRAC', 'PORTE-CONTENEUR', 'PLATEAU', 'PORTE-CHAR']
        ).select_related('chauffeur', 'transitaire', 'camion')

        data = []
        for mouvement in mouvements:
            duration_seconds = (now() - mouvement.date_entree).total_seconds()
            hours = int(duration_seconds // 3600)
            minutes = int((duration_seconds % 3600) // 60)
            seconds = int(duration_seconds % 60)
            heures_passees = f"{hours:02}:{minutes:02}:{seconds:02}"
            data.append({
                'id_mouvement': mouvement.id_mvt,
                'heures_passees': heures_passees,
                'camion': {
                    'id_cam': mouvement.camion.id_cam,
                    'immatriculation': mouvement.camion.immatriculation,
                    'transporteur': mouvement.camion.transporteur,
                    'type': mouvement.camion.type
                },
                'chauffeur': {
                    'fullname': mouvement.chauffeur.fullname if mouvement.chauffeur else None,
                    'telephone': mouvement.chauffeur.telephone if mouvement.chauffeur else None
                },
                'transitaire': {
                    'fullname': mouvement.transitaire.fullname if mouvement.transitaire else None,
                    'telephone': mouvement.transitaire.telephone if mouvement.transitaire else None
                }
            })
        return JsonResponse({'mouvements': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)       ######Gestion Camion
#################### FETCH OBSERVATION ####################################################

def fetch_observation(request):
    try:
        observations = Observation.objects.select_related('user', 'camion')
        data = []
        for observation in observations:
            observation_data = {
                'id_observation': observation.id_observation,
                'date': observation.date.strftime("%Y-%m-%d %H:%M:%S"),
                'bd': observation.bd,
                'dd': observation.dd,
                'enpanne': observation.enpanne,
                'motif_stationnement': observation.motif_stationnement,
                'camion': {
                    'id_cam': observation.camion.id_cam,
                    'immatriculation': observation.camion.immatriculation,
                    'transporteur': observation.camion.transporteur,
                    'type': observation.camion.type
                },
                'utilisateurs': {
                    'id_user': observation.user.id_user,
                    'fullname': observation.user.fullname,
                },
            }
            data.append(observation_data)

        return JsonResponse({'observations': data})

    except Exception as e:
        print(f"Error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
        ########################################AGENT CONTROLER#######################################################
def ajout_observation(request):
    if request.method == 'POST':
        id_mvt = int(request.POST.get('id_mvt'))
        id_cam = request.POST.get('id_cam')
        dd = request.POST.get('dd')
        bd = request.POST.get('bd')
        enpanne = request.POST.get('enpanne')
        commentaire = request.POST.get('commentaire')
        motif_stationnement = request.POST.get('motif_stationnement')
        statut = request.POST.get('statut')
        user_id = request.POST.get('id_user')
        try:
            user = get_object_or_404(Utilisateurs, id_user=user_id)
            # Check if an observation exists
            observation, created = Observation.objects.update_or_create(
                id_mvt_0=Mouvement0.objects.get(id_mvt=id_mvt),
                camion=Camion.objects.get(id_cam=id_cam),
                defaults={
                    'bd': int(bd) if bd else None,
                    'dd': int(dd) if dd else None,
                    'enpanne': int(enpanne) if enpanne else None,
                    'commentaire': str(commentaire) if commentaire else None,
                    'motif_stationnement': str(motif_stationnement) if motif_stationnement else None,
                    'user': user,
                }
            )
            # Update the status of the movement
            mouvement0 = get_object_or_404(Mouvement0, id_mvt=id_mvt)
            mouvement0.statut_entree = statut
            mouvement0.save()
            return redirect(f"/index_controller/{user_id}?success=true")
        except Exception as e:
            error_message = str(e)
            return redirect(f"/index_controller/{user_id}?error={error_message}")
################################### AJOUT OBSERVATION #########################################
def liste_camion(request, id):
 user = Utilisateurs.objects.get(id_user=id)
 listeboite = Camion()
 if(user.poste=='A1'):
     listeboite=Camion.objects.filter(type='SEMI-REMORQUE')
 elif (user.poste=='A2'):
     listeboite=Camion.objects.filter(type='VRAC')
 elif (user.poste == 'A0' or user.poste == 'A3'):
    listeboite=Camion.objects.all()

 #user=2
 return render(request, 'pages/gestion_camion.html', {'boite': listeboite,'util': user})
#####Gestion User

 def Crer_agent_page(request, id):
   # forms = UtilisateurForm(request.POST)
   user = Utilisateurs.objects.get(id_utilisateur=id)
   # return render(request, 'templatetra/crer_agent.html', {'forms': forms, 'util': user})
   return render(request, 'templatetra/crer_agent.html', {'util': user})


def enregistrer_agent(request):
  if request.method == 'POST':
    # form = UtilisateuwrForm(request.POST)
    form = 1
    user = Utilisateurs.objects.get(id_utilisateur=request.POST["id_user"])
    if form.is_valid() or 1:
      cleaned_data = form.cleaned_data
      instance = Utilisateurs(
        nom=request.POST['nom'],
        prenom=request.POST['prenom'],
        email=request.POST['email'],
        password='passer',
        telephone=request.POST['telephone'],
        role=request.POST['rolee']

        # form.cleaned_data['direction']
      )
      instance.save()
      # return redirect('/users/login/')
      return redirect('/liste_user/{user.id_utilisateur}')
    else:
      # forms = UtilisateurForm()
      return Crer_agent_page(request, request.POST["id_user"])
      # return render(request, 'templatetra/crer_agent.html', {'forms': forms})


def update_agent_page(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  return render(request, 'templatetra/update_agent.html', {'agt': agt, 'util': user, 'user': agt})


def upadte_agent(request):
  agt = Utilisateurs.objects.get(id_utilisateur=request.POST['id_agt'])
  user = Utilisateurs.objects.get(id_utilisateur=request.POST['id_user'])
  if request.POST['prenom']:
    agt.prenom = request.POST['prenom']
  if request.POST['nom']:
    agt.email = request.POST['nom']
  if request.POST['telephone']:
    agt.telephone = request.POST['telephone'],

  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')


def desactiver_agent(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  agt.etat = 0
  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')


def activer_agent(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  agt.etat = 1
  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')
def reinitiliaser_mdp(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  agt.password = 'reinit'
  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')
##############Import Fichier Camion
@csrf_exempt
def importer_donnees_camion_excel(fichier_excel):
    """
    Importe les données d'un fichier Excel dans la table Camion.

    Args:
        fichier_excel (UploadedFile): Le fichier Excel à importer.

    Raises:
        ValueError: Si le fichier Excel n'est pas valide.
    """

    if not fichier_excel.content_type.startswith('application/vnd.openxmlformats'):
        raise ValueError('Le fichier Excel n\'est pas valide.')

    # Lire le fichier Excel dans un DataFrame Pandas
    df = pd.read_excel(fichier_excel)

    # Vérifier si les colonnes du DataFrame correspondent aux champs du modèle Camion
    champs_camion = ['immatriculation', 'transporteur','type']
    if not df.columns.isin(champs_camion).all():
        raise ValueError('Le fichier Excel ne contient pas les colonnes requises.')

    # Enregistrer les données dans la base de données
    with transaction.atomic():
        for index, row in df.iterrows():
            camion = Camion(
                immatriculation=row['immatriculation'],
                transporteur=row['transporteur'],
                type=row['type'],
            )
            camion.save()


@csrf_exempt
def importer_donnees_camion(request):
    if request.method == 'POST':
        fichier_excel = request.FILES['fichier_excel']
        try:
            importer_donnees_camion_excel(fichier_excel)
            message = 'Données importées avec succès.'
        except ValueError as e:
            message = e.args[0]
    else:
        fichier_excel = None
        message = ''
    #user=Utilisateurs.objects.get(poste='admin')
    user = Utilisateurs.objects.get(id_user=1)
    return  redirect("/liste_amion/" +str(user.id_user))
def ajouter_camion_page(request):
  return render(request, 'pages/ajouter_camion.html',)
@csrf_exempt
def ajouter_camion(request):
  if request.method == 'POST':
    #form = BoiteForm(request.POST)
    #id_user = request.session.get('user_id')
    if Camion.objects.filter(immatriculation=request.POST['imat']).exists():
      return redirect()
      # cleaned_data = form.cleaned_data
    else:
      camion = Camion(
        immatriculation=request.POST['imat'],
        transporteur=request.POST['trans'],
          type=request.POST['type'],
      )
      camion.save()
      return redirect("/liste_amion/" + str(request.POST['id_user']))
  else:
    # return ajouterboite_page(request, request.POST['id_user'])
    return redirect("/doc/jouterboite_page/" + str(request.POST['id_user']))
def modifier_boite(request):
    if request.method == 'POST':
      boite_id = request.POST['boite_id']
      boite = get_object_or_404(Camion, id_cam=boite_id)
      boite.immatriculation = request.POST['immat']
      boite.transporteur = request.POST['trans']
      boite.type = request.POST['type']
      boite.save()
      #return redirect(reverse('index'))  # Redirigez vers la page index après la modification
      return redirect("/liste_amion/" + str(request.POST['id_user']))
    return render(request, 'index.html')
#####Gestion Utilisateur
def list_user(request, id):
     listeboite=Utilisateurs.objects.all()
     user=Utilisateurs.objects.get(id_user=id)
     return render(request, 'pages/gestion_utilisateurs.html', {'boite': listeboite,'util' : user})
##############Import Fichier Camion
def importer_donnees_utilisateur_excel(fichier_excel):
    """
    Importe les données d'un fichier Excel dans la table Camion.
    Args:
        fichier_excel (UploadedFile): Le fichier Excel à importer.

    Raises:
        ValueError: Si le fichier Excel n'est pas valide.
    """

    if not fichier_excel.content_type.startswith('application/vnd.openxmlformats'):
        raise ValueError('Le fichier Excel n\'est pas valide.')
    # Lire le fichier Excel dans un DataFrame Pandas
    df = pd.read_excel(fichier_excel)
    # Vérifier si les colonnes du DataFrame correspondent aux champs du modèle Camion
    champs_camion = ['fullname', 'matricule', 'email', 'password', 'poste','statut']
    if not df.columns.isin(champs_camion).all():
        raise ValueError('Le fichier Excel ne contient pas les colonnes requises.')
    # Enregistrer les données dans la base de données
    with transaction.atomic():
        for index, row in df.iterrows():
            camion = Utilisateurs(
              fullname=row['fullname'],
                matricule=row['matricule'],
              email=row['email'],
              password=row['password'],
              poste=row['poste'],
                status=row['statut'],

            )
            camion.save()
def importer_donnees_utilisateur(request):
    if request.method == 'POST':
        fichier_excel = request.FILES['fichier_excel']
        try:
            importer_donnees_utilisateur_excel(fichier_excel)
            message = 'Données importées avec succès.'
        except ValueError as e:
            message = e.args[0]
    else:
        fichier_excel = None
        message = ''
    return  redirect("/liste_user/2")
def ajouter_utilisateur_page(request):
  return render(request, 'pages/ajouter_camion.html',)
@csrf_exempt
def ajouter_user(request):
  if request.method == 'POST':
    user = Utilisateurs.objects.get(id_user=request.POST['id_user'])
    if Utilisateurs.objects.filter(email=request.POST['email']).exists():
        return redirect("/liste_user/" + str(user.id_user))
    else:
      camion = Utilisateurs(
        #nom=request.POST['nom'],
        fullname=request.POST['prenom'],
        email=request.POST['email'],
          matricule=request.POST['matricule'],
        password=request.POST['password'],
        poste=request.POST['poste'],
      )
      camion.save()
  user=Utilisateurs.objects.get(id_user=request.POST['id_user'])
  return redirect("/liste_user/" + str(user.id_user))
  #else:
    # return ajouterboite_page(request, request.POST['id_user'])
    #return redirect("/doc/jouterboite_page/" + str(request.POST['id_user']))
def modifier_user(request):
    if request.method == 'POST':
      boite_id = request.POST['boite_id']
      boite = get_object_or_404(Utilisateurs, id_user=boite_id)
      #boite.nom = request.POST['nom']
      boite.fullname= request.POST['prenom']
      boite.email = request.POST['email']
      boite.password = request.POST['password']
      boite.poste = request.POST['poste']
      boite.matricule = request.POST['matricule']
      boite.save()
      #return redirect(reverse('index'))  # Redirigez vers la page index après la modification
      #user=Utilisateurs.objects.get(poste='superadmin')
      return redirect("/liste_user/" + str(request.POST['id_user']))
    return redirect("/liste_user/"+ str(request.POST['id_user']))
############Gestion Detail
def mouvements_superieur_20_minutes(request, id):
    vingt_minutes = timedelta(minutes=20)
    maintenant = timezone.now()
    mouvements = Mouvement1.objects.filter(date_sortie__isnull=True).values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur',
                                                 'remorque','date_entree','pointeur_entree_id')
    mouvements = list( mouvements)
    mouvements_filtrés = [mvt for mvt in mouvements if (maintenant - mvt['date_entree']) > vingt_minutes]
    ma = len(mouvements_filtrés)
    mouvements_filtrés=list(mouvements_filtrés)
    for doc in  mouvements_filtrés:
      camion = Camion.objects.get(id_cam=doc['camion_id'])
      doc['camion'] = camion
      ptr_etr = Utilisateurs.objects.get(id_user=doc['pointeur_entree_id'])
      doc['pointeur ']= ptr_etr
    return JsonResponse(list(mouvements_filtrés), safe=False)
    #return render(request, "pages/detail_urgent.html", {'mvt' :mouvements_filtrés ,'lg' :lg})
    #return mouvements_filtrés
def detail_urgent(request,id):
  duree_dk = 30
  delais_urg = 20
  maintenant = timezone.now()
  try:
      para = ParametrageDelais.objects.get(entite='icdtom', type='SEMI-REMORQUE')
      duree_dk = timedelta(minutes=int(para.delais_maximal))
      delais_urg = timedelta(minutes=int(para.delais_urgent))
  except:
      # pass
      duree_dk = timedelta(minutes=30)
      delais_urg = timedelta(minutes=20)
  mouvements = Mouvement1.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)
  mouvements_filtrés = [mvt for mvt in mouvements if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk ]
  for mvt in mouvements_filtrés :
      chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
      mvt.chauffeur_name = chauf.fullname
      poineur= Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
      mvt.pointeur=poineur.fullname
      camion=Camion.objects.get(id_cam=mvt.camion_id)
      mvt.trans=camion.transporteur
      mvt.imat=camion.immatriculation
  lg=len(mouvements_filtrés)
  total_encours=len(mouvements)
  if total_encours >= 1 :
    pourcentage_urgent=(lg/total_encours) * 100
  else :
    pourcentage_urgent=0
  user = Utilisateurs.objects.get(id_user=id)
  context = {
    'segment': 'index',
    'mvt' : mouvements_filtrés,
    # 'products' : Product.objects.all()
      'lg' :lg,
      'pourcentage_urgent' :int(pourcentage_urgent),
      'util': user
  }
  return render(request, "pages/detail_urgent.html", context)
def detail_depassemnt(request, id):

    duree_dk = 3
    delais_urg = 20
    maintenant = timezone.now()

    try:

        para = ParametrageDelais.objects.get(entite='icdtom', type='SEMI-REMORQUE')
        duree_dk = timedelta(minutes=int(para.delais_maximal))
        delais_urg = timedelta(minutes=int(para.delais_urgent))
    except:
        # pass
        duree_dk = timedelta(minutes=30)
        delais_urg = timedelta(minutes=20)
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)
    mouvements_filtrés = [mvt for mvt in mouvements if (maintenant - mvt.date_entree) > duree_dk]
    for mvt in mouvements_filtrés:
        chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauf.fullname
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
    lg = len(mouvements_filtrés)
    total_depassemnt=len(mouvements)
    if total_depassemnt >= 1:
       pourcentage_depassement = (lg / total_depassemnt) * 100
    else :
      pourcentage_depassement= 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'pourcentage_depassement' : int(pourcentage_depassement),
        'util': user
    }

    return render(request, "pages/detail_depassement.html", context)

def detail_plus_30(request, id):

    duree_dk = 30
    delais_urg = 20
    maintenant = timezone.now()

    try:

        para = ParametrageDelais.objects.get(entite='icdtom', type='SEMI-REMORQUE')
        duree_dk = timedelta(minutes=int(para.delais_maximal))
        delais_urg = timedelta(minutes=int(para.delais_urgent))
    except:
        # pass
        duree_dk = timedelta(minutes=30)
        delais_urg = timedelta(minutes=20)
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False,date_sortie__isnull=False)
    mouvements_filtrés = [mvt for mvt in mouvements if (mvt.date_sortie - mvt.date_entree) > duree_dk]
    for mvt in mouvements_filtrés:
        chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauf.fullname
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
    lg = len(mouvements_filtrés)
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    if total_plus_30 >= 1:
      pourcentage_plus_30 = (lg / total_plus_30) * 100
    else :
      pourcentage_plus_30=0
    user = Utilisateurs.objects.get(id_user=id)

    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'util': user,
        'pourcentage_plus_30' : int( pourcentage_plus_30)
    }

    return render(request, "pages/detail_plus_30.html", context)
def detail_moins_30(request, id):

    duree_dk = 30
    delais_urg = 20
    maintenant = timezone.now()

    try:

        para = ParametrageDelais.objects.get(entite='icdtom', type='SEMI-REMORQUE')
        duree_dk = timedelta(minutes=int(para.delais_maximal))
        delais_urg = timedelta(minutes=int(para.delais_urgent))
    except:
        # pass
        duree_dk = timedelta(minutes=30)
        delais_urg = timedelta(minutes=20)
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False,date_sortie__isnull=False)
    mouvements_filtrés = [mvt for mvt in mouvements if (mvt.date_sortie - mvt.date_entree) <= duree_dk]
    for mvt in mouvements_filtrés:
        chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauf.fullname
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    if  total_moins_30 >= 1 :
       pourcentage_moins_30 = (lg / total_moins_30) * 100
    else :
        pourcentage_moins_30=0
    user = Utilisateurs.objects.get(id_user=id)

    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'pourcentage_moins_30' :int(pourcentage_moins_30),
        'util': user
    }

    #return render(request, "pages/detail_depassement.html", context)
    return render(request, "pages/detail_moins_30.html", context)

@csrf_exempt
###Gesrion Connexion
def seconnecter(request):
    forms = LoginForm()
    if request.method == 'POST':
        username = request.POST['email']
        password = request.POST['password']
        user = None
        try:
            user = Utilisateurs.objects.get(email=username, password=password)
            if user.status == 'active':
                ################### SUPER ADMIN ##################
                if user.poste == 'SA':
                    return redirect("/liste_user/" + str(user.id_user))
                #################### UTILISATEURS DKLOG ###############
                elif user.poste == 'E0':
                    return redirect("/entreedecalog_view/" + str(user.id_user))
                elif user.poste == 'A0':
                    return redirect("/index_dk_log0/" + str(user.id_user))

                elif user.poste == 'S0':
                    return redirect("/sortie_decalog_view/" + str(user.id_user))
                elif user.poste == 'C0':
                    return redirect("/index_controller/" + str(user.id_user))
                ##################### ICD TOM ###########################
                elif user.poste == 'A1':
                    return redirect("/index/" + str(user.id_user))
                if user.poste == 'E1':
                    return redirect("/index_entree_icdtom/" + str(user.id_user))
                elif user.poste == 'S1':
                    return redirect("/index_sortie_icdtom/" + str(user.id_user))
                ######################## SACHERIE ##############################
                elif user.poste == 'E2':
                    return redirect("/index2_view/" + str(user.id_user))
                elif user.poste == 'S2':
                    return redirect("/index3_view/" + str(user.id_user))
                elif user.poste == 'A2':
                    return redirect("/index_sacherie/" + str(user.id_user))
                ######################## ZUD ##############################
                elif user.poste == 'S3':
                    return redirect("/index_sortie_zud/" + str(user.id_user))
                elif user.poste == 'E3':
                    return redirect("/index_entree_zud/" + str(user.id_user))
                elif user.poste == 'A3':
                    return redirect("/index_admin_zud/" + str(user.id_user))
                #################### ICD CMA ######################
                elif user.poste == 'E4':
                    return redirect("/index_entree_icdcma/" + str(user.id_user))
                elif user.poste == 'S4':
                    return redirect("/index_sortie_icdcma/" + str(user.id_user))
                else:
                    return render(request, 'pages/login2.html')
            else:
                messages.error(request, 'Votre compte est désactivé !! ')
                return render(request, 'pages/login2.html')
        except Utilisateurs.DoesNotExist:
            messages.error(request, 'Login ou mot de passe incorrect!!')
            return render(request, 'pages/login2.html')
    else:
        return render(request, 'pages/login2.html')
        #################### La vue qui affiche la page login au démarrage de l'application #######################
def login_page(request):
  forms = LoginForm()
  if request.method == 'POST':
    forms = LoginForm(request.POST)
    if forms.is_valid() or 1:
      email = forms.cleaned_data['email']
      password = forms.cleaned_data['password']
      user = authenticate(email=email, password=password,status='active')
      if user:
        login(request, user)
        return redirect('dashboard')
  context = {'form': forms}
  return render(request, 'pages/login2.html', context)

######Gestio Post Entree
def liste_mouvements(request, id_user):
    mouvements = Mouvement1.objects.filter(date_entree__isnull=True, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Vérifier camion_id
        camion_id = mouvement.get('camion_id')
        if camion_id:
            camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        else:
            camion = {'id_cam': 'non assignée', 'immatriculation': 'non assignée', 'transporteur': 'non assignée', 'type': 'non assignée'}
        mouvement['camion'] = camion

        # Vérifier chauffeur_id
        chauffeur_id = mouvement.get('chauffeur_id')
        if chauffeur_id:
            chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        else:
            chauffeur = {'id_chauffeur': 'non assignée', 'fullname': 'non assignée', 'permis': 'non assignée'}
        mouvement['chauffeur'] = chauffeur

        # Remplacer les autres champs nulls par 'non assignée'
        for key, value in mouvement.items():
            if value is None:
                mouvement[key] = 'non assignée'

    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)
################# LISTE MOUVEMENTS ENTREE CMA #########################
def liste_mouvements_entree_cma(request, id_user):
    mouvements = Mouvement4.objects.filter(date_entree__isnull=True, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Vérifier camion_id
        camion_id = mouvement.get('camion_id')
        if camion_id:
            camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        else:
            camion = {'id_cam': 'non assignée', 'immatriculation': 'non assignée', 'transporteur': 'non assignée', 'type': 'non assignée'}
        mouvement['camion'] = camion

        # Vérifier chauffeur_id
        chauffeur_id = mouvement.get('chauffeur_id')
        if chauffeur_id:
            chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        else:
            chauffeur = {'id_chauffeur': 'non assignée', 'fullname': 'non assignée', 'permis': 'non assignée'}
        mouvement['chauffeur'] = chauffeur

        # Remplacer les autres champs nulls par 'non assignée'
        for key, value in mouvement.items():
            if value is None:
                mouvement[key] = 'non assignée'

    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)
#################### LISTE MOUVEMENTS SORTIE CMA ################################
def liste_mouvements_sortie_cma(request, id_user):
    # Récupérer les mouvements avec des filtres sur date_entree et date_sortie
    mouvements = Mouvement4.objects.filter(date_entree__isnull=False, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Vérifier camion_id et assigner les valeurs du camion ou 'non assignée'
        camion_id = mouvement.get('camion_id')
        if camion_id:
            camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                    'type').first()
        else:
            camion = {'id_cam': 'non assignée', 'immatriculation': 'non assignée', 'transporteur': 'non assignée',
                      'type': 'non assignée'}
        mouvement['camion'] = camion

        # Vérifier chauffeur_id et assigner les valeurs du chauffeur ou 'non assignée'
        chauffeur_id = mouvement.get('chauffeur_id')
        if chauffeur_id:
            chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                                  'permis').first()
        else:
            chauffeur = {'id_chauffeur': 'non assignée', 'fullname': 'non assignée', 'permis': 'non assignée'}
        mouvement['chauffeur'] = chauffeur

        # Remplacer toutes les autres valeurs nulles ou vides par 'non assignée'
        for key, value in mouvement.items():
            if value is None or value == '':
                mouvement[key] = 'non assignée'

    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }

    return JsonResponse(response_data)

############# RENDER POINTEURS ICD TOM ###################
def index_entree_icdtom(request, id_user):
    user = Utilisateurs.objects.get(id_user=id_user)
    return render(request, 'pages/index1.html', {'util': user, })

def index_sortie_icdtom(request, id_user):
    user = Utilisateurs.objects.get(id_user=id_user)
    return render(request, 'pages/index2.html', {'util': user, })
################# RENDER POINTEURS ICD CMA ###################
def index_entree_icdcma(request, id_user):
    user = Utilisateurs.objects.get(id_user=id_user)
    return render(request, 'pages/index_entree_icdcma.html', {'util': user, })

def index_sortie_icdcma(request, id_user):
    user = Utilisateurs.objects.get(id_user=id_user)
    return render(request, 'pages/index_sortie_icdcma.html', {'util': user, })
######################### FETCH LISTE MOUVEMENTS PARTICULIERS ######################
def liste_mouvements_particulier(request):
    mouvements = Mouvement8.objects.filter(date_sortie__isnull=True).values('id_mvt', 'destination', 'vehicule_id', 'date_entree')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        vehicule_id = mouvement['vehicule_id']
        vehicule = Vehicule.objects.filter(id_veh=vehicule_id).values('id_veh', 'immatriculation').first()
        mouvement['vehicule'] = vehicule
    response_data = {
        'mouvements': mouvement_list,
     # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)


def index1_view(request, id_user):
  util = Utilisateurs.objects.get(id_user=id_user)
  camions = Camion.objects.all()
  return render(request, 'pages/index1.html', {'util': util, 'camions': camions})
def sortie_tom(request, id_user):
  util = Utilisateurs.objects.get(id_user=id_user)
  return render(request, 'pages/index2.html', {'util': util})
def entredecalon_view(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)
    mouvements = Mouvement0.objects.filter(destination__icontains='icd', date_sortie__isnull=True).values(
        'id_mvt', 'zone_entree', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupérer les informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}

        # Récupérer les informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        if user_sortie_id:
            user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first()
        else:
            user_sortie = {'fullname': 'null'}
        mouvement['user_srt'] = user_sortie

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

        # Assigner des valeurs par défaut pour les autres champs
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')
        mouvement['date_entree'] = mouvement.get('date_entree', 'Non assigné')
        mouvement['date_sortie'] = mouvement.get('date_sortie', 'Non assigné')
        mouvement['destination'] = mouvement.get('destination', 'Non assigné')

    # Passer la liste des mouvements au template
    return render(request, 'pages/mouvement_entre_0.html', {'util': util, 'mouvements': mouvement_list})
def entreedecalog_particulier(request, id_user):
      util = Utilisateurs.objects.get(id_user=id_user)
      mouvements = Mouvement8.objects.filter(date_sortie__isnull=True).values('id_mvt', 'destination', 'vehicule_id', 'zone_entree',
                                                                              'date_entree')
      mouvement_list = list(mouvements)
      for mouvement in mouvement_list:
          vehicule_id = mouvement['vehicule_id']
          vehicule = Vehicule.objects.filter(id_veh=vehicule_id).values('id_veh', 'immatriculation').first()
          mouvement['vehicule'] = vehicule
      return render(request, 'pages/mouvement_entre_particulier.html', {'util': util, 'mouvements': mouvements})
def entredecalon_view1(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)
    mouvements = Mouvement0.objects.filter(destination__icontains='hangar', date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'zone_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupérer les informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}

        # Récupérer les informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        if user_sortie_id:
            user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first()
        else:
            # Valeur par défaut si pointeur_sortie_id est null
            user_sortie = {'fullname': 'Non assigné'}
        mouvement['user_srt'] = user_sortie

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

        # Assigner des valeurs par défaut pour les autres champs
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')
        mouvement['date_entree'] = mouvement.get('date_entree', 'Non assigné')
        mouvement['date_sortie'] = mouvement.get('date_sortie', 'Non assigné')
        mouvement['destination'] = mouvement.get('destination', 'Non assigné')

    # Passer la liste des mouvements au template
    return render(request, 'pages/mouvement_entre_01.html', {'util': util, 'mouvements': mouvement_list})

def entredecalon_view2(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)
    # Récupérer la liste des mouvements actuels qui remplissent les critères
    mouvements = Mouvement0.objects.filter(
        destination__icontains='zud',
        date_sortie__isnull=True
    ).values(
        'id_mvt', 'camion_id', 'statut_entree', 'zone_entree', 'statut_sortie', 'chauffeur_id', 'remorque',
        'date_entree', 'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'numconteneur1', 'typeconteneur1',
        'numconteneur2', 'typeconteneur2', 'numconteneur3', 'typeconteneur3', 'numconteneur4', 'typeconteneur4', 'numconteneur5', 'typeconteneur5','destination', 'client_id',
        'transitaire_id', 'representant_id'
    )

    # Parcourir chaque mouvement
    for mouvement in mouvements:
        # Vérifier si numconteneur1 est déjà présent dans d'autres mouvements
        mouvement['real_numconteneur1'] = mouvement['numconteneur1']
        if Mouvement0.objects.filter(numconteneur1=mouvement['numconteneur1']).exclude(
                id_mvt=mouvement['id_mvt']).exists():
            mouvement['numconteneur1'] = None  # Affecter à null si le conteneur est déjà affecté
        mouvement['real_numconteneur2'] = mouvement['numconteneur2']
        # Vérifier si numconteneur2 est déjà présent dans d'autres mouvements
        if Mouvement0.objects.filter(numconteneur2=mouvement['numconteneur2']).exclude(
                id_mvt=mouvement['id_mvt']).exists():
            mouvement['numconteneur2'] = None  # Affecter à null si le conteneur est déjà affecté
        mouvement['real_numconteneur3'] = mouvement['numconteneur3']
        # Vérifier si numconteneur3 est déjà présent dans d'autres mouvements
        if Mouvement0.objects.filter(numconteneur3=mouvement['numconteneur3']).exclude(
                id_mvt=mouvement['id_mvt']).exists():
            mouvement['numconteneur3'] = None  # Affecter à null si le conteneur est déjà affecté
        mouvement['real_numconteneur4'] = mouvement['numconteneur4']
        if Mouvement0.objects.filter(numconteneur4=mouvement['numconteneur4']).exclude(
                id_mvt=mouvement['id_mvt']).exists():
            mouvement['numconteneur4'] = None  # Affecter à null si le conteneur est déjà affecté
        mouvement['real_numconteneur5'] = mouvement['numconteneur5']
        if Mouvement0.objects.filter(numconteneur5=mouvement['numconteneur5']).exclude(
                id_mvt=mouvement['id_mvt']).exists():
            mouvement['numconteneur5'] = None  # Affecter à null si le conteneur est déjà affecté

    # Convertir les résultats en liste si nécessaire
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }
        # Récupérer les informations du Client
        client_id = mouvement.get('client_id')
        client = Client.objects.filter(id_client=client_id).values('id_client', 'fullname', 'telephone'
                                                                ).first()
        mouvement['client'] = client if client else {
            'id_client': None,
            'fullname': 'Non assigné',
            'telephone': 'Non assigné'
        }
        # Récuperer les informations du transitaire
        transitaire_id = mouvement.get('transitaire_id')
        transitaire = Transitaire.objects.filter(id_transit=transitaire_id).values('id_transit', 'fullname', 'telephone'
                                                                   ).first()
        mouvement['transitaire'] = transitaire if transitaire else {
            'id_transit': None,
            'fullname': 'Non assigné',
            'telephone': 'Non assigné'
        }
        # Récuperer les informations du représentant
        representant_id = mouvement.get('representant_id')
        representant = Transitaire.objects.filter(id_transit=representant_id).values('id_transit', 'fullname', 'telephone'
                                                                                   ).first()
        mouvement['representant'] = representant if representant else {
            'id_transit': None,
            'fullname': 'Non assigné',
            'telephone': 'Non assigné'
        }
        # Récupérer les informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}

        # Récupérer les informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        if user_sortie_id:
            user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first()
        else:
            # Valeur par défaut si pointeur_sortie_id est null
            user_sortie = {'fullname': 'Non assigné'}
        mouvement['user_srt'] = user_sortie

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

        # Assigner des valeurs par défaut pour les autres champs
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')
        mouvement['date_entree'] = mouvement.get('date_entree', 'Non assigné')
        mouvement['date_sortie'] = mouvement.get('date_sortie', 'Non assigné')
        mouvement['destination'] = mouvement.get('destination', 'Non assigné')

    # Passer la liste des mouvements au template
    return render(request, 'pages/mouvement_entre_02.html', {'util': util, 'mouvements': mouvement_list})

################################# RENDER MODIFICATIONS DES MOUVEMENTS ###############
def modif_mvt(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)

    # Récupérer les mouvements filtrés
    mouvements = Mouvement0.objects.filter(destination__icontains='icd', date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination', 'mission', 'client_id'
    )
    mouvement_list = list(mouvements)

    # Boucle pour enrichir les données des mouvements
    for mouvement in mouvement_list:
        # Récupération des informations du client
        client_id = mouvement.get('client_id')
        client = Client.objects.filter(id_client=client_id).values('id_client', 'fullname', 'telephone').first()
        mouvement['client'] = client if client else {
            'id_client': None,
            'fullname': 'Non Assigné',
            'telephone': 'Non Assigné'
        }

        # Récupération des informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupération des informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}

        # Récupération des informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first() if user_sortie_id else {'fullname': 'null'}
        mouvement['user_srt'] = user_sortie

        # Récupération des informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

    # Préparation du contexte pour le rendu
    context = {
        'util': util,
        'mouvement_list': mouvement_list
    }

    return render(request, 'pages/modif_mvt.html', context)
def modif_mvt1(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)

    # Récupérer les mouvements filtrés
    mouvements = Mouvement0.objects.filter(destination__icontains='hangar', date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination', 'mission', 'client_id'
    )
    mouvement_list = list(mouvements)

    # Boucle pour enrichir les données des mouvements
    for mouvement in mouvement_list:
        # Récupération des informations du client
        client_id = mouvement.get('client_id')
        client = Client.objects.filter(id_client=client_id).values('id_client', 'fullname', 'telephone').first()
        mouvement['client'] = client if client else {
            'id_client': None,
            'fullname': 'Non Assigné',
            'telephone': 'Non Assigné'
        }

        # Récupération des informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupération des informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first() if user_sortie_id else {'fullname': 'null'}
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }
    context = {
        'util': util,
        'mouvement_list': mouvement_list
    }
    return render(request, 'pages/modif_mvt1.html', context)


def modif_mvt2(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)
    mouvements = Mouvement0.objects.filter(date_sortie__isnull=True, destination__contains='zud').values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'destination',
        'marchandise', 'mission', 'client_id', 'representant_id'
    )

    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        # Ajout des détails pour chaque relation avec les noms d'objets fournis
        camion = Camion.objects.filter(id_cam=mouvement['camion_id']).values('id_cam', 'immatriculation',
                                                                             'transporteur', 'type').first()
        mouvement['camion'] = camion

        chauffeur = Chaffeur.objects.filter(id_chauffeur=mouvement['chauffeur_id']).values('id_chauffeur', 'fullname',
                                                                                           'permis').first()
        mouvement['chauffeur'] = chauffeur

        client = Client.objects.filter(id_client=mouvement['client_id']).values('id_client', 'telephone').first()
        mouvement['client'] = client

        representant = Transitaire.objects.filter(id_transit=mouvement['representant_id']).values('id_transit',
                                                                                              'telephone').first()
        mouvement['representant'] = representant

    # Passage des données au template
    return render(request, 'pages/modif_mvt2.html', {'util': util, 'mouvements': mouvement_list})


def modif_mvtpar(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)
    mouvements = Mouvement8.objects.filter(date_sortie__isnull=True).values('id_mvt', 'destination', 'vehicule_id', 'date_entree')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        vehicule_id = mouvement['vehicule_id']
        vehicule = Vehicule.objects.filter(id_veh=vehicule_id).values('id_veh', 'immatriculation').first()
        mouvement['vehicule'] = vehicule
    return render(request, 'pages/modif_mvtpar.html', {'util': util, 'mouvements': mouvement_list})
@csrf_exempt
def ajoutmouvement(request, id_user):
  try:
    util = Utilisateurs.objects.get(id_user=id_user)
  except Utilisateurs.DoesNotExist:
    return HttpResponse("L'utilisateur spécifié n'existe pas.", status=404)
  if request.method == 'POST':
    form = MouvementForm(request.POST)
    if form.is_valid():
      mouvement = form.save(commit=False)
      mouvement.date_entree = timezone.now()
      mouvement.camion_id = request.POST.get('camion')
      mouvement.statut_entree = request.POST.get('statut_entree')
      mouvement.remorque = request.POST.get('remorque')
      mouvement.permis = request.POST.get('permis')
      mouvement.mission = request.POST.get('mission')
      mouvement.pointeur_entree_id = util.id_user
      mouvement.save()
      return redirect(f'/index1_view/{util.id_user}')
  else:
    form = MouvementForm()
  return render(request, 'pages/ajouter_mouvement.html', {'form': form, 'util': util})
#################################### GESTION DES FETCH SUR LE DASHBOARD ##################
####################### FETCH DASHBORD INFORMATIONS PARTICULIERS ###########################
def fetch_statsparticulier(request, id_user):
    maintenant = timezone.now()
    mouvements = (Mouvement8.objects.filter(date_sortie__isnull=True)
                  | Mouvement8.objects.filter(date_sortie__isnull=False))
    urg = 0
    dep = 0
    lg_30 = 0
    lg_moins = 0
    total_cours = 0
    for mvt in mouvements:
        try:
            para = ParametrageDelais.objects.get(entite='dklog', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)
        if mvt.date_sortie is None:
            if(mvt.date_entree):
                pass
            else:
                mvt.date_entree = maintenant
            total_cours += 1
            if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk:
                urg += 1
            elif (maintenant - mvt.date_entree) >= duree_dk:
                dep += 1
        else:
            duree_mouvement = mvt.date_sortie - mvt.date_entree
            if duree_mouvement >= duree_dk:
                lg_30 += 1
            else:
                lg_moins += 1
    totalter = lg_30 + lg_moins
    return JsonResponse({
        'urg': urg,
        'dep': dep,
        'lg_30': lg_30,
        'lg_mois': lg_moins,
        'total_cours': total_cours,
        'total_ter': totalter,
    })
####################### FETCH DASHBORD INFORMATIONS PLT ###########################
def fetch_stats0(request, id_user):
    maintenant = timezone.now()
    # Filtrer les mouvements
    mouvements = (Mouvement0.objects.filter(date_sortie__isnull=True).order_by('date_entree')[:50]
                  | Mouvement0.objects.filter(date_sortie__isnull=False).order_by('date_entree')[:50])  # Si vous avez besoin d'ordonner les résultats, vous pouvez trier la liste
    urg = 0
    dep = 0
    lg_30 = 0
    lg_moins = 0
    total_cours = 0
    for mvt in mouvements:
        try:
            camion = Camion.objects.get(id_cam=mvt.camion_id)
        except:
            camion = Camion(
                id_cam = None,
                immatriculation = 'Non Assigné',
                transporteur = 'Non Assigné',
                type = 'Non Assigné'
            )
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='dklog', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=10080)
            delais_urg = timedelta(minutes=7200)
        if mvt.date_sortie is None:
            # Mouvements en cours
            total_cours += 1
            if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk:
                urg += 1
            elif (maintenant - mvt.date_entree) >= duree_dk:
                dep += 1
        else:
            # Mouvements sortis
            duree_mouvement = mvt.date_sortie - mvt.date_entree
            if duree_mouvement >= duree_dk:
                lg_30 += 1
            else:
                lg_moins += 1

    totalter = lg_30 + lg_moins

    return JsonResponse({
        'urg': urg,
        'dep': dep,
        'lg_30': lg_30,
        'lg_mois': lg_moins,
        'total_cours': total_cours,
        'total_ter': totalter,
    })
####################### FETCH DASHBORD INFORMATIONS ICD TOM ###########################

def fetch_stats(request, id_user):
    maintenant = timezone.now()
    # Filtrer les mouvements
    mouvements = (Mouvement1.objects.filter(date_sortie__isnull=True).order_by('date_entree')[:50]
                  | Mouvement1.objects.filter(date_sortie__isnull=False).order_by('date_entree')[:50])
    urg = 0
    dep = 0
    lg_30 = 0
    lg_moins = 0
    total_cours = 0
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='icd', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        if mvt.date_sortie is None:
            total_cours += 1
            if mvt.date_entree:
                pass
            else:
                mvt.date_entree = maintenant
            if mvt.date_sortie:
                pass
            else:
                mvt.date_sortie = maintenant
            if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk:
                urg += 1
            elif (maintenant - mvt.date_entree) >= duree_dk:
                dep += 1
        else:
            if mvt.date_entree:
                pass
            else:
                mvt.date_entree = maintenant
            if mvt.date_sortie:
                pass
            else:
                mvt.date_sortie = maintenant
            duree_mouvement = mvt.date_sortie - mvt.date_entree
            if duree_mouvement >= duree_dk:
                lg_30 += 1
            else:
                lg_moins += 1
    totalter = lg_30 + lg_moins
    return JsonResponse({
        'urg': urg,
        'dep': dep,
        'lg_30': lg_30,
        'lg_mois': lg_moins,
        'total_cours': total_cours,
        'total_ter': totalter,
    })
####################### FETCH DASHBORD INFORMATIONS CMA ICD ###########################

def fetch_stats1(request, id_user):
    maintenant = timezone.now()

    # Filtrer les mouvements
    mouvements = (Mouvement4.objects.filter(date_sortie__isnull=True).order_by('date_entree')[:50]
                  | Mouvement4.objects.filter(date_sortie__isnull=False).order_by('date_entree')[:50])
    urg = 0
    dep = 0
    lg_30 = 0
    lg_moins = 0
    total_cours = 0
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='icd', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        if mvt.date_sortie is None:
            # Mouvements en cours
            total_cours += 1

            if mvt.date_entree:
                pass
            else:
                mvt.date_entree = maintenant
            if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk:
                urg += 1
            elif (maintenant - mvt.date_entree) >= duree_dk:
                dep += 1
        else:
            # Mouvements sortis
            duree_mouvement = mvt.date_sortie - mvt.date_entree
            if duree_mouvement >= duree_dk:
                lg_30 += 1
            else:
                lg_moins += 1

    totalter = lg_30 + lg_moins

    return JsonResponse({
        'urg': urg,
        'dep': dep,
        'lg_30': lg_30,
        'lg_mois': lg_moins,
        'total_cours': total_cours,
        'total_ter': totalter,
    })
####################### FETCH DASHBORD INFORMATIONS SACHERIE ###########################
def fetch_stats2(request, id_user):
    maintenant = timezone.now()

    # Filtrer les mouvements (Mouvement2, Mouvement6, Mouvement7)
    mouvements2 = Mouvement2.objects.filter(date_sortie__isnull=True).order_by('date_entree')[:50] | Mouvement2.objects.filter(date_sortie__isnull=False).order_by('date_entree')[:50]
    mouvements6 = Mouvement6.objects.filter(date_sortie__isnull=True).order_by('date_entree')[:50] | Mouvement6.objects.filter(date_sortie__isnull=False).order_by('date_entree')[:50]
    mouvements7 = Mouvement7.objects.filter(date_sortie__isnull=True).order_by('date_entree')[:50] | Mouvement7.objects.filter(date_sortie__isnull=False).order_by('date_entree')[:50]
    # Combiner les mouvements des trois modèles
    mouvements_combines = chain(mouvements2, mouvements6, mouvements7)
    urg = 0
    dep = 0
    lg_30 = 0
    lg_moins = 0
    total_cours = 0
    destination = "Non Assigné"
    for mvt in mouvements_combines:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        type = camion.type

        # Gestion des délais via ParametrageDelais
        try:
            para = ParametrageDelais.objects.get(entite='sacherie', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        # Gestion de la variable destination selon le modèle de mouvement
        if isinstance(mvt, Mouvement2):
            destination = "TOM"
        elif isinstance(mvt, Mouvement6):
            destination = "ITS"
        elif isinstance(mvt, Mouvement7):
            destination = "TRANSEXPRESS"
        else:
            destination = "Non Assigné"

        # Mouvements en cours
        if mvt.date_sortie is None:
            total_cours += 1

            if not mvt.date_entree:
                mvt.date_entree = maintenant

            if delais_urg <= (maintenant - mvt.date_entree) < duree_dk:
                urg += 1
            elif (maintenant - mvt.date_entree) >= duree_dk:
                dep += 1

        # Mouvements sortis
        else:
            duree_mouvement = mvt.date_sortie - mvt.date_entree
            if duree_mouvement >= duree_dk:
                lg_30 += 1
            else:
                lg_moins += 1

    totalter = lg_30 + lg_moins

    return JsonResponse({
        'urg': urg,
        'dep': dep,
        'lg_30': lg_30,
        'lg_mois': lg_moins,
        'total_cours': total_cours,
        'total_ter': totalter,
        'destination': destination,  # Inclure la variable destination
    })
####################### FETCH DASHBORD INFORMATIONS ZUD ###########################
def fetch_stats3(request, id_user):
    maintenant = timezone.now()
    # Filtrer les mouvements
    mouvements = (Mouvement3.objects.filter(date_sortie__isnull=True).order_by('date_entree')[:50]
                  | Mouvement3.objects.filter(date_sortie__isnull=False).order_by('date_entree')[:50])
    urg = 0
    dep = 0
    lg_30 = 0
    lg_moins = 0
    total_cours = 0
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        if mvt.date_sortie is None:
            total_cours += 1
            if mvt.date_entree:
                pass
            else:
                mvt.date_entree = maintenant
            if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk:
                urg += 1
            elif (maintenant - mvt.date_entree) >= duree_dk:
                dep += 1
        else:
            # Mouvements sortis
            duree_mouvement = mvt.date_sortie - mvt.date_entree
            if duree_mouvement >= duree_dk:
                lg_30 += 1
            else:
                lg_moins += 1
    totalter = lg_30 + lg_moins
    return JsonResponse({
        'urg': urg,
        'dep': dep,
        'lg_30': lg_30,
        'lg_mois': lg_moins,
        'total_cours': total_cours,
        'total_ter': totalter,
    })
#####Gestion Sortie
def liste_mouvements1(request, id_user):
    # Récupérer les mouvements avec des filtres sur date_entree et date_sortie
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Vérifier camion_id et assigner les valeurs du camion ou 'non assignée'
        camion_id = mouvement.get('camion_id')
        if camion_id:
            camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                    'type').first()
        else:
            camion = {'id_cam': 'non assignée', 'immatriculation': 'non assignée', 'transporteur': 'non assignée',
                      'type': 'non assignée'}
        mouvement['camion'] = camion

        # Vérifier chauffeur_id et assigner les valeurs du chauffeur ou 'non assignée'
        chauffeur_id = mouvement.get('chauffeur_id')
        if chauffeur_id:
            chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                                  'permis').first()
        else:
            chauffeur = {'id_chauffeur': 'non assignée', 'fullname': 'non assignée', 'permis': 'non assignée'}
        mouvement['chauffeur'] = chauffeur

        # Remplacer toutes les autres valeurs nulles ou vides par 'non assignée'
        for key, value in mouvement.items():
            if value is None or value == '':
                mouvement[key] = 'non assignée'

    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }

    return JsonResponse(response_data)


############################### FETCH SORTICE PARTICULIER ################
################### FETCH SORTIE ICD TOM ########################
def liste_mouvementsdkparticulier(request, id_user):
    mouvements = Mouvement8.objects.filter(date_sortie__isnull=True).values('id_mvt', 'vehicule_id', 'destination')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
      camion_id = mouvement['vehicule_id']
      camion = Vehicule.objects.filter(id_veh=camion_id).values('id_veh', 'immatriculation').first()
      mouvement['camion'] = camion
      response_data = {
      'mouvements': mouvement_list,
      'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)
################### FETCH SORTIE ICD TOM  DKLOG ########################
def liste_mouvementsdk1(request, id_user):
    mouvements = Mouvement0.objects.filter(date_sortie__isnull=True, destination__contains='icdtom').values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur', 'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

        # Assigner des valeurs par défaut pour les autres champs
        mouvement['statut_entree'] = mouvement.get('statut_entree', 'Non assigné')
        mouvement['statut_sortie'] = mouvement.get('statut_sortie', 'Non assigné')
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')

    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)

#### pip
################### FETCH SORTIE ICD CMA  DKLOG########################
def liste_mouvementsdk01(request, id_user):
    mouvements = Mouvement0.objects.filter(date_sortie__isnull=True, destination__contains='cmaicd').values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

        # Assigner des valeurs par défaut pour les autres champs
        mouvement['statut_entree'] = mouvement.get('statut_entree', 'Non assigné')
        mouvement['statut_sortie'] = mouvement.get('statut_sortie', 'Non assigné')
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')

    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)


################### FETCH SORTIE SACHERIE  DKLOG########################
def liste_mouvementsdk11(request, id_user):
    from django.http import JsonResponse

    # Récupération des mouvements
    mouvements = Mouvement0.objects.filter(date_sortie__isnull=True, destination__contains='hangar').values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'destination'
    )
    mouvement_list = list(mouvements)

    # Traitement des mouvements
    for mouvement in mouvement_list:
        # Gestion des informations du camion
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        if camion:
            mouvement['camion'] = camion
        else:
            mouvement['camion'] = {'id_cam': 'Non assigné', 'immatriculation': 'Non assigné',
                                   'transporteur': 'Non assigné', 'type': 'Non assigné'}

        # Gestion des informations du chauffeur
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        if chauffeur:
            mouvement['chauffeur'] = chauffeur
        else:
            mouvement['chauffeur'] = {'id_chauffeur': 'Non assigné', 'fullname': 'Non assigné', 'permis': 'Non assigné'}

    # Préparation des données pour la réponse JSON
    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }

    return JsonResponse(response_data)


################### FETCH SORTIE ZUD ########################
def liste_mouvementsdk12(request, id_user):
    mouvements = Mouvement0.objects.filter(date_sortie__isnull=True, destination__contains='zud').values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id'
                                            , 'remorque')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
      camion_id = mouvement['camion_id']
      camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur','type').first()
      mouvement['camion'] = camion
      chauffeur_id = mouvement['chauffeur_id']
      chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                 'permis').first()
      mouvement['chauffeur'] = chauffeur
    response_data = {
      'mouvements': mouvement_list,
      'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)
def ajoutsortie(request):
    if request.method == 'POST':
        form = SortieForm(request.POST)
        if form.is_valid():
            id_mvt = request.POST.get('id_mvt')
            id_user = request.POST.get('id_user')
            source =  request.POST.get('source')
            destination = request.POST.get('destination')
            # Récupérer les objets Utilisateurs et Mouvement en fonction des IDs fournis
            util = get_object_or_404(Utilisateurs, id_user=id_user)
            if (source == 'icdtom'):
                mouvement = get_object_or_404(Mouvement1, id_mvt=id_mvt)
            elif (source == 'icdcma'):
                mouvement = get_object_or_404(Mouvement4, id_mvt=id_mvt)
            elif (source == 'zud'):
                mouvement = get_object_or_404(Mouvement3, id_mvt=id_mvt)
            elif (source == 'sacherie'):
                if (destination== 'TOM'):
                    mouvement = get_object_or_404(Mouvement2, id_mvt=id_mvt)
                if (destination== 'ITS'):
                    mouvement = get_object_or_404(Mouvement6, id_mvt=id_mvt)
                elif (destination== 'TRANSEXPRESS'):
                    mouvement = get_object_or_404(Mouvement7, id_mvt=id_mvt)

            # Mettre à jour les informations du mouvement
            mouvement.date_sortie = timezone.now()
            mouvement.pointeur_sortie_id = util.id_user
            mouvement.statut_sortie = form.cleaned_data.get(
                'statut_sortie')  # Récupérer le statut depuis les données validées du formulaire
            mouvement.save()

            # Rediriger vers une vue après sauvegarde
            #return redirect(f'/index2_view/{util.id_user}')
            if (source=='icdtom'):
                return redirect(f"/sortie_tom/{id_user.strip()}?success=true")
            elif(source == 'zud'):
                return redirect(f"/index_sortie_zud/{id_user.strip()}?success=true")
            elif (source == 'icdcma'):
                return redirect(f"/index_sortie_icdcma/{id_user.strip()}?success=true")
            elif (source == 'sacherie'):
                return redirect(f"/index3_view/{id_user.strip()}?success=true")
    else:
        form = SortieForm()
    # Assurez-vous que 'mouvement' et 'util' sont définis si vous les utilisez dans le template
    return render(request, 'pages/ajoutsortiedp.html', {'form': form})

from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from .models import Utilisateurs, Mouvement0
from .forms import SortieForm
from django.http import HttpResponseRedirect
from urllib.parse import urlencode

def ajoutsortiepar(request):
    id_user = request.POST.get('id_user')
    util = Utilisateurs.objects.get(id_user=id_user)

    if request.method == 'POST':
        try:
            id_mvt = request.POST.get('id_mvt')
            zone_sortie = request.POST.get('zone')
            mouvement8 = Mouvement8.objects.get(id_mvt=id_mvt)
            mouvement8.date_sortie = timezone.now()
            mouvement8.zone_sortie = zone_sortie
            mouvement8.pointeur_sortie_id = util.id_user
            mouvement8.save()
            # Si tout se passe bien, on renvoie un paramètre de succès
            query_params = urlencode({'success': 'true'})
            return HttpResponseRedirect(f'/sortie_decalog_viewparticulier/{util.id_user}?{query_params}')
        except Mouvement8.DoesNotExist:
            # Gérer le cas où le mouvement n'existe pas
            query_params = urlencode({'error': 'Le mouvement spécifié n\'existe pas.'})
            return HttpResponseRedirect(f'/sortie_decalog_viewparticulier/{util.id_user}?{query_params}')
        except Exception as e:
            # Gérer toute autre erreur
            query_params = urlencode({'error': 'Une erreur s\'est produite: ' + str(e)})
            return HttpResponseRedirect(f'/sortie_decalog_viewparticulier/{util.id_user}?{query_params}')
    # Si la méthode n'est pas POST, on renvoie simplement la vue sans action
    return render(request, 'pages/ajoutsortiepar.html')
def ajoutsortiedk(request):
    id_user = request.POST.get('id_user')
    util = get_object_or_404(Utilisateurs, id_user=id_user)
    if request.method == 'POST':
        form = SortieForm(request.POST)
        if form.is_valid():
            zone_sortie = request.POST.get('zone')
            code_camion = request.POST.get('code_camion')
            poids = request.POST.get('poids') or 0
            poids_autorise = request.POST.get('poinds_autorise')
            pont_bascule = request.POST.get('pont_bascule')
            remorque =  request.POST.get('remorque')
            dest = request.POST.get('destination')
            pays =  request.POST.get('pays')
            if not pays == dest:
                destination = f"{pays} ({dest})"
            else:
                destination =  pays
            id_mvt = request.POST.get('id_mvt')
            source = request.POST.get('source')
            num_ticket = request.POST.get('num_ticket')
            if source == 'zud':
                mouvement0 = Mouvement0.objects.get(id_mvt=id_mvt)
                mouvement0.date_sortie = timezone.now()
                mouvement0.pointeur_sortie_id = util.id_user
                mouvement0.zone_sortie = zone_sortie
                mouvement0.num_ticket = num_ticket
                mouvement0.save()
                mouvement = Mouvement3.objects.get(id_mvt_0_id=id_mvt)
                mouvement.num_ticket = num_ticket
                mouvement.pointeur_sortie_id = util.id_user
                mouvement.zone_sortie = zone_sortie
                if not mouvement.date_entree:
                    mouvement.date_entree = timezone.now()
                if not mouvement.date_sortie:
                    mouvement.date_sortie = timezone.now()
                mouvement.save()
            elif source == 'sacherie':
                mouvement0 = Mouvement0.objects.get(id_mvt=id_mvt)
                mouvement0.poids = poids
                mouvement0.poids_autorise = poids_autorise
                mouvement0.pont_bascule = pont_bascule
                mouvement0.remorque = remorque
                mouvement0.code_camion = code_camion
                mouvement0.date_sortie = timezone.now()
                mouvement0.destination_final = destination
                mouvement0.save()
                if 'tom' in mouvement0.destination:
                    mouvement = Mouvement2.objects.filter(id_mvt_0_id=id_mvt).first()
                    mouvement.poids = poids
                    mouvement.poids_autorise = poids_autorise
                    mouvement.pont_bascule = pont_bascule
                    mouvement.remorque = remorque
                    mouvement.code_camion = code_camion
                    mouvement.zone_sortie = zone_sortie
                    mouvement.destination_final = destination
                    if not mouvement.date_entree:
                        mouvement.date_entree = timezone.now()
                    if not mouvement.date_sortie:
                        mouvement.date_sortie = timezone.now()
                    mouvement.save()
                elif 'transexpress' in mouvement0.destination:
                    mouvement = Mouvement7.objects.filter(id_mvt_0_id=id_mvt).first()
                    mouvement.poids = poids
                    mouvement.poids_autorise = poids_autorise
                    mouvement.pont_bascule = pont_bascule
                    mouvement.zone_sortie = zone_sortie
                    mouvement.remorque = remorque
                    mouvement.code_camion = code_camion
                    mouvement.destination_final = destination
                    if not mouvement.date_entree:
                        mouvement.date_entree = timezone.now()
                    if not mouvement.date_sortie:
                        mouvement.date_sortie = timezone.now()
                    mouvement.save()
                elif 'its' in mouvement0.destination:
                    mouvement = Mouvement6.objects.filter(id_mvt_0_id=id_mvt).first()
                    mouvement.poids = poids
                    mouvement.poids_autorise = poids_autorise
                    mouvement.pont_bascule = pont_bascule
                    mouvement.remorque = remorque
                    mouvement.code_camion = code_camion
                    mouvement.destination_final = destination
                    mouvement.zone_sortie = zone_sortie
                    if not mouvement.date_entree:
                        mouvement.date_entree = timezone.now()
                    if not mouvement.date_sortie:
                        mouvement.date_sortie = timezone.now()
                    mouvement.save()
                pass  # Ajouter votre code ici
            elif 'icd' in source:
                # Comportement spécifique pour 'icd'
                pass  # Ajouter votre code ici

            # Récupérer les objets Utilisateurs et Mouvement en fonction des IDs fournis
                movement_map = {
                    'icdtom': Mouvement1,
                    'icdcma': Mouvement4,
                }
                mouvement = get_object_or_404(movement_map.get(source), id_mvt_0_id=id_mvt)
            # Mettre à jour les informations du mouvement
                mouvement.date_sortie = timezone.now()
                mouvement.pointeur_sortie_id = util.id_user
                mouvement.zone_sortie = zone_sortie
                if not mouvement.date_entree:
                    mouvement.date_entree = timezone.now()
                if not mouvement.date_sortie:
                    mouvement.date_sortie = timezone.now()
                # Logique en fonction de la source
                mouvement.save()
                mouvement0=Mouvement0.objects.get(id_mvt=mouvement.id_mvt_0_id)
                mouvement0.date_sortie = timezone.now()
                mouvement0.pointeur_sortie_id = util.id_user
                mouvement0.zone_sortie = zone_sortie
                mouvement0.save()
            # Rediriger vers une vue après sauvegarde avec un paramètre de succès
            query_params = urlencode({'success': 'true'})
            if source == 'sacherie':
                return HttpResponseRedirect(f'/sortie_decalog_view1/{util.id_user}?{query_params}')
            elif source == 'zud':
                return HttpResponseRedirect(f'/sortie_decalog_view2/{util.id_user}?{query_params}')
            else:
                return HttpResponseRedirect(f'/sortie_decalog_view/{util.id_user}?{query_params}')
        else:
            # Rediriger avec un message d'erreur si le formulaire est invalide
            query_params = urlencode({'error': 'Le formulaire est invalide.'})
        return HttpResponseRedirect(f'/sortie_decalog_view/{util.id_user}?{query_params}')
    else:
        form = SortieForm()
    return render(request, 'pages/ajoutsortiedk.html', {'form': form})
def tout_mouvement(request, id_mvt):
    user=Utilisateurs.objects.get(id_user=id_mvt)
    return render(request, 'pages/liste_mouvement.html', {'util': user})
########################## TOUT MOUVEMENT ADMIN DKLOG PLT #####################
def tout_mouvement00(request, id_mvt):
    user=Utilisateurs.objects.get(id_user=id_mvt)
    mouvements = Mouvement0.objects.filter(date_sortie__isnull=True).values(
        'id_mvt', 'destination', 'camion_id', 'statut_entree', 'statut_sortie',
        'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
        'pointeur_sortie_id', 'pointeur_entree_id'
    )
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        # Camion details
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion or {'id_cam': 'Non Assigné', 'immatriculation': 'Non Assigné',
                                         'transporteur': 'Non Assigné'}
        # Pointeur entrée
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre or {'fullname': 'Non Assigné'}
        # Pointeur sortie
        user_sortie = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_sortie_id')).values('fullname').first()
        mouvement['user_srt'] = user_sortie or {'fullname': 'Non Assigné'}
        # Chauffeur details
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur or {'id_chauffeur': 'Non Assigné', 'fullname': 'Non Assigné',
                                               'permis': 'Non Assigné'}
    return render(request, 'pages/liste_mouvements00.html', {'util': user, 'mouvement_list': mouvement_list})########################## TOUT MOUVEMENT ADMIN DKLOG TOM ICD #####################
########################## TOUT MOUVEMENT ADMIN DKLOG TOM ICD #####################
def tout_mouvement0(request, id_mvt):
    user = Utilisateurs.objects.get(id_user=id_mvt)
    # movements query update (only if destination exists)
    mouvements = Mouvement1.objects.filter(date_sortie__isnull=True).values(
        'id_mvt',  'camion_id', 'statut_entree', 'statut_sortie',
        'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
        'pointeur_sortie_id', 'pointeur_entree_id'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Camion details
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion or {'id_cam': 'Non Assigné', 'immatriculation': 'Non Assigné',
                                         'transporteur': 'Non Assigné'}

        # Pointeur entrée
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre or {'fullname': 'Non Assigné'}

        # Pointeur sortie
        if mouvement.get('pointeur_sortie_id'):
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            user_sortie = {'fullname': 'Non Assigné'}

        mouvement['user_srt'] = user_sortie

        # Chauffeur details
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()

        mouvement['chauffeur'] = chauffeur or {'id_chauffeur': 'Non Assigné', 'fullname': 'Non Assigné',
                                               'permis': 'Non Assigné'}

        # Remorque
        mouvement['remorque'] = mouvement.get('remorque') or 'Non Assigné'

        # Date d'entrée et sortie
        mouvement['date_entree'] = mouvement.get('date_entree') or 'Non Assigné'
        mouvement['date_sortie'] = mouvement.get('date_sortie') or 'Non Assigné'
    return render(request, 'pages/liste_mouvements0.html', {'util': user, 'mouvement_list': mouvement_list})
########################## TOUT MOUVEMENT ADMIN DKLOG TOM CMA #####################
def tout_mouvement01(request, id_mvt):
    user = Utilisateurs.objects.get(id_user=id_mvt)

    # Query Mouvement4 and get necessary fields as dictionaries
    mouvements = Mouvement4.objects.filter(date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
        'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
        'pointeur_sortie_id', 'pointeur_entree_id'
    )

    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Camion management
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion or {
            'id_cam': 'Non Assigné',
            'immatriculation': 'Non Assigné',
            'transporteur': 'Non Assigné'
        }

        # Pointeur entrée management
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre or {'fullname': 'Non Assigné'}

        # Pointeur sortie management
        user_sortie = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_sortie_id')).values(
            'fullname').first() if mouvement.get('pointeur_sortie_id') else {'fullname': 'Non Assigné'}
        mouvement['user_srt'] = user_sortie

        # Chauffeur management (correcting the model name from 'Chaffeur' to 'Chauffeur')
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                               'permis').first()
        mouvement['chauffeur'] = chauffeur or {
            'id_chauffeur': 'Non Assigné',
            'fullname': 'Non Assigné',
            'permis': 'Non Assigné'
        }

        # Remorque management
        mouvement['remorque'] = mouvement.get('remorque') or 'Non Assigné'

        # Dates management
        mouvement['date_entree'] = mouvement.get('date_entree') or 'Non Assigné'
        mouvement['date_sortie'] = mouvement.get('date_sortie') or 'Non Assigné'

    return render(request, 'pages/liste_mouvements01.html', {'util': user, 'mouvements': mouvement_list})
######################### TOUT MOUVEMENT ADMIN DKLOG SACHERIE #####################
def tout_mouvement02(request, id_mvt):
    user = Utilisateurs.objects.get(id_user=id_mvt)

    # Fetch data from Mouvement2, Mouvement6, and Mouvement7
    mouvements = Mouvement2.objects.filter(date_sortie__isnull=True).values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                 'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                 'pointeur_sortie_id', 'pointeur_entree_id')

    mouvements6 = Mouvement6.objects.all().values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                  'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                  'pointeur_sortie_id', 'pointeur_entree_id')

    mouvements7 = Mouvement7.objects.all().values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                  'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                  'pointeur_sortie_id', 'pointeur_entree_id')

    # Combine the movements from different models
    mouvement_list = list(mouvements) + list(mouvements6) + list(mouvements7)

    for mouvement in mouvement_list:
        # Camion management
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion or {'id_cam': 'Non Assigné', 'immatriculation': 'Non Assigné',
                                         'transporteur': 'Non Assigné'}

        # Pointeur entrée management
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre or {'fullname': 'Non Assigné'}

        # Pointeur sortie management
        user_sortie = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_sortie_id')).values(
            'fullname').first()
        mouvement['user_srt'] = user_sortie or {'fullname': 'Non Assigné'}

        # Chauffeur management (corrected model name to 'Chauffeur')
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                               'permis').first()
        mouvement['chauffeur'] = chauffeur or {'id_chauffeur': 'Non Assigné', 'fullname': 'Non Assigné',
                                               'permis': 'Non Assigné'}

        # Add the destination
        if mouvement in mouvements:
            mouvement['destination'] = "TOM"
        elif mouvement in mouvements6:
            mouvement['destination'] = "ITS"
        elif mouvement in mouvements7:
            mouvement['destination'] = "TRANSEXPRESS"
        else:
            mouvement['destination'] = "Non Assigné"

    # Correct the context when rendering
    return render(request, 'pages/liste_mouvements02.html', {'util': user, 'mouvements': mouvement_list})
########################## TOUT MOUVEMENT ADMIN DKLOG ZUD #####################
def tout_mouvement03(request, id_mvt):
    user=Utilisateurs.objects.get(id_user=id_mvt)
    mouvements = Mouvement3.objects.filter(date_sortie__isnull=True).values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                 'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                 'pointeur_sortie_id', 'pointeur_entree_id')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion
        user_entre = Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        if mouvement['pointeur_sortie_id']:
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_sortie_id', 1)).values('fullname').first()
            # for user_sortie in user_sortie :

            user_sortie = {'fullname': 'original_name'}
            user_sortie['fullname'] = 'null'
            # user_sortie='oo'
        mouvement['user_ert'] = user_entre
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    return render(request, 'pages/liste_mouvements03.html', {'util': user, 'mouvements': mouvement_list})
########################## TOUT MOUVEMENT ADMIN DKLOG PARTICULIERS #####################
def tout_mouvementpar(request, id_mvt):
    user = Utilisateurs.objects.get(id_user=id_mvt)
    mouvements = Mouvement8.objects.filter(date_sortie__isnull=True).values('id_mvt', 'vehicule_id', 'destination', 'date_entree', 'date_sortie',
                                                 'pointeur_sortie_id', 'pointeur_entree_id')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        # Gestion du vehicule
        vehicule_id = mouvement.get('vehicule_id')
        vehicule = Vehicule.objects.filter(id_veh=vehicule_id).values('id_veh', 'immatriculation').first()
        mouvement['vehicule'] = vehicule if vehicule else {'immatriculation': 'non assigné'}
        # Gestion du pointeur d'entrée
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'non assigné'}

        # Gestion du pointeur de sortie
        if mouvement.get('pointeur_sortie_id'):
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            user_sortie = {'fullname': 'Non Assigné'}
        mouvement['user_srt'] = user_sortie if user_sortie else {'fullname': 'non assigné'}

    return render(request, 'pages/liste_mouvementspar.html', {'util': user, 'mouvements': mouvement_list})


def tout_mouvement2(request, id_mvt):
    user=Utilisateurs.objects.get(id_user=id_mvt)
    return render(request, 'pages/liste_mouvement2.html', {'util': user})


def liste_mouvements2(request):
    mouvements = Mouvement1.objects.all().values('id_mvt','mission','camion_id', 'statut_entree', 'statut_sortie','chauffeur_id', 'remorque','date_entree','date_sortie','pointeur_sortie_id','pointeur_entree_id')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion
        user_entre=Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        if mouvement['pointeur_sortie_id']  :
            user_sortie= Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values('fullname').first()
        else:
            user_sortie = Utilisateurs.objects.filter(id_user=1).values('fullname').first()
            #for user_sortie in user_sortie :
                # Mettre l'attribut fullname vide
            user_sortie = {'fullname': 'original_name'}
            user_sortie['fullname'] = 'null'

            #user_sortie='oo'

        mouvement['user_ert'] = user_entre
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    return JsonResponse(list(mouvement_list), safe=False)
################################# FETCH TOUTU MOUVEMENTS ADMIN ICD TOM #########################
def liste_mouvements_dk0(request):
    mouvements = Mouvement0.objects.all().values(
        'id_mvt', 'destination', 'camion_id', 'statut_entree', 'statut_sortie',
        'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
        'pointeur_sortie_id', 'pointeur_entree_id'
    )

    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Camion details
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion or {'id_cam': 'Non Assigné', 'immatriculation': 'Non Assigné',
                                         'transporteur': 'Non Assigné'}

        # Pointeur entrée
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre or {'fullname': 'Non Assigné'}

        # Pointeur sortie
        if mouvement.get('pointeur_sortie_id'):
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            user_sortie = {'fullname': 'Non Assigné'}

        mouvement['user_srt'] = user_sortie

        # Chauffeur details
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur or {'id_chauffeur': 'Non Assigné', 'fullname': 'Non Assigné',
                                               'permis': 'Non Assigné'}

    return JsonResponse(mouvement_list, safe=False)


################################# FETCH TOUTU MOUVEMENTS ADMIN ICD TOM #########################
def liste_mouvements_dk(request):
    mouvements = Mouvement1.objects.all().values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
        'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
        'pointeur_sortie_id', 'pointeur_entree_id'
    )

    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Camion details
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion or {'id_cam': 'Non Assigné', 'immatriculation': 'Non Assigné',
                                         'transporteur': 'Non Assigné'}

        # Pointeur entrée
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre or {'fullname': 'Non Assigné'}

        # Pointeur sortie
        if mouvement.get('pointeur_sortie_id'):
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            user_sortie = {'fullname': 'Non Assigné'}

        mouvement['user_srt'] = user_sortie

        # Chauffeur details
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur or {'id_chauffeur': 'Non Assigné', 'fullname': 'Non Assigné',
                                               'permis': 'Non Assigné'}

        # Remorque
        mouvement['remorque'] = mouvement.get('remorque') or 'Non Assigné'

        # Date d'entrée et sortie
        mouvement['date_entree'] = mouvement.get('date_entree') or 'Non Assigné'
        mouvement['date_sortie'] = mouvement.get('date_sortie') or 'Non Assigné'

    return JsonResponse(mouvement_list, safe=False)


################################# FETCH TOUTU MOUVEMENTS ADMIN ICD CMA #########################

def liste_mouvements_dk1(request):
    mouvements = Mouvement4.objects.all().values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
        'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
        'pointeur_sortie_id', 'pointeur_entree_id'
    )

    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Gestion du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion or {'id_cam': 'Non Assigné', 'immatriculation': 'Non Assigné',
                                         'transporteur': 'Non Assigné'}

        # Gestion du pointeur d'entrée
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre or {'fullname': 'Non Assigné'}

        # Gestion du pointeur de sortie
        if mouvement.get('pointeur_sortie_id'):
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            user_sortie = {'fullname': 'Non Assigné'}

        mouvement['user_srt'] = user_sortie

        # Gestion du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur or {'id_chauffeur': 'Non Assigné', 'fullname': 'Non Assigné',
                                               'permis': 'Non Assigné'}

        # Gestion de la remorque
        mouvement['remorque'] = mouvement.get('remorque') or 'Non Assigné'

        # Gestion des dates d'entrée et de sortie
        mouvement['date_entree'] = mouvement.get('date_entree') or 'Non Assigné'
        mouvement['date_sortie'] = mouvement.get('date_sortie') or 'Non Assigné'

    return JsonResponse(mouvement_list, safe=False)


################################# FETCH TOUTU MOUVEMENTS ADMIN SACHERIE #########################

from django.db.models import Q

def liste_mouvements_dk2(request):
    # Combiner les valeurs de Mouvement2, Mouvement6, et Mouvement7
    mouvements = Mouvement2.objects.all().values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                 'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                 'pointeur_sortie_id', 'pointeur_entree_id')

    mouvements6 = Mouvement6.objects.all().values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                  'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                  'pointeur_sortie_id', 'pointeur_entree_id')

    mouvements7 = Mouvement7.objects.all().values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                  'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                  'pointeur_sortie_id', 'pointeur_entree_id')

    # Fusionner les mouvements
    mouvement_list = list(mouvements) + list(mouvements6) + list(mouvements7)

    # Ajouter la variable destination et remplir les champs
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion if camion else "Non Assigné"

        # Récupérer les informations du pointeur d'entrée et de sortie
        user_entre = Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        user_sortie = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_sortie_id', 1)).values('fullname').first() or {'fullname': 'Non Assigné'}

        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non Assigné'}
        mouvement['user_srt'] = user_sortie

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname', 'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else "Non Assigné"

        # Ajouter la variable destination
        if mouvement in mouvements:
            mouvement['destination'] = "TOM"
        elif mouvement in mouvements6:
            mouvement['destination'] = "ITS"
        elif mouvement in mouvements7:
            mouvement['destination'] = "TRANSEXPRESS"
        else:
            mouvement['destination'] = "Non Assigné"

    return JsonResponse(list(mouvement_list), safe=False)

################################# FETCH TOUTU MOUVEMENTS ADMIN ZUD #########################

def liste_mouvements_dk3(request):
    mouvements = Mouvement3.objects.all().values('id_mvt', 'camion_id', 'statut_entree', 'statut_sortie',
                                                 'chauffeur_id', 'remorque', 'date_entree', 'date_sortie',
                                                 'pointeur_sortie_id', 'pointeur_entree_id')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion
        user_entre = Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        if mouvement['pointeur_sortie_id']:
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            user_sortie = Utilisateurs.objects.filter(id_user=1).values('fullname').first()
            # for user_sortie in user_sortie :
            # Mettre l'attribut fullname vide
            user_sortie = {'fullname': 'original_name'}
            user_sortie['fullname'] = 'null'
            # user_sortie='oo'
        mouvement['user_ert'] = user_entre
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    return JsonResponse(list(mouvement_list), safe=False)
####################### FETCH TOUT MOUVEMENTS PARTICULIERS ##############################
################################# FETCH TOUTU MOUVEMENTS ADMIN ZUD #########################

def liste_mouvements_dkpar(request):
    mouvements = Mouvement8.objects.all().values('id_mvt', 'vehicule_id',
                                                 'destination', 'date_entree', 'date_sortie',
                                                 'pointeur_sortie_id', 'pointeur_entree_id')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        # Gestion du vehicule
        vehicule_id = mouvement.get('vehicule_id')
        vehicule = Vehicule.objects.filter(id_veh=vehicule_id).values('id_veh', 'immatriculation').first()
        mouvement['vehicule'] = vehicule if vehicule else {'immatriculation': 'non assigné'}
        # Gestion du pointeur d'entrée
        user_entre = Utilisateurs.objects.filter(id_user=mouvement.get('pointeur_entree_id')).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'non assigné'}
        # Gestion du pointeur de sortie
        if mouvement.get('pointeur_sortie_id'):
            user_sortie = Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values(
                'fullname').first()
        else:
            # Valeur par défaut si pointeur_sortie_id est null
            user_sortie = {'fullname': 'non assigné'}
        mouvement['user_srt'] = user_sortie if user_sortie else {'fullname': 'non assigné'}
    return JsonResponse(list(mouvement_list), safe=False)
####################### FETCH TOUT MOUVEMENTS SACHERIE ##############################
def liste_mouvements_sacherie(request):
    mouvements = Mouvement2.objects.all().values('id_mvt','mission','camion_id', 'statut_entree', 'statut_sortie','chauffeur_id', 'remorque','date_entree','date_sortie','pointeur_sortie_id','pointeur_entree_id')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion
        user_entre=Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        if mouvement['pointeur_sortie_id']  :
            user_sortie= Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values('fullname').first()
        else:
            user_sortie = Utilisateurs.objects.filter(id_user=1).values('fullname').first()
            #for user_sortie in user_sortie :
                # Mettre l'attribut fullname vide
            user_sortie = {'fullname': 'original_name'}
            user_sortie['fullname'] = 'null'

            #user_sortie='oo'

        mouvement['user_ert'] = user_entre
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    return JsonResponse(list(mouvement_list), safe=False)
################# MODIFICATION MOUVEMENTS PARTICULIER ~########################
def modifier_mouvementpar(request):
    if request.method == 'POST':
        try:
            id_mvt = request.POST.get('id_mouvement')
            id_user = request.POST.get('id_user')
            destination = request.POST.get('destination')
            vehicule_id = request.POST.get('vehicule_id')
            mouvement = Mouvement8.objects.filter(id_mvt=id_mvt).first()
            mouvement.destination = destination
            mouvement.vehicule_id = vehicule_id
            mouvement.save()
            return redirect(f"/modif_mvtpar/{id_user}?success=true")
        except Exception as e:
            error_message = str(e)
            return redirect(f"/modif_mvtpar/{id_user}?error={error_message}")
##################### MODIFIATION MOUVEMENTS DKLOG #############################

@csrf_exempt
def modifier_mouvement(request):
    if request.method == 'POST':
        try:
            ################# RECUPERATION DES VALEURS ###############
            id_mvt = request.POST.get('id_mouvement')
            id_user = request.POST.get('id_user')
            navire = request.POST.get('navire')
            camion_id = request.POST.get('camion')
            chauffeur_id = request.POST.get('chauffeur')
            client_id = request.POST.get('client')
            representant_id = request.POST.get('representant')
            marchandise = request.POST.get('marchandise')
            mission = request.POST.get('mission')
            date_validite = request.POST.get('date_validite')
            tonnage = request.POST.get('tonnage') or 0
            typeconteneur1 = request.POST.get('typeconteneur1')
            numconteneur1 = request.POST.get('numconteneur1')
            typeconteneur2 = request.POST.get('typeconteneur2')
            numconteneur2 = request.POST.get('numconteneur2')
            bl1 = request.POST.get('bl1')
            bl2 = request.POST.get('bl2')
            nbrcolis = request.POST.get('nbrcolis') or 0
            entite = request.POST.get('entite')
            destination = request.POST.get('destination')
            remorque = request.POST.get('remorque')
            transitaire_id = request.POST.get('transitaire')
            mouvement0= Mouvement0.objects.filter(id_mvt=id_mvt).first()
            # Mise à jour de Mouvement0
            if typeconteneur1:
                mouvement0.typeconteneur1 = typeconteneur1
            if remorque:
                mouvement0.remorque = remorque
            if navire:
                mouvement0.navire = navire
            if numconteneur1:
                mouvement0.numconteneur1 = numconteneur1
            if typeconteneur2:
                mouvement0.typeconteneur2 = typeconteneur2
            if numconteneur2:
                mouvement0.numconteneur2 = numconteneur2
            if bl1:
                mouvement0.bl1 = bl1
            if date_validite:
                mouvement0.date_validite = date_validite
            if bl2:
                mouvement0.bl2 = bl2
            if nbrcolis:
                mouvement0.nbrcolis = nbrcolis
            if tonnage:
                mouvement0.tonnage = tonnage
            if mission:
                mouvement0.mission = mission
            if camion_id:
                mouvement0.camion_id = camion_id
            if transitaire_id:
                mouvement0.transitaire_id = transitaire_id
            if representant_id:
                mouvement0.representant_id = representant_id
            if chauffeur_id:
                mouvement0.chauffeur_id = chauffeur_id
            if client_id:
                mouvement0.client_id = client_id
            if marchandise:
                mouvement0.marchandise = marchandise
            if entite != destination:
                destination = f"{entite} ({destination})"
            else:
                destination = entite
            mouvement0.destination = destination
            #camion = Camion.objects.get(id_cam=camion_id)
            mouvement0.save()
            # Mise à jour ou création des mouvements enfants en fonction du destination
            movement_type = None
            destination = destination
            if 'icdtom' in destination:
                movement_type = Mouvement1
            elif 'cmaicd' in destination:
                movement_type = Mouvement4
            elif 'tom' in destination and 'hangar' in destination:
                movement_type = Mouvement2
            elif 'its' in destination and 'hangar' in destination:
                movement_type = Mouvement6
            elif 'transexpress' in destination and 'hangar' in destination:
                movement_type = Mouvement7
            elif 'zud' in destination:
                movement_type = Mouvement3

            if movement_type:
                # Vérification si le mouvement enfant existe déjà
                # Vérification si le mouvement enfant existe déjà
                mouvement = movement_type.objects.filter(id_mvt_0=mouvement0).first()

                if mouvement is not None:
                    # Mise à jour de l'enfant
                    if camion_id:
                        mouvement.camion_id = camion_id
                    if chauffeur_id:
                        mouvement.chauffeur_id = chauffeur_id
                    if client_id:
                        mouvement.client_id = client_id
                    if mission:
                        mouvement.mission = mission
                    if remorque:
                        mouvement.remorque = remorque
                else:
                    # Si le mouvement n'existe pas, il faut le créer
                    mouvement = movement_type(
                        camion_id=camion_id,
                        chauffeur_id=chauffeur_id,
                        client_id=client_id,
                        mission=mission,
                        remorque=remorque,
                        id_mvt_0=mouvement0
                    )

                # Mise à jour des champs spécifiques à Zud
                if destination == 'zud':
                    if marchandise:
                        mouvement.marchandise = marchandise
                    if bl1:
                        mouvement.bl1 = bl1
                    if bl2:
                        mouvement.bl2 = bl2
                    if transitaire_id:
                        mouvement.transitaire_id = transitaire_id
                    if representant_id:
                        mouvement.representant_id = representant_id
                    if navire:
                        mouvement.navire = navire
                    if tonnage:
                        mouvement.tonnage = tonnage
                    if nbrcolis:
                        mouvement.nbrcolis = nbrcolis

                # Sauvegarder l'objet mouvement enfant
                mouvement.save()
            if 'hangar' in destination:
                return redirect(f"/modif_mvt1/{id_user}?success=true")
            elif destination == 'zud':
                return redirect(f"/modif_mvt2/{id_user}?success=true")
            else:
                return redirect(f"/modif_mvt/{id_user}?success=true")
        except Exception as e:
            error_message = str(e)
            return redirect(f"/modif_mvt/{id_user}?error={error_message}")

    else:
        chauffeurs = Chaffeur.objects.all()
        camions = Camion.objects.all()
        return render(request, 'pages/mouvement_entre_0.html', {
            'camions': camions,
            'chauffeurs': chauffeurs,
        })
@csrf_exempt
def modifier_mouvement0(request):
    if request.method == 'POST':
      boite_id = request.POST['mouvementId']
      boite = get_object_or_404(Mouvement0, id_mvt=boite_id)
      boite.destination = request.POST['destination']
      boite.remorque = request.POST['remorque']
      boite.statut_entree = request.POST['statut_entree']
      boite.statut_sortie = request.POST['statut_sortie']

      if  request.POST['camion'] :
          try :
              cam_mof=Camion.objects.filter(immatriculation= request.POST['camion']).first()
              boite.camion_id=cam_mof.id_cam
          except :
           pass
      if  request.POST['permis'] :
          try :
              chau_mof=Chaffeur.objects.filter(permis= request.POST['permis']).first()
              boite.chauffeur_id=chau_mof.id_chauffeur
          except :
           pass



      boite.save()
      #return redirect(reverse('index'))  # Redirigez vers la page index après la modification
      return redirect("/liste_tt_mvt0/" + str(request.POST['id_user']))

    return render(request, 'index.html')

@csrf_exempt
def modifier_mouvement2(request):
    if request.method == 'POST':
      boite_id = request.POST['mouvementId']

      boite = get_object_or_404(Mouvement2, id_mvt=boite_id)
      status_sortie = boite.statut_sortie
      status_entree = boite.statut_entree
      #boite.immatriculation = request.POST['camion']
      #boite.transporteur = request.POST['trans']
      boite.mission = request.POST['mission']
      #boite.permis = request.POST['permis']
      boite.remorque = request.POST['remorque']
      boite.statut_entree= request.POST['statut_entree']
      boite.statut_sortie= request.POST['statut_sortie']

      if  request.POST['camion'] :
          try :
              cam_mof=Camion.objects.filter(immatriculation= request.POST['camion']).first()
              boite.camion_id=cam_mof.id_cam
          except :
           pass
      if  request.POST['permis'] :
          try :
              chau_mof=Chaffeur.objects.filter(permis= request.POST['permis']).first()
              boite.chauffeur_id=chau_mof.id_chauffeur
          except :
           pass



      boite.save()
      #return redirect(reverse('index'))  # Redirigez vers la page index après la modification
      return redirect("/liste_tt_mvt2/" + str(request.POST['id_user']))

    return render(request, 'index.html')


def activer_desactiver_user(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        status = request.POST.get('status')

        user = Utilisateurs.objects.get(id_user=user_id)  # Remplacez User par votre modèle utilisateur
        user.status = status
        user.save()

        return redirect(
        "/liste_user/" + str(request.POST.get('userid')))  # Remplacez 'your_view_name' par le nom de la vue que vous souhaitez rediriger après la mise à jour

##################3Gestion
def liste_mouvements_0(request):
    mouvements = Mouvement0.objects.filter(destination__icontains='icd', date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupérer les informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}

        # Récupérer les informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        if user_sortie_id:
            user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first()
        else:
            user_sortie = {'fullname': 'null'}
        mouvement['user_srt'] = user_sortie

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

        # Assigner des valeurs par défaut pour les autres champs
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')
        mouvement['date_entree'] = mouvement.get('date_entree', 'Non assigné')
        mouvement['date_sortie'] = mouvement.get('date_sortie', 'Non assigné')
        mouvement['destination'] = mouvement.get('destination', 'Non assigné')

    return JsonResponse(mouvement_list, safe=False)


def liste_mouvements_01(request):
    mouvements = Mouvement0.objects.filter(destination__icontains='hangar', date_sortie__isnull=True).values('id_mvt','camion_id', 'statut_entree', 'statut_sortie','chauffeur_id', 'remorque','date_entree','date_sortie','pointeur_sortie_id','pointeur_entree_id','destination')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur','type').first()
        mouvement['camion'] = camion
        user_entre=Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        if mouvement['pointeur_sortie_id']  :
            user_sortie= Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values('fullname').first()
        else:
            user_sortie = Utilisateurs.objects.filter(id_user=1).values('fullname').first()
            #for user_sortie in user_sortie :
                # Mettre l'attribut fullname vide
            user_sortie = {'fullname': 'original_name'}
            user_sortie['fullname'] = 'null'

            #user_sortie='oo'

        mouvement['user_ert'] = user_entre
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    return JsonResponse(list(mouvement_list), safe=False)
####################### FETCHE POUR LA SACHERIE #################

def liste_mouvements_02(request):
    mouvements = Mouvement0.objects.filter(destination__icontains='zud', date_sortie__isnull=True).values('id_mvt','camion_id', 'statut_entree', 'statut_sortie','chauffeur_id', 'remorque','date_entree','date_sortie','pointeur_sortie_id','pointeur_entree_id','destination')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur','type').first()
        mouvement['camion'] = camion
        user_entre=Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        if mouvement['pointeur_sortie_id']  :
            user_sortie= Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values('fullname').first()
        else:
            user_sortie = Utilisateurs.objects.filter(id_user=1).values('fullname').first()
            #for user_sortie in user_sortie :
                # Mettre l'attribut fullname vide
            user_sortie = {'fullname': 'original_name'}
            user_sortie['fullname'] = 'null'

            #user_sortie='oo'

        mouvement['user_ert'] = user_entre
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    return JsonResponse(list(mouvement_list), safe=False)
def ajouter_mouvements_0(request):
    ###ajout mvt 0
    ####copie mvt 1
    return  redirect("/entreedecalog_view/" + str(request.POST['id_user']))
from django.shortcuts import redirect, get_object_or_404
from django.utils import timezone
from .models import *
from django.http import HttpResponse
###################### AJOUT LIAISON MOUVEMENT ZUD #########################
def liaisonmouvement0(request, id_user):
    util = get_object_or_404(Utilisateurs, id_user=id_user)
    if request.method == 'POST':
        try:
            zone = request.POST.get('zone')
            date_validite = request.POST.get('date_validite', '').strip()
            conteneur1 = request.POST.get('numcont1', '').strip()
            type1 = request.POST.get('typecont1', '').strip()
            conteneur2 = request.POST.get('numcont2', '').strip()
            type2 = request.POST.get('typecont2', '').strip()
            conteneur3 = request.POST.get('numcont3', '').strip()
            conteneur4 = request.POST.get('numcont4', '').strip()
            conteneur5 = request.POST.get('numcont5', '').strip()
            remorque = request.POST.get('remorque', '').strip()
            type3 = request.POST.get('typecont3', '').strip()
            type4 = request.POST.get('typecont4', '').strip()
            type5 = request.POST.get('typecont5', '').strip()
            client = request.POST.get('client', '').strip()
            chauffeur = request.POST.get('chauffeur', '').strip()
            max_size = 0
            transitaire = request.POST.get('transitaire', '').strip()
            representant = request.POST.get('representant', '').strip()
            id_mvt1 = request.POST.get('id_mvt', '').strip()  # This is an ID, not an instance
            typecam = request.POST.get('typecam', '').strip()
            camion = request.POST.get('camion', '').strip()
            mouvement_instance = Mouvement0.objects.get(id_mvt=id_mvt1)
            mouvement0 = Mouvement0.objects.create()
            mouvement = Mouvement3.objects.create()
            if conteneur1:
                mouvement0.numconteneur1 = conteneur1
                mouvement.numconteneur1 = conteneur1
                mouvement0.typeconteneur1 = type1
                mouvement.typeconteneur1 = type1
                max_size+= int(type1) or 0
            if conteneur2:
                mouvement0.numconteneur2 = conteneur2
                mouvement.numconteneur2 = conteneur2
                mouvement0.typeconteneur2 = type2
                mouvement.typeconteneur2 = type2
                max_size += int(type2) or 0
            if conteneur3:
                mouvement0.numconteneur3 = conteneur3
                mouvement.numconteneur3 = conteneur3
                mouvement0.typeconteneur3 = type3
                mouvement.typeconteneur3 = type3
                max_size += int(type3) or 0
            if conteneur4:
                mouvement0.numconteneur4 = conteneur4
                mouvement.numconteneur4 = conteneur4
                mouvement0.typeconteneur4 = type4
                mouvement.typeconteneur4 = type4
                max_size += int(type4) or 0
            if conteneur5:
                mouvement0.numconteneur5 = conteneur5
                mouvement.numconteneur5 = conteneur5
                mouvement0.typeconteneur5 = type5
                mouvement.typeconteneur5 = type5
                max_size += int(type5) or 0
            if typecam == 'VRAC' and max_size > 40:
                error_message = 'Taille maximum dépassée !!!'
                messages.error(request, error_message)
                return redirect(f'/entreedecalog_view2/{id_user}?error={error_message}')
            mouvement0.marchandise = mouvement_instance.marchandise
            mouvement.marchandise = mouvement_instance.marchandise
            mouvement0.bl1 = mouvement_instance.bl1
            mouvement.bl1 = mouvement_instance.bl1
            mouvement0.pointeur_entree_id = util.id_user
            mouvement.pointeur_entree_id = util.id_user
            mouvement0.bl2 = mouvement_instance.bl2
            mouvement.bl2 = mouvement_instance.bl2
            mouvement0.remorque = remorque
            mouvement.remorque = remorque
            mouvement0.destination = 'zud'
            mouvement.destination = 'zud'
            mouvement0.nbrcolis = mouvement_instance.nbrcolis
            mouvement.nbrcolis = mouvement_instance.nbrcolis
            mouvement0.tonnage = mouvement_instance.tonnage
            mouvement.tonnage = mouvement_instance.tonnage
            mouvement0.date_validite = date_validite
            mouvement.date_validite = date_validite
            mouvement0.navire = mouvement_instance.navire
            mouvement.navire = mouvement_instance.navire
            mouvement0.zone_entree = zone
            mouvement.zone_entree = zone
            mouvement0.date_entree = datetime.now()
            mouvement.date_entree = datetime.now()
            mouvement0.client_id = client
            mouvement.client_id = client
            mouvement0.camion_id = camion
            mouvement.camion_id = camion
            mouvement0.chauffeur_id = chauffeur
            mouvement.chauffeur_id = chauffeur
            mouvement0.transitaire_id = transitaire
            mouvement.transitaire_id = transitaire
            mouvement0.representant_id = representant
            mouvement.representant_id = representant
            mouvement0.save()
            mouvement.save()
            liaison = Liaison.objects.create()
            if conteneur1:
                liaison.conteneur1 = conteneur1
                liaison.type1 = type1
            if conteneur2:
                liaison.conteneur2 = conteneur2
                liaison.type2 = type2
            if conteneur3:
                liaison.conteneur3 = conteneur3
                liaison.type3 = type3
            if conteneur4:
                liaison.conteneur4 = conteneur4
                liaison.type4 = type4
                if conteneur3:
                    liaison.conteneur5 = conteneur5
                    liaison.type5 = type5
            if typecam == 'VRAC':
                liaison.id_mvt_vrac = mouvement_instance
                liaison.id_mvt_sm = mouvement0  # Assigning the newly created Mouvement0 instance
                liaison.save()
                return redirect(f"/entreedecalog_view2/{util.id_user}?success=true")
            elif typecam == 'SEMI-REMORQUE':
                liaison = Liaison.objects.create()
                liaison.id_mvt_vrac = mouvement0
                liaison.id_mvt_sm = mouvement_instance
                liaison.save()
                return redirect(f"/entreedecalog_view2/{util.id_user}?success=true")
            return redirect(f"/entreedecalog_view2/{util.id_user}?error=Invalid typecam")
        except Exception as e:
            error_message = str(e)
            return redirect(f"/entreedecalog_view2/{util.id_user}?error={error_message}")
    return redirect(f"/entreedecalog_view2/{util.id_user}")
    #return render("/")
######################### AJOUT MOUVEMENT PARTICULIER #########################
def ajoutmouvementparticulier(request, id_user):
        util = get_object_or_404(Utilisateurs, id_user=id_user)
        if request.method == 'POST':
            try:
                vehicule = request.POST.get('vehicule')
                destination = request.POST.get('destination')
                zone_entree = request.POST.get('zone')
              #  mouvement0 = Mouvement0(
               #     date_entree =  datetime.now(),
                #    pointeur_entree_id= util.id_user,
                 #   vehicule_id=vehicule,
                  #  destination=destination,
                #)
                #mouvement0.save()
                mouvement8 = Mouvement8(
                    date_entree=datetime.now(),
                    zone_entree = zone_entree,
                    pointeur_entree_id=util.id_user,
                    vehicule_id=vehicule,
                    destination=destination,
             #       id_mvt_0_id=mouvement0.id_mvt
                )
                mouvement8.save()
                # Optional: You might want to redirect to a success page or do something else after saving
                return redirect(f"/entreedecalog_particulier/{util.id_user}?success=true")
            except Exception as e:
                error_message = str(e)
                return redirect(f"/entreedecalog_particulier/{util.id_user}?error={error_message}")
########################## AJOUT MOUVEMENT 0 #####################################
def ajoutmouvement0(request, id_user):
    util = get_object_or_404(Utilisateurs, id_user=id_user)

    if request.method == 'POST':
        try:
            zone = request.POST.get('zone')
            navire = request.POST.get('navire')
            camion_id = request.POST.get('camion')
            chauffeur_id = request.POST.get('chauffeur')
            client_id = request.POST.get('client')
            representant_id =  request.POST.get('representant')
            marchandise = request.POST.get('marchandise')
            mission = request.POST.get('mission')
            date_validite = request.POST.get('date_validite')
            tonnage = request.POST.get('tonnage') or 0
            numconteneur1 = request.POST.get('numconteneur1')
            typeconteneur1 = request.POST.get('typeconteneur1')
            numconteneur2 = request.POST.get('numconteneur2')
            typeconteneur2 = request.POST.get('typeconteneur2')
            numconteneur3 = request.POST.get('numconteneur3')
            typeconteneur3 = request.POST.get('typeconteneur3')
            numconteneur4 = request.POST.get('numconteneur4')
            typeconteneur4 = request.POST.get('typeconteneur4')
            numconteneur5 = request.POST.get('numconteneur5')
            typeconteneur5 = request.POST.get('typeconteneur5')
            bl1 = request.POST.get('bl1')
            bl2 = request.POST.get('bl2')
            nbrcolis = request.POST.get('nbrcolis') or 0
            entite = request.POST.get('entite')
            destination = request.POST.get('destination')
            remorque = request.POST.get('remorque')
            transitaire_id = request.POST.get('transitaire')
            # Création d'un Mouvement0
            mouvement0 = Mouvement0(
                remorque= remorque,
                navire = navire,
                numconteneur1=numconteneur1,
                typeconteneur1=typeconteneur1,
                numconteneur2=numconteneur2,
                typeconteneur2=typeconteneur2,
                numconteneur3=numconteneur3,
                typeconteneur3=typeconteneur3,
                numconteneur4=numconteneur4,
                typeconteneur4=typeconteneur4,
                numconteneur5=numconteneur5,
                typeconteneur5=typeconteneur5,
                bl1=bl1,
                zone_entree = zone,
                date_validite=date_validite,
                bl2=bl2,
                nbrcolis=nbrcolis,
                tonnage=tonnage,
                mission=mission,
                date_entree=timezone.now(),
                camion_id=camion_id,
                transitaire_id=transitaire_id,
                representant_id=representant_id,
                chauffeur_id=chauffeur_id,
                client_id=client_id,
                marchandise=marchandise,
                pointeur_entree_id=util.id_user
            )
            if entite != destination:
                destination = f"{entite} ({destination})"
            else:
                destination = entite
            mouvement0.destination = destination
            camion = Camion.objects.get(id_cam=camion_id)
            if camion.type != 'SEMI-REMORQUE':
                mouvement0.remorque = camion.immatriculation
            else:
                mouvement0.remorque = remorque
            mouvement0.save()

            # Obtenir le dernier mouvement0 pour l'entité donnée
           # mvt0 = Mouvement0.objects.filter(destination__startswith=entite).last()
            # Obtenir le dernier mouvement0 pour l'entité donnée
            mvt0 = mouvement0.id_mvt

            # Initialize common attributes
            common_attrs = {
                'camion_id':camion_id,
                'chauffeur_id':chauffeur_id,
                'zone_entree' : zone,
                'client_id':client_id,
                'mission': mission,
                'remorque': remorque,
                'id_mvt_0_id': mvt0,
                'pointeur_entree_id': util.id_user
            }
            # Determine movement type based on destination
            movement_type = None
            destination = destination
            if 'icdtom' in destination:
                movement_type = Mouvement1
            elif 'cmaicd' in destination:
                movement_type = Mouvement4
            elif 'tom' in destination and 'hangar' in destination:
                movement_type = Mouvement2
            elif 'its' in destination and 'hangar' in destination:
                movement_type = Mouvement6
            elif 'transexpress' in destination and 'hangar' in destination:
                movement_type = Mouvement7
            elif 'zud' in destination:
                movement_type = Mouvement3
            if movement_type:
                # Initialize mouvement with common attributes
                mouvement = movement_type(**common_attrs)
                # Add specific attributes for certain movement types
                if destination ==  'zud':
                    mouvement.marchandise = marchandise
                    mouvement.bl1 = bl1
                    mouvement.bl2 = bl2
                    mouvement.numconteneur1 = numconteneur3
                    mouvement.typeconteneur1 = typeconteneur3
                    mouvement.numconteneur2 = numconteneur3
                    mouvement.typeconteneur2 = typeconteneur3
                    mouvement.numconteneur3 = numconteneur3
                    mouvement.typeconteneur3 = typeconteneur3
                    mouvement.numconteneur4 = numconteneur4
                    mouvement.typeconteneur4 = typeconteneur4
                    mouvement.numconteneur5 = numconteneur5
                    mouvement.typeconteneur5 = typeconteneur5
                    mouvement.transitaire_id = transitaire_id
                    mouvement.representant_id = representant_id
                    mouvement.navire = navire
                    mouvement.tonnage = tonnage
                    mouvement.nbrcolis = nbrcolis
                mouvement.save()
                if 'hangar' in destination:
                    mouvement.marchandise = marchandise
            if 'hangar' in destination :
                return redirect(f"/entreedecalog_view1/{util.id_user}?success=true")
            elif destination == 'zud':
                return redirect(f"/entreedecalog_view2/{util.id_user}?success=true")
            else :
                return redirect(f"/entreedecalog_view/{util.id_user}?success=true")
        except Exception as e:
            error_message = str(e)
            return redirect(f"/entreedecalog_view/{util.id_user}?error={error_message}")
    else:

        return render(request, 'pages/mouvement_entre_0.html', {
            'util': util,
        })


def entree_tom(request):
    if request.method == 'POST':
        id_mvt = request.POST.get('id_mvt')
        id_user = request.POST.get('id_user')
        util = get_object_or_404(Utilisateurs, id_user=id_user)
        mouvement = get_object_or_404(Mouvement1, id_mvt=id_mvt)
        mouvement.pointeur_entree_id = util.id_user
        mouvement.date_entree = timezone.now()
        mouvement.mission = request.POST.get('mission')
        mouvement.save()
        return redirect('/index1_view/' + str(util.id_user))
def entree_camion(request):
    if request.method == 'POST':
        form = SortieForm(request.POST)
        if form.is_valid():
            id_mvt = request.POST.get('id_mvt')
            id_user = request.POST.get('id_user')
            source = request.POST.get('source')
            destination = request.POST.get('destination')
            # Récupérer les objets Utilisateurs et Mouvement en fonction des IDs fournis
            util = get_object_or_404(Utilisateurs, id_user=id_user)
            if (source=='icdtom'):
                mouvement = get_object_or_404(Mouvement1, id_mvt=id_mvt)
            elif (source=='icdcma'):
                mouvement = get_object_or_404(Mouvement4, id_mvt=id_mvt)
            elif (source == 'zud'):
                mouvement = get_object_or_404(Mouvement3, id_mvt=id_mvt)
            elif (source == 'sacherie'):
                if (destination) == 'TOM':
                    mouvement = get_object_or_404(Mouvement2, id_mvt=id_mvt)
                elif (destination) == 'ITS':
                    mouvement = get_object_or_404(Mouvement6, id_mvt=id_mvt)
                elif (destination) == 'TRANSEXPRESS':
                    mouvement = get_object_or_404(Mouvement7, id_mvt=id_mvt)
            mouvement.pointeur_entree_id=id_user
            mouvement.date_entree = timezone.now()
            mouvement.mission =  request.POST.get(
                'mission')
            mouvement.save()
            # Rediriger vers une vue après sauvegarde
            if (source == 'icdtom'):
                return redirect(f'/index1_view/{util.id_user}?success=true')
            elif (source =='zud'):
                return redirect(f'/index_entree_zud/{util.id_user}?success=true')
            elif(source == 'icdcma'):
                return redirect(f'/index_entree_icdcma/{util.id_user}?success=true')
            elif (source == 'sacherie'):
                return redirect(f'/index2_view/{util.id_user}?success=true')
    else:
        form = SortieForm()

    # Assurez-vous que 'mouvement' et 'util' sont définis si vous les utilisez dans le template
    return render(request, 'pages/ajoutsortiedp.html', {'form': form})
################################# SORTIE DKLOG POUR PARTICULIER ##################
def sortie_decalog_viewparticulier(request, id_user):
  util = Utilisateurs.objects.get(id_user=id_user)
  return render(request, 'pages/mouvement_sortie_particulier.html', {'util': util})
################################# SORTIE DKLOG POUR ICD TOM ##################
def sortie_decalog_view(request, id_user):
  util = Utilisateurs.objects.get(id_user=id_user)
  return render(request, 'pages/mouvement_sortie_0.html', {'util': util})
################################# SORTIE DKLOG POUR ICD CMA ##################
def sortie_decalog_view01(request, id_user):
  util = Utilisateurs.objects.get(id_user=id_user)
  return render(request, 'pages/mouvement_sortie_001.html', {'util': util})
################################# SORTIE DKLOG POUR SACHERIE ##################
def sortie_decalog_view1(request, id_user):
  util = Utilisateurs.objects.get(id_user=id_user)
  return render(request, 'pages/mouvement_sortie_01.html', {'util': util})
################################# SORTIE DKLOG POUR ZUD ##################
def sortie_decalog_view2(request, id_user):
    util = Utilisateurs.objects.get(id_user=id_user)
    return render(request, 'pages/mouvement_sortie_02.html', {'util': util})
@csrf_exempt
def ajouter_client(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    return JsonResponse({'success': False, 'errors': 'Méthode non autorisée'})
@csrf_exempt

def ajouter_camion_dk_log(request):
    if request.method == 'POST':
        form = CamionForm(request.POST)
        if form.is_valid() or 1:
            camion = form.save()
            # Retourner une réponse JSON indiquant le succès de l'enregistrement
            return JsonResponse({'success': True})
        else:
            # Retourner une réponse JSON avec des erreurs de formulaire
            return JsonResponse({'success': False, 'errors': form.errors})

    return JsonResponse({'success': False, 'errors': 'Méthode non autorisée'})
def ajouter_vehicule_particulier(request):
            if request.method == 'POST':
                form = VehiculeForm(request.POST)
                if form.is_valid() or 1:
                    camion = form.save()
                    # Retourner une réponse JSON indiquant le succès de l'enregistrement
                    return JsonResponse({'success': True})
                else:
                    # Retourner une réponse JSON avec des erreurs de formulaire
                    return JsonResponse({'success': False, 'errors': form.errors})

            return JsonResponse({'success': False, 'errors': 'Méthode non autorisée'})


@csrf_exempt

def ajouter_chauffeurs(request):
    if request.method == 'POST':
        form = ChauffeurForm(request.POST)
        if form.is_valid():
            chauffeur = form.save()
            # Retourner une réponse JSON indiquant le succès de l'enregistrement
            #return JsonResponse({'success': True, 'chauffeur': {'id': chauffeur.id_chauffeur, 'fullname': chauffeur.fullname}})
            return JsonResponse({
                'success': True,
                'chauffeur': {
                    'id': chauffeur.id_chauffeur,
                    'fullname': chauffeur.fullname
                }
            })

        else:
            # Retourner une réponse JSON avec des erreurs de formulaire
            return JsonResponse({'success': False, 'errors': form.errors})

    return JsonResponse({'success': False, 'errors': 'Méthode non autorisée'})



def fetch_camion(request):
    camions=Camion.objects.all().values('id_cam','transporteur','immatriculation','type')
    return JsonResponse(list(camions), safe=False)
def fetch_vehicule(request):
    vehicules=Vehicule.objects.all().values('id_veh','immatriculation')
    return JsonResponse(list(vehicules), safe=False)
def fetch_camion_sacherie(request):
    camions = Camion.objects.filter(type__in=['VRAC', 'PLATEAU']).values('id_cam', 'transporteur', 'immatriculation', 'type')
    return JsonResponse(list(camions), safe=False)

def fetch_chauffeur(request):
    chauffeurs=Chaffeur.objects.all().values('id_chauffeur','fullname','permis')
    return JsonResponse(list(chauffeurs), safe=False)
def fetch_transitaire(request):
    transitaires = Transitaire.objects.filter(mission='transitaire').values('id_transit','fullname', 'telephone')
    return JsonResponse(list(transitaires), safe=False)
def fetch_representant(request):
    transitaires = Transitaire.objects.filter(mission='representant').values('id_transit', 'fullname', 'telephone')
    return JsonResponse(list(transitaires), safe=False)
########################### GESTION DES FETCHS DES MOUVEMENTS SUR LE DASHBOARD #############
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta, datetime


######################### GESTION DES FETCHS DES MOUVEMENTS SUR LE DASHBOARD ##################
#################### PLT ################################
def fetch_mouvementsdash0(request):
    maintenant = timezone.now()
    mouvements = Mouvement0.objects.filter(date_sortie__isnull=True).order_by('-date_entree')[:50]
    mouvements_avec_attributs = []
    for mvt in mouvements:
        try:
            camion = Camion.objects.get(id_cam=mvt.camion_id)
        except Camion.DoesNotExist:
            # Si le camion n'existe pas, on crée un camion par défaut
            camion = Camion(
                id_cam=None,  # ou vous pouvez laisser `None` si vous ne voulez pas affecter un ID spécifique
                immatriculation="Non assigné",
                transporteur="Non assigné",
                type="Non assigné",
            )
        mvt.cam = camion.immatriculation
        try:
            para = ParametrageDelais.objects.get(entite='dklog', type=camion.type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=10080)
            delais_urg = timedelta(minutes=7200)

        diff = (maintenant - mvt.date_entree).total_seconds() / 60
        diff_seconds = (maintenant - mvt.date_entree).total_seconds()
        hours, remainder = divmod(diff_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        mvt.chrono = f'{int(hours):02}:{int(minutes):02}:{int(seconds):02}'

        if timedelta(minutes=diff) >= duree_dk:
            mvt.etat = 3
        elif timedelta(minutes=diff) < duree_dk and timedelta(minutes=diff) >= delais_urg:
            mvt.etat = 2
        else:
            mvt.etat = 1

        mouvements_avec_attributs.append({
            'cam': mvt.cam,
            'destination': mvt.destination.upper(),
            'chrono': mvt.chrono,
            'etat': mvt.etat
        })

    return JsonResponse({'mouvements': mouvements_avec_attributs})
#################### ICD TOM  #################################
def fetch_mouvementsdash(request):
    maintenant = timezone.now()
    mouvements = Mouvement1.objects.filter(date_sortie__isnull=True).order_by('-date_entree')[:50]
    mouvements_avec_attributs = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation

        try:
            para = ParametrageDelais.objects.get(entite='icd', type=camion.type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)
        if (mvt.date_entree):
            pass
        else:
            mvt.date_entree=maintenant
        diff = (maintenant - mvt.date_entree).total_seconds() / 60
        diff_seconds = (maintenant - mvt.date_entree).total_seconds()
        hours, remainder = divmod(diff_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        mvt.chrono = f'{int(hours):02}:{int(minutes):02}:{int(seconds):02}'

        if timedelta(minutes=diff) >= duree_dk:
            mvt.etat = 3
        elif timedelta(minutes=diff) < duree_dk and timedelta(minutes=diff) >= delais_urg:
            mvt.etat = 2
        else:
            mvt.etat = 1

        mouvements_avec_attributs.append({
            'cam': mvt.cam,
            'chrono': mvt.chrono,
            'etat': mvt.etat
        })
    return JsonResponse({'mouvements': mouvements_avec_attributs})
#################### ICD CMA  #################################
def fetch_mouvementsdash1(request):
    maintenant = timezone.now()
    mouvements = Mouvement4.objects.filter(date_sortie__isnull=True).order_by('-date_entree')[:50]
    mouvements_avec_attributs = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        try:
            para = ParametrageDelais.objects.get(entite='icd', type=camion.type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        if mvt.date_entree:
            pass
        else:
            mvt.date_entree=maintenant
        diff = (maintenant - mvt.date_entree).total_seconds() / 60
        diff_seconds = (maintenant - mvt.date_entree).total_seconds()
        hours, remainder = divmod(diff_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        mvt.chrono = f'{int(hours):02}:{int(minutes):02}:{int(seconds):02}'
        if timedelta(minutes=diff) >= duree_dk:
            mvt.etat = 3
        elif timedelta(minutes=diff) < duree_dk and timedelta(minutes=diff) >= delais_urg:
            mvt.etat = 2
        else:
            mvt.etat = 1

        mouvements_avec_attributs.append({
            'cam': mvt.cam,
            'chrono': mvt.chrono,
            'etat': mvt.etat
        })

    return JsonResponse({'mouvements': mouvements_avec_attributs})
#################### SACHERIE  #################################
def fetch_mouvementsdash2(request):
    maintenant = timezone.now()
    # Récupérer les mouvements des trois modèles
    mouvements2 = Mouvement2.objects.filter(date_sortie__isnull=True).order_by('-date_entree')[:50]
    mouvements6 = Mouvement6.objects.filter(date_sortie__isnull=True).order_by('-date_entree')[:50]
    mouvements7 = Mouvement7.objects.filter(date_sortie__isnull=True).order_by('-date_entree')[:50]

    # Combiner les requêtes en une seule liste
    mouvements_combines = chain(mouvements2, mouvements6, mouvements7)

    mouvements_avec_attributs = []
    for mvt in mouvements_combines:
        # Gestion du camion
        try:
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation if camion.immatriculation else "Non Assigné"
        except Camion.DoesNotExist:
            mvt.cam = "Non Assigné"

        # Gestion des délais via ParametrageDelais
        try:
            para = ParametrageDelais.objects.get(entite='sacherie', type=camion.type)
            duree_dk = timedelta(minutes=int(para.delais_maximal)) if para.delais_maximal else timedelta(minutes=30)
            delais_urg = timedelta(minutes=int(para.delais_urgent)) if para.delais_urgent else timedelta(minutes=20)
        except ParametrageDelais.DoesNotExist:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        # Calculer la date d'entrée si elle est manquante
        mvt.date_entree = mvt.date_entree if mvt.date_entree else maintenant
        # Calculer la différence en minutes et en secondes
        diff = (maintenant - mvt.date_entree).total_seconds() / 60
        diff_seconds = (maintenant - mvt.date_entree).total_seconds()
        hours, remainder = divmod(diff_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        mvt.chrono = f'{int(hours):02}:{int(minutes):02}:{int(seconds):02}'

        # Définir l'état en fonction des délais
        if timedelta(minutes=diff) >= duree_dk:
            mvt.etat = 3  # Urgence
        elif timedelta(minutes=diff) < duree_dk and timedelta(minutes=diff) >= delais_urg:
            mvt.etat = 2  # Alerte
        else:
            mvt.etat = 1  # Normal

        # Ajouter la variable Destination selon le modèle du mouvement
        if isinstance(mvt, Mouvement2):
            mvt.destination = "TOM"
        elif isinstance(mvt, Mouvement6):
            mvt.destination = "ITS"
        elif isinstance(mvt, Mouvement7):
            mvt.destination = "TRANSEXPRESS"
        else:
            mvt.destination = "Non Assigné"

        # Ajouter les informations à la liste
        mouvements_avec_attributs.append({
            'cam': mvt.cam,
            'chrono': mvt.chrono,
            'etat': mvt.etat,
            'destination': mvt.destination  # Inclure la destination
        })

    # Limiter les résultats à 50 mouvements au total
    mouvements_avec_attributs = mouvements_avec_attributs[:50]

    return JsonResponse({'mouvements': mouvements_avec_attributs})
#################### ZUD  #################################
def fetch_mouvementsdash3(request):
    maintenant = timezone.now()
    mouvements = Mouvement3.objects.filter(date_sortie__isnull=True).order_by('-date_entree')[:50]
    mouvements_avec_attributs = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation

        try:
            para = ParametrageDelais.objects.get(entite='zud', type=camion.type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        if (mvt.date_entree):
            pass
        else:
            mvt.date_entree = maintenant
        diff = (maintenant - mvt.date_entree).total_seconds() / 60
        diff_seconds = (maintenant - mvt.date_entree).total_seconds()
        hours, remainder = divmod(diff_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        mvt.chrono = f'{int(hours):02}:{int(minutes):02}:{int(seconds):02}'

        if timedelta(minutes=diff) >= duree_dk:
            mvt.etat = 3
        elif timedelta(minutes=diff) < duree_dk and timedelta(minutes=diff) >= delais_urg:
            mvt.etat = 2
        else:
            mvt.etat = 1

        mouvements_avec_attributs.append({
            'cam': mvt.cam,
            'chrono': mvt.chrono,
            'etat': mvt.etat
        })
    return JsonResponse({'mouvements': mouvements_avec_attributs})
def fetch_mouvementsdashparticulier(request):
    maintenant = timezone.now()
    mouvements = Mouvement8.objects.filter(date_sortie__isnull=True)[:50]
    mouvements_avec_attributs = []
    for mvt in mouvements:
        vehicule = Vehicule.objects.get(id_veh=mvt.vehicule_id)
        mvt.cam = vehicule.immatriculation
        try:
            para = 30
            duree_dk = timedelta(minutes=int(para))
            delais_urg = timedelta(minutes=int(para))
        except:
            duree_dk = timedelta(minutes=480)
            delais_urg = timedelta(minutes=20)
        if (mvt.date_entree):
            pass
        else:
            mvt.date_entree=maintenant
        diff = (maintenant - mvt.date_entree).total_seconds() / 60
        diff_seconds = (maintenant - mvt.date_entree).total_seconds()
        hours, remainder = divmod(diff_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        mvt.chrono = f'{int(hours):02}:{int(minutes):02}:{int(seconds):02}'
        if timedelta(minutes=diff) >= duree_dk:
            mvt.etat = 3
        elif timedelta(minutes=diff) < duree_dk and timedelta(minutes=diff) >= delais_urg:
            mvt.etat = 2
        else:
            mvt.etat = 1

        mouvements_avec_attributs.append({
            'cam': mvt.cam,
            'chrono': mvt.chrono,
            'etat': mvt.etat
        })

    return JsonResponse({'mouvements': mouvements_avec_attributs})


from datetime import timedelta
from django.utils import timezone
############################ DASHBOARD ICD TOM ##########################
def index_dk0(request, id):
    user = Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index_dk_log0.html',
                  { 'util': user})

############################ DASHBOARD ICD TOM ##########################
def index_dk(request, id):
    user = Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index_dk_log.html',
                  {'util': user})

######################################## DASHBOARD ICD CMA #####################################
def index_dk1(request, id):
    user = Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index_dk_log1.html',
                  {'util': user})
from itertools import chain
################################## DASHBOARD SACHERIE #################################
def index_dk2(request, id):
    user = Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index_dk_log2.html',
                  {'util': user})

################################### DASHBOARD ZUD ################################
def index_dk3(request, id):
    user = Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index_dk_log3.html',
                  {'util': user})
################################### DASHBOARD PARTICULIER ################################
def index_dkparticulier(request, id):
        user = Utilisateurs.objects.get(id_user=id)
        return render(request, 'pages/index_dk_logparticulier.html',
                  {'util': user})
######################## RENDER FOR SACHERIE USERS ########################
################ ENTREE ###########################"
def index2_view(request, id):
  util = Utilisateurs.objects.get(id_user=id)
  return render(request, 'pages/entree_mouvement_1.html', {'util': util})
####################### SORTIE ##########################
def index3_view(request, id):
  util = Utilisateurs.objects.get(id_user=id)
  return render(request, 'pages/sortie_mouvement_1.html', {'util': util})
def entree_sacherie(request):
    if request.method == 'POST':
        form = SortieForm(request.POST)
        if form.is_valid():
            id_mvt = request.POST.get('id_mvt')
            id_user = request.POST.get('id_user')

            # Récupérer les objets Utilisateurs et Mouvement en fonction des IDs fournis
            util = get_object_or_404(Utilisateurs, id_user=id_user)
            mouvement = get_object_or_404(Mouvement2, id_mvt=id_mvt)

            mouvement.date_entree = timezone.now()
            mouvement.pointeur_entree_id = id_user
            mouvement.mission =  request.POST.get(
                'mission')  # Récupérer le statut depuis les données validées du formulaire
            mouvement.save()

            # Rediriger vers une vue après sauvegarde
            return redirect(f'/index2_view/{util.id_user}')
    else:
        form = SortieForm()

    # Assurez-vous que 'mouvement' et 'util' sont définis si vous les utilisez dans le template
    return render(request, 'pages/ajoutsortiedp.html', {'form': form})
def ajoutsortie_sacherie(request):
        if request.method == 'POST':
            form = SortieForm(request.POST)
            if form.is_valid():
                id_mvt = request.POST.get('id_mvt')
                id_user = request.POST.get('id_user')
                # Récupérer les objets Utilisateurs et Mouvement en fonction des IDs fournis
                util = get_object_or_404(Utilisateurs, id_user=id_user)
                mouvement = get_object_or_404(Mouvement2, id_mvt=id_mvt)
                # Mettre à jour les informations du mouvement
                mouvement.date_sortie = timezone.now()
                mouvement.pointeur_sortie_id = util.id_user
                mouvement.statut_sortie = form.cleaned_data.get(
                    'statut_sortie')  # Récupérer le statut depuis les données validées du formulaire
                mouvement.save()
                # Rediriger vers une vue après sauvegarde
                return redirect(f'/index3_view/{util.id_user}')
        else:
            form = SortieForm()
        # Assurez-vous que 'mouvement' et 'util' sont définis si vous les utilisez dans le template
        return render(request, 'pages/ajoutsortiedp.html', {'form': form})


def liste_mouvements_3(request, id_user):
    # Récupérer les mouvements de Mouvement2, Mouvement6 et Mouvement7
    mouvements2 = Mouvement2.objects.filter(date_entree__isnull=True, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvements6 = Mouvement6.objects.filter(date_entree__isnull=True, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvements7 = Mouvement7.objects.filter(date_entree__isnull=True, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )

    # Convertir les QuerySets en listes et ajouter la clé `destination`
    mouvement_list = list(mouvements2) + list(mouvements6) + list(mouvements7)

    # Ajouter la variable `destination` et gérer les valeurs manquantes
    for mouvement in mouvement_list:
        # Vérifier dans quelle liste se trouve le mouvement et ajouter la destination correspondante
        if mouvement in mouvements2:
            mouvement['destination'] = 'TOM'
        elif mouvement in mouvements6:
            mouvement['destination'] = 'ITS'
        elif mouvement in mouvements7:
            mouvement['destination'] = 'TRANSEXPRESS'

        # Ajouter les informations sur le camion, sinon "Non assigné"
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        if camion:
            mouvement['camion'] = camion
        else:
            mouvement['camion'] = {
                'id_cam': 'Non assigné',
                'immatriculation': 'Non assigné',
                'transporteur': 'Non assigné',
                'type': 'Non assigné'
            }

        # Ajouter les informations sur le chauffeur, sinon "Non assigné"
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        if chauffeur:
            mouvement['chauffeur'] = chauffeur
        else:
            mouvement['chauffeur'] = {
                'id_chauffeur': 'Non assigné',
                'fullname': 'Non assigné',
                'permis': 'Non assigné'
            }

        # Vérifier et ajouter les autres champs manquants avec "Non assigné"
        mouvement['mission'] = mouvement.get('mission', 'Non assigné')
        mouvement['statut_entree'] = mouvement.get('statut_entree', 'Non assigné')
        mouvement['statut_sortie'] = mouvement.get('statut_sortie', 'Non assigné')
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')

    # Préparer la réponse JSON avec les mouvements combinés
    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }

    return JsonResponse(response_data)
def liste_mouvements4(request, id_user):
    # Récupérer les mouvements de Mouvement2, Mouvement6 et Mouvement7
    mouvements2 = Mouvement2.objects.filter(date_entree__isnull=False, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvements6 = Mouvement6.objects.filter(date_entree__isnull=False, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )
    mouvements7 = Mouvement7.objects.filter(date_entree__isnull=False, date_sortie__isnull=True).values(
        'id_mvt', 'mission', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque'
    )

    # Convertir les QuerySets en listes
    mouvement_list = list(mouvements2) + list(mouvements6) + list(mouvements7)

    for mouvement in mouvement_list:
        # Déterminer la destination en fonction du modèle d'origine
        if mouvement in mouvements2:
            mouvement['destination'] = 'TOM'
        elif mouvement in mouvements6:
            mouvement['destination'] = 'ITS'
        elif mouvement in mouvements7:
            mouvement['destination'] = 'TRANSEXPRESS'

        # Ajouter les informations sur le camion, sinon "Non assigné"
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        if camion:
            mouvement['camion'] = camion
        else:
            mouvement['camion'] = {
                'id_cam': 'Non assigné',
                'immatriculation': 'Non assigné',
                'transporteur': 'Non assigné',
                'type': 'Non assigné'
            }

        # Ajouter les informations sur le chauffeur, sinon "Non assigné"
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        if chauffeur:
            mouvement['chauffeur'] = chauffeur
        else:
            mouvement['chauffeur'] = {
                'id_chauffeur': 'Non assigné',
                'fullname': 'Non assigné',
                'permis': 'Non assigné'
            }

        # Vérifier et ajouter les autres champs manquants avec "Non assigné"
        mouvement['mission'] = mouvement.get('mission', 'Non assigné')
        mouvement['statut_entree'] = mouvement.get('statut_entree', 'Non assigné')
        mouvement['statut_sortie'] = mouvement.get('statut_sortie', 'Non assigné')
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')

    # Préparer la réponse JSON avec les mouvements combinés
    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }

    return JsonResponse(response_data)

def fetch_camion(request):
    camions=Camion.objects.all().values('id_cam','transporteur','immatriculation','type')
    return JsonResponse(list(camions), safe=False)

def fetch_client(request):
    clients = Client.objects.all().values('id_client', 'fullname', 'telephone')
    return JsonResponse(list(clients), safe=False)
def fetch_chauffeur(request):
    chauffeurs=Chaffeur.objects.all().values('id_chauffeur','fullname','permis')
    return JsonResponse(list(chauffeurs), safe=False)
def index_sacherie(request,id):
    duree_dk = 30
    delais_urg=20
    maintenant = timezone.now()
    try:
        para = ParametrageDelais.objects.get(entite='sacherie', type='VRAC')
        duree_dk = timedelta(minutes=int(para.delais_maximal))
        delais_urg=timedelta(minutes=int(para.delais_urgent))
    except:
        # pass
        duree_dk = timedelta(minutes=30)
        delais_urg = timedelta(minutes=20)
    mouvements = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)
    mouvements_filtrés = [mvt for mvt in mouvements if (maintenant - mvt.date_entree) >= delais_urg and (
                maintenant - mvt.date_entree) < duree_dk]
    urg = len(mouvements_filtrés)
   ###Depassement
    mouvements_dep = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)
    mouvements_filtrés_dep = [mvt for mvt in mouvements_dep if (maintenant - mvt.date_entree) >= duree_dk]
    dep = len(mouvements_filtrés_dep)
    ###Plus 30
    vingt_minutes_30 = timedelta(minutes=30)
    maintenant = timezone.now()
    maintenant = timezone.now()
    mouvements_30 = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=False)
    mouvements_filtrés_30 = [mvt for mvt in mouvements_30 if (mvt.date_sortie - mvt.date_entree) >= duree_dk]
    lg_30 = len(mouvements_filtrés_30)
    #####OIMS 30
    vingt_minutes_moin = timedelta(minutes=30)
    maintenant = timezone.now()
    mouvements_moins = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=False)
    mouvements_filtrés_moins = [mvt for mvt in mouvements_moins if
                                (mvt.date_sortie - mvt.date_entree) < duree_dk]
    lg_moins = len(mouvements_filtrés_moins)
    mvt_total=Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)
    total_cours =len(mvt_total)
    totalter = lg_30 + lg_moins
    tout_mouvements= Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)[:50]
    vingt_minutes = timedelta(minutes=30)
    trt = timedelta(minutes=20)
    for mvt in tout_mouvements :
        diff = (maintenant - mvt.date_entree).total_seconds()/60
        mvt.chrono=int(diff)
        #duree=math.ceil((mvt.date_sortie - mvt.date_entree).total_seconds() / 3600)
        duree=float(diff)
        if timedelta(minutes=duree) >= duree_dk:
            mvt.etat= 3
        if timedelta(minutes=duree) < duree_dk and timedelta(minutes=duree) >= delais_urg :
            mvt.etat= 2
        if timedelta(minutes=duree) <= delais_urg :
          mvt.etat= 1
        camion=Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam=camion.immatriculation
    user=Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index_sacherie.html',
                  { 'urg': urg, 'dep': dep, 'lg_30': lg_30, 'lg_mois': lg_moins,
                   'total_cours': total_cours, 'total_ter': totalter,'mvt':tout_mouvements,'util' :user})
###############Gestion details Sacherie
############Gestion Detail
def detail_urgent_sacherie(request,id):
   duree_dk = 30
   delais_urg = 20
   maintenant = timezone.now()

   try:

        para = ParametrageDelais.objects.get(entite='sacherie', type='VRAC')
        duree_dk = timedelta(minutes=int(para.delais_maximal))
        delais_urg = timedelta(minutes=int(para.delais_urgent))
   except:
        # pass
        duree_dk = timedelta(minutes=30)
        delais_urg = timedelta(minutes=20)

   mouvements = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)
   mouvements_filtrés = [mvt for mvt in mouvements if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk ]
   for mvt in mouvements_filtrés :
      chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
      mvt.chauffeur_name = chauf.fullname
      poineur= Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
      mvt.pointeur=poineur.fullname
      camion=Camion.objects.get(id_cam=mvt.camion_id)
      mvt.trans=camion.transporteur
      mvt.imat=camion.immatriculation
   lg=len(mouvements_filtrés)
   total_encours=len(mouvements)
   if total_encours >= 1 :
    pourcentage_urgent=(lg/total_encours) * 100
   else :
    pourcentage_urgent=0

   user = Utilisateurs.objects.get(id_user=id)
   context = {
    'segment': 'index',
    'mvt' : mouvements_filtrés,
    # 'products' : Product.objects.all()
      'lg' :lg,
      'pourcentage_urgent' :int(pourcentage_urgent),
      'util': user
  }

   return render(request, "pages/detail_urgent_sacherie.html", context)

def detail_depassement_sacherie(request, id):
    duree_dk = 30
    delais_urg = 20
    maintenant = timezone.now()

    try:

        para = ParametrageDelais.objects.get(entite='sacherie', type='VRAC')
        duree_dk = timedelta(minutes=int(para.delais_maximal))
        delais_urg = timedelta(minutes=int(para.delais_urgent))
    except:
        # pass
        duree_dk = timedelta(minutes=30)
        delais_urg = timedelta(minutes=20)
    mouvements = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=True)
    mouvements_filtrés = [mvt for mvt in mouvements if (maintenant - mvt.date_entree) > duree_dk]
    for mvt in mouvements_filtrés:
        chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauf.fullname
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
    lg = len(mouvements_filtrés)
    total_depassemnt=len(mouvements)
    if total_depassemnt >= 1:
       pourcentage_depassement = (lg / total_depassemnt) * 100
    else :
      pourcentage_depassement= 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'pourcentage_depassement' : int(pourcentage_depassement),
        'util': user
    }

    return render(request, "pages/detail_depassement_sacherie.html", context)

def detail_plus_30_sacherie(request, id):
    duree_dk = 30
    delais_urg = 20
    maintenant = timezone.now()

    try:

        para = ParametrageDelais.objects.get(entite='sacherie', type='VRAC')
        duree_dk = timedelta(minutes=int(para.delais_maximal))
        delais_urg = timedelta(minutes=int(para.delais_urgent))
    except:
        # pass
        duree_dk = timedelta(minutes=30)
        delais_urg = timedelta(minutes=20)
    mouvements = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=False)
    mouvements_filtrés = [mvt for mvt in mouvements if (mvt.date_sortie - mvt.date_entree) > duree_dk]
    for mvt in mouvements_filtrés:
        chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauf.fullname
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname

        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)

        mvt.pointeur_srt = poineur_srt.fullname
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
    lg = len(mouvements_filtrés)
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    if total_plus_30 >= 1:
      pourcentage_plus_30 = (lg / total_plus_30) * 100
    else :
      pourcentage_plus_30=0
    user = Utilisateurs.objects.get(id_user=id)

    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'util': user,
        'pourcentage_plus_30' : int( pourcentage_plus_30)
    }

    return render(request, "pages/detail_plus_30_sacherie.html", context)
def detail_moins_30_sacherie(request, id):

    duree_dk = 30

    try:

        para = ParametrageDelais.objects.get(entite='sacherie', type='VRAC')
        duree_dk = timedelta(minutes=int(para.duree))
    except:
        # pass
        duree_dk = timedelta(minutes=30)
    # duree_trt = timedelta(minutes=30)
    vingt_minutes = timedelta(minutes=20)
    delais_urg = duree_dk - vingt_minutes
    vingt_minutes = timedelta(minutes=30)
    maintenant = timezone.now()
    mouvements = Mouvement2.objects.filter(date_entree__isnull=False,date_sortie__isnull=False)
    mouvements_filtrés = [mvt for mvt in mouvements if (mvt.date_sortie - mvt.date_entree) <= duree_dk]
    for mvt in mouvements_filtrés:
        chauf = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauf.fullname
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    if  total_moins_30 >= 1 :
       pourcentage_moins_30 = (lg / total_moins_30) * 100
    else :
        pourcentage_moins_30=0
    user = Utilisateurs.objects.get(id_user=id)

    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'pourcentage_moins_30' :int(pourcentage_moins_30),
        'util': user
    }

    #return render(request, "pages/detail_depassement.html", context)
    return render(request, "pages/detail_moins_30_sacherie.html", context)

################Detail DK LOG
############Gestion Detail


###################### DETAIL URGENT PARTICULIER #####################################
def detail_urgent_dk_logpar(request, id):
    maintenant = timezone.now()
    delais_urg = timedelta(minutes=15)  # Par défaut 15 minutes pour l'urgence
    duree_dk = timedelta(minutes=30)  # Par défaut 30 minutes pour le délai maximal

    # Récupération des mouvements non sortis
    mouvements = Mouvement8.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []

    for mvt in mouvements:
        try:
            # Récupérer les informations du véhicule associé au mouvement
            camion = Vehicule.objects.get(id_veh=mvt.vehicule_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
        except Vehicule.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.imat = "Non assigné"

        try:
            # Récupérer le pointeur d'entrée
            pointeur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = pointeur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"

        # Calculer la durée du mouvement par rapport à l'urgence
        temps_ecoule = maintenant - mvt.date_entree
        if delais_urg <= temps_ecoule < duree_dk:
            mouvements_filtrés.append(mvt)

    # Calcul du nombre total de mouvements et du pourcentage urgent
    total_encours = len(mouvements)
    lg = len(mouvements_filtrés)
    pourcentage_urgent = (lg / total_encours) * 100 if total_encours >= 1 else 0

    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)

    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_urgent': int(pourcentage_urgent),
        'util': user
    }

    # Rendu du template avec le contexte
    return render(request, "pages/detail_urgent_dk_logpar.html", context)
###################### DETAIL URGENT PLT #####################################
def detail_urgent_dk_log0(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []

    for mvt in mouvements:
        try:
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            type = camion.type or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            type = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"
        try:
            para = ParametrageDelais.objects.get(entite='dklog', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            duree_dk = timedelta(minutes=10080)
            delais_urg = timedelta(minutes=7200)
        try:
            pointeur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = pointeur.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"

        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"

        # Filtrage basé sur les délais
        if delais_urg <= (maintenant - mvt.date_entree) < duree_dk:
            mouvements_filtrés.append(mvt)

    # Calcul des statistiques
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    pourcentage_urgent = (lg / total_encours) * 100 if total_encours >= 1 else 0

    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)

    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_urgent': int(pourcentage_urgent),
        'util': user
    }

    return render(request, "pages/detail_urgent_dk_log0.html", context)
###################### DETAIL URGENT ICD TOM #####################################
def detail_urgent_dk_log(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_avec_attributs = []
    mouvements_filtrés=[]
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='icd', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        try:
            chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        except Chaffeur.DoesNotExist:
            chauffeur_name = {'id_chauffeur': None, 'fullname': 'Non Assigné', 'permis': 'Non Assigné',
                              'telephone': 'Non Assigné'}

        # Vérifiez si chauffeur_name est un dictionnaire ou un objet Chauffeur
        mvt.chauffeur_name = chauffeur_name['fullname'] if isinstance(chauffeur_name, dict) else chauffeur_name.fullname

        if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk :
            mouvements_filtrés.append(mvt)
        # Ajoutez l'objet mouvement enrichi à la liste
        #mouvements_avec_attributs.append(mvt)
    # Appliquez le filtre après avoir enrichi les objets mouvement
    #mouvements_filtrés = [mvt for mvt in mouvements_avec_attributs if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk]
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_urgent = (lg / total_encours) * 100
    else:
        pourcentage_urgent = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_urgent': int(pourcentage_urgent),
        'util': user
    }

    return render(request, "pages/detail_urgent_dk_log.html", context)
###################### DETAIL URGENT ICD CMA #####################################
from datetime import timedelta
from django.utils import timezone

def detail_urgent_dk_log1(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=60)  # Délai par défaut
    delais_urg = timedelta(minutes=30)  # Délai urgent par défaut
    # Récupérer les mouvements en cours (date_sortie est null)
    mouvements = Mouvement4.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            # Récupérer les informations du camion associé au mouvement
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
            type = camion.type
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"
            type = None
        # Récupérer les délais spécifiques pour ce type de camion
        try:
            if type:
                para = ParametrageDelais.objects.get(entite='icd', type=type)
                duree_dk = timedelta(minutes=int(para.delais_maximal))
                delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            pass  # Utiliser les valeurs par défaut

        # Récupérer les informations du pointeur d'entrée
        try:
            poineur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = poineur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"

        # Récupérer les informations du chauffeur
        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"
        # Calculer la durée passée depuis l'entrée
        temps_dans_zone = maintenant - mvt.date_entree
        # Filtrer les mouvements en attente urgente
        if delais_urg <= temps_dans_zone < duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul du pourcentage de mouvements urgents
    total_encours = len(mouvements)
    lg = len(mouvements_filtrés)
    pourcentage_urgent = (lg / total_encours) * 100 if total_encours >= 1 else 0
    # Récupérer les informations de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_urgent': int(pourcentage_urgent),
        'util': user
    }
    # Rendu du template
    return render(request, "pages/detail_urgent_dk_log.html", context)
###################### DETAIL URGENT SACHERIE #####################################
def detail_urgent_dk_log2(request, id):
    maintenant = timezone.now()
    duree_dk = 60
    delais_urg = 30
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False, date_sortie__isnull=True, destination__contains='hangar')
    mouvements_avec_attributs = []
    mouvements_filtrés=[]
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='sacherie', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk :
            mouvements_filtrés.append(mvt)
        lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_urgent = (lg / total_encours) * 100
    else:
        pourcentage_urgent = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_urgent': int(pourcentage_urgent),
        'util': user
    }
    return render(request, "pages/detail_urgent_dk_log0.html", context)
###################### DETAIL URGENT ZUD #####################################
def detail_urgent_dk_log3(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_avec_attributs = []
    mouvements_filtrés=[]

    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk :
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_urgent = (lg / total_encours) * 100
    else:
        pourcentage_urgent = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_urgent': int(pourcentage_urgent),
        'util': user
    }
    return render(request, "pages/detail_urgent_dk_log.html", context)
####################### DETAIL DEPASSEMENT PLT ############################
def detail_depassement_dk_logpar(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=480)  # Par défaut 30 minutes pour le délai maximal
    # Récupération des mouvements non sortis
    mouvements = Mouvement8.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            # Récupérer les informations du véhicule associé au mouvement
            camion = Vehicule.objects.get(id_veh=mvt.vehicule_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
        except Vehicule.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.imat = "Non assigné"
        try:
            # Récupérer le pointeur d'entrée
            pointeur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = pointeur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"
        # Vérifier si la durée écoulée dépasse le délai maximal (dépassement)
        if (maintenant - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul du nombre total de mouvements et du pourcentage de dépassement
    total_encours = len(mouvements)
    lg = len(mouvements_filtrés)
    pourcentage_depassement = (lg / total_encours) * 100 if total_encours >= 1 else 0
    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_depassement': int(pourcentage_depassement),
        'util': user
    }
    # Rendu du template avec le contexte
    return render(request, "pages/detail_depasseme_dk_logpar.html", context)
####################### DETAIL DEPASSEMENT PLT ############################
def detail_depassement_dk_log0(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []

    for mvt in mouvements:
        try:
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            type = camion.type or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            type = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"

        try:
            para = ParametrageDelais.objects.get(entite='dklog', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            duree_dk = timedelta(minutes=10080)
            delais_urg = timedelta(minutes=7200)
        try:
            pointeur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = pointeur.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"
        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"
        # Filtrage des mouvements selon la durée dépassée
        if (maintenant - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul des statistiques
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    pourcentage_depassement = (lg / total_encours) * 100 if total_encours >= 1 else 0
    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_depassement': int(pourcentage_depassement),
        'util': user
    }
    return render(request, "pages/detail_depasseme_dk_log0.html", context)
####################### DETAIL DEPASSEMENT ICD TOM ############################
def detail_depassement_dk_log(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='icd', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        try:
            chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        except Chaffeur.DoesNotExist:
            chauffeur_name = {'id_chauffeur': None, 'fullname': 'Non Assigné', 'permis': 'Non Assigné',
                              'telephone': 'Non Assigné'}
        # Vérifiez si chauffeur_name est un dictionnaire ou un objet Chauffeur
        mvt.chauffeur_name = chauffeur_name['fullname'] if isinstance(chauffeur_name, dict) else chauffeur_name.fullname
        if (maintenant - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_depassement = (lg / total_encours) * 100
    else:
        pourcentage_depassement = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'pourcentage_depassement' : int(pourcentage_depassement),
        'util': user
    }
    return render(request, "pages/detail_depasseme_dk_log.html", context)
####################### DETAIL DEPASSEMENT ICD CMA ############################
def detail_depassement_dk_log1(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=60)  # Valeur par défaut pour le délai maximal
    delais_urg = timedelta(minutes=30)  # Valeur par défaut pour le délai urgent
    # Récupérer les mouvements en cours (date_sortie est null)
    mouvements = Mouvement4.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            # Récupérer les informations du camion
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
            type = camion.type
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"
            type = None
        # Récupérer les délais spécifiques en fonction du type de camion
        try:
            if type:
                para = ParametrageDelais.objects.get(entite='icd', type=type)
                duree_dk = timedelta(minutes=int(para.delais_maximal))
                delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            pass  # Utiliser les valeurs par défaut

        # Récupérer les informations du pointeur d'entrée
        try:
            poineur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = poineur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"

        # Récupérer les informations du chauffeur
        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"

        # Vérifier si le mouvement dépasse la durée maximale autorisée
        if (maintenant - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)

    # Calcul du nombre de mouvements en dépassement et du pourcentage
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    pourcentage_depassement = (lg / total_encours) * 100 if total_encours >= 1 else 0
    # Récupérer les informations de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte pour le rendu du template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_depassement': int(pourcentage_depassement),
        'util': user
    }
    # Rendu du template
    return render(request, "pages/detail_depasseme_dk_log.html", context)
####################### DETAIL DEPASSEMENT SACHERIE ############################
def detail_depassement_dk_log2(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False, date_sortie__isnull=True, destination__contains='hangar')
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='sacherie', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (maintenant - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_depassement = (lg / total_encours) * 100
    else:
        pourcentage_depassement = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_depassement' : int(pourcentage_depassement),
        'util': user
    }
    return render(request, "pages/detail_depasseme_dk_log0.html", context)
####################### DETAIL DEPASSEMENT ZUD ############################
def detail_depassement_dk_log3(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (maintenant - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_depassement = (lg / total_encours) * 100
    else:
        pourcentage_depassement = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_depassement' : int(pourcentage_depassement),
        'util': user
    }
    return render(request, "pages/detail_depasseme_dk_log.html", context)
######################### DETAILS PLUS DE 30 MINUTES PLT  #######################
def detail_plus_30_dk_log0(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=30)
    delais_urg = timedelta(minutes=15)
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            type = camion.type or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            type = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"
        try:
            para = ParametrageDelais.objects.get(entite='dklog', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            duree_dk = timedelta(minutes=10080)
            delais_urg = timedelta(minutes=7200)
        try:
            pointeur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = pointeur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"
        try:
            pointeur_sortie = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
            mvt.pointeur_srt = pointeur_sortie.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur_srt = "Non assigné"
        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"
        # Filtrer les mouvements dont la durée entre l'entrée et la sortie dépasse duree_dk
        if (mvt.date_sortie - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul des statistiques
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    pourcentage_plus_30 = (lg / total_plus_30) * 100 if total_plus_30 >= 1 else 0
    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_plus_30': int(pourcentage_plus_30),
        'util': user
    }
    # Rendu du template avec le contexte
    return render(request, "pages/detail_plus_30_dk_log0.html", context)
######################### DETAILS PLUS DE 30 MINUTES ICD TOM  #######################
def detail_plus_30_dk_log(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=60)  # Par défaut, 30 minutes pour le délai maximal
    delais_urg = timedelta(minutes=30)  # Par défaut, 15 minutes pour le délai urgent

    # Récupération des mouvements avec date d'entrée et de sortie
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            # Récupérer les informations du camion associé au mouvement
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            type = camion.type
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.imat = "Non assigné"
            mvt.trans = "Non assigné"
            type = None

        # Récupérer les délais spécifiques, sinon utiliser les valeurs par défaut
        try:
            if type:
                para = ParametrageDelais.objects.get(entite='icd', type=type)
                duree_dk = timedelta(minutes=int(para.delais_maximal))
                delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            pass  # Garder les valeurs par défaut
        # Récupérer le pointeur d'entrée et de sortie
        try:
            poineur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = poineur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"
        try:
            poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
            mvt.pointeur_srt = poineur_srt.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur_srt = "Non assigné"
        # Récupérer le chauffeur
        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"
        # Ajouter les mouvements dont la durée dépasse duree_dk
        if (mvt.date_sortie - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul du nombre total de mouvements et du pourcentage de dépassement des 30 minutes
    total_plus_30 = len(mouvements)
    lg = len(mouvements_filtrés)
    pourcentage_plus_30 = (lg / total_plus_30) * 100 if total_plus_30 >= 1 else 0
    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'util': user,
        'pourcentage_plus_30': int(pourcentage_plus_30)
    }
    # Rendu du template avec le contexte
    return render(request, "pages/detail_plus_30_dk_log.html", context)
######################### DETAILS PLUS DE 30 MINUTES ICD CMA  #######################
def detail_plus_30_dk_log1(request, id):
    maintenant = timezone.now()
    duree_dk_default = timedelta(minutes=60)  # Valeur par défaut pour le délai maximal
    delais_urg_default = timedelta(minutes=30)  # Valeur par défaut pour le délai urgent
    # Récupérer les mouvements avec dates d'entrée et de sortie définies
    mouvements = Mouvement4.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            # Récupérer les informations du camion
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
            type = camion.type
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"
            type = None

        try:
            # Récupérer les paramètres de délai spécifiques
            if type:
                para = ParametrageDelais.objects.get(entite='icd', type=type)
                duree_dk = timedelta(minutes=int(para.delais_maximal))
                delais_urg = timedelta(minutes=int(para.delais_urgent))
            else:
                duree_dk = duree_dk_default
                delais_urg = delais_urg_default
        except ParametrageDelais.DoesNotExist:
            duree_dk = duree_dk_default
            delais_urg = delais_urg_default

        try:
            # Récupérer les informations du pointeur d'entrée
            poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = poineur.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"

        try:
            # Récupérer les informations du pointeur de sortie
            poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
            mvt.pointeur_srt = poineur_srt.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur_srt = "Non assigné"
        try:
            # Récupérer les informations du chauffeur
            chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur_name.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"

        # Ajouter le mouvement à la liste s'il dépasse la durée maximale
        if (mvt.date_sortie - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)

    # Calcul du nombre de mouvements filtrés et du pourcentage
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    pourcentage_plus_30 = (lg / total_plus_30) * 100 if total_plus_30 >= 1 else 0

    # Récupérer les informations de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)

    # Contexte pour le rendu du template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'util': user,
        'pourcentage_plus_30': int(pourcentage_plus_30)
    }

    return render(request, "pages/detail_plus_30_dk_log.html", context)
######################### DETAILS PLUS DE 30 MINUTES SACHERIE  #######################
def detail_plus_30_dk_log2(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False,date_sortie__isnull=False, destination__contains='hangar')
    mouvements_filtrés = []

    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation

        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='sacherie', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (mvt.date_sortie - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    if total_plus_30 >= 1:
      pourcentage_plus_30 = (lg / total_plus_30) * 100
    else :
      pourcentage_plus_30=0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'util': user,
        'pourcentage_plus_30' : int( pourcentage_plus_30)
    }
    return render(request, "pages/detail_plus_30_dk_log0.html", context)
######################### DETAILS PLUS DE 30 MINUTES ZUD   #######################
def detail_plus_30_dk_log3(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False,date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (mvt.date_sortie - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    if total_plus_30 >= 1:
      pourcentage_plus_30 = (lg / total_plus_30) * 100
    else :
      pourcentage_plus_30=0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'util': user,
        'pourcentage_plus_30' : int( pourcentage_plus_30)
    }
    return render(request, "pages/detail_plus_30_dk_log.html", context)
###################### DETAILS MOINS DE 30 MINUTES ICD TOM ############################
def detail_moins_30_dk_logpar(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement8.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Vehicule.objects.get(id_veh=mvt.vehicule_id)
        mvt.cam = camion.immatriculation
        #type = camion.type
        try:
            para = 30
            duree_dk = timedelta(minutes=int(para))
            delais_urg = timedelta(minutes=int(para))
        except:
            duree_dk = timedelta(minutes=480)
            delais_urg = timedelta(minutes=20)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.imat = camion.immatriculation
        if (mvt.date_sortie - mvt.date_entree) <= duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    if  total_moins_30 >= 1 :
       pourcentage_moins_30 = (lg / total_moins_30) * 100
    else :
        pourcentage_moins_30=0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_moins_30' :int(pourcentage_moins_30),
        'util': user
    }
    #return render(request, "pages/detail_depassement.html", context)
    return render(request, "pages/detail_moins_30_dk_logpar.html", context)
######################### DETAILS PLUS DE 30 MINUTES PARTICULIER  #######################
def detail_plus_30_dk_logpar(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement8.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Vehicule.objects.get(id_veh=mvt.vehicule_id)
        mvt.cam = camion.immatriculation
        try:
            duree_dk = timedelta(minutes=int(480))
            delais_urg = timedelta(minutes=int(420))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.imat = camion.immatriculation
        if (mvt.date_sortie - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    if total_plus_30 >= 1:
        pourcentage_plus_30 = (lg / total_plus_30) * 100
    else:
        pourcentage_plus_30 = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'util': user,
        'pourcentage_plus_30': int(pourcentage_plus_30)
    }
    return render(request, "pages/detail_plus_30_dk_logpar.html", context)
###################### DETAILS MOINS DE 30 MINUTES ICD TOM ############################
def detail_moins_30_dk_log0(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=30)
    delais_urg = timedelta(minutes=15)
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            type = camion.type or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            type = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"
        try:
            para = ParametrageDelais.objects.get(entite='dklog', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            duree_dk = timedelta(minutes=10080)
            delais_urg = timedelta(minutes=7200)
        try:
            pointeur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = pointeur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"
        try:
            pointeur_sortie = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
            mvt.pointeur_srt = pointeur_sortie.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur_srt = "Non assigné"
        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"
        # Filtrer les mouvements dont la durée entre l'entrée et la sortie est inférieure ou égale à duree_dk
        if (mvt.date_sortie - mvt.date_entree) <= duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul des statistiques
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    pourcentage_moins_30 = (lg / total_moins_30) * 100 if total_moins_30 >= 1 else 0
    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_moins_30': int(pourcentage_moins_30),
        'util': user
    }
    # Rendu du template avec le contexte
    return render(request, "pages/detail_moins_30_dk_log0.html", context)
###################### DETAILS MOINS DE 30 MINUTES ICD TOM ############################
def detail_moins_30_dk_log(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=60)  # Par défaut, 30 minutes pour le délai maximal
    delais_urg = timedelta(minutes=30)  # Par défaut, 15 minutes pour le délai urgent
    # Récupération des mouvements avec date d'entrée et de sortie
    mouvements = Mouvement1.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            # Récupérer les informations du camion associé au mouvement
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            type = camion.type
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.imat = "Non assigné"
            mvt.trans = "Non assigné"
            type = None
        # Récupérer les délais spécifiques, sinon utiliser les valeurs par défaut
        try:
            if type:
                para = ParametrageDelais.objects.get(entite='icd', type=type)
                duree_dk = timedelta(minutes=int(para.delais_maximal))
                delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            pass  # Garde les valeurs par défaut
        # Récupérer le pointeur d'entrée et de sortie
        try:
            poineur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = poineur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"
        try:
            poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
            mvt.pointeur_srt = poineur_srt.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur_srt = "Non assigné"
        # Récupérer le chauffeur
        try:
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"
        # Ajouter les mouvements qui respectent la condition (durée <= duree_dk)
        if (mvt.date_sortie - mvt.date_entree) <= duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul du nombre total de mouvements et du pourcentage de moins de 30 minutes
    total_moins_30 = len(mouvements)
    lg = len(mouvements_filtrés)
    pourcentage_moins_30 = (lg / total_moins_30) * 100 if total_moins_30 >= 1 else 0
    # Récupération de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte à passer au template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_moins_30': int(pourcentage_moins_30),
        'util': user
    }
    # Rendu du template avec le contexte
    return render(request, "pages/detail_moins_30_dk_log.html", context)
###################### DETAILS MOINS DE 30 MINUTES ICD CMA ############################
def detail_moins_30_dk_log1(request, id):
    maintenant = timezone.now()
    duree_dk = timedelta(minutes=60)  # Valeur par défaut pour le délai maximal
    delais_urg = timedelta(minutes=30)  # Valeur par défaut pour le délai urgent
    # Récupérer les mouvements avec dates d'entrée et de sortie définies
    mouvements = Mouvement4.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        try:
            # Récupérer les informations du camion
            camion = Camion.objects.get(id_cam=mvt.camion_id)
            mvt.cam = camion.immatriculation or "Non assigné"
            mvt.trans = camion.transporteur or "Non assigné"
            mvt.imat = camion.immatriculation or "Non assigné"
            type = camion.type
        except Camion.DoesNotExist:
            mvt.cam = "Non assigné"
            mvt.trans = "Non assigné"
            mvt.imat = "Non assigné"
            type = None

        try:
            # Récupérer les paramètres de délai spécifiques
            if type:
                para = ParametrageDelais.objects.get(entite='icd', type=type)
                duree_dk = timedelta(minutes=int(para.delais_maximal))
                delais_urg = timedelta(minutes=int(para.delais_urgent))
        except ParametrageDelais.DoesNotExist:
            pass  # Utiliser les valeurs par défaut
        try:
            # Récupérer les informations du pointeur d'entrée
            poineur_entree = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
            mvt.pointeur = poineur_entree.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur = "Non assigné"
        try:
            # Récupérer les informations du pointeur de sortie
            poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
            mvt.pointeur_srt = poineur_srt.fullname or "Non assigné"
        except Utilisateurs.DoesNotExist:
            mvt.pointeur_srt = "Non assigné"
        try:
            # Récupérer les informations du chauffeur
            chauffeur = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
            mvt.chauffeur_name = chauffeur.fullname or "Non assigné"
        except Chaffeur.DoesNotExist:
            mvt.chauffeur_name = "Non assigné"
        # Ajouter le mouvement à la liste s'il est dans le délai maximal
        if (mvt.date_sortie - mvt.date_entree) <= duree_dk:
            mouvements_filtrés.append(mvt)
    # Calcul du nombre de mouvements filtrés et du pourcentage
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    pourcentage_moins_30 = (lg / total_moins_30) * 100 if total_moins_30 >= 1 else 0
    # Récupérer les informations de l'utilisateur
    user = Utilisateurs.objects.get(id_user=id)
    # Contexte pour le rendu du template
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_moins_30': int(pourcentage_moins_30),
        'util': user
    }
    return render(request, "pages/detail_moins_30_dk_log.html", context)
###################### DETAILS MOINS DE 30 MINUTES SACHERIE ############################
def detail_moins_30_dk_log2(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement0.objects.filter(date_entree__isnull=False, date_sortie__isnull=False, destination__contains='hangar')
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='sacherie', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (mvt.date_sortie - mvt.date_entree) <= duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    if  total_moins_30 >= 1 :
       pourcentage_moins_30 = (lg / total_moins_30) * 100
    else :
        pourcentage_moins_30=0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_moins_30' :int(pourcentage_moins_30),
        'util': user
    }
    return render(request, "pages/detail_moins_30_dk_log0.html", context)
###################### DETAILS MOINS DE 30 MINUTES ZUD  ############################
def detail_moins_30_dk_log3(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=60)
            delais_urg = timedelta(minutes=30)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname

        if (mvt.date_sortie - mvt.date_entree) <= duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    if  total_moins_30 >= 1 :
       pourcentage_moins_30 = (lg / total_moins_30) * 100
    else :
        pourcentage_moins_30=0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_moins_30' :int(pourcentage_moins_30),
        'util': user
    }
    #return render(request, "pages/detail_depassement.html", context)
    return render(request, "pages/detail_moins_30_dk_log.html", context)
#####################Extraction
def export_camions(request):
  if request.method == 'POST':
    # Access camion data from your model (replace with your logic)
    camions = Camion.objects.all()
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Camions"
    # Write table header
    ws.append(["IMMATRICULATION", "Transporteur", "Type"])

    # Write camion data to rows
    for camion in camions:
      ws.append([camion.immatriculation, camion.transporteur, camion.type])

    # Create a response object and set content type to Excel
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=camions.xlsx'

    # Write the workbook to the response
    wb.save(response)
    return response

  else:
    # Handle invalid request method (optional)
    return HttpResponseBadRequest()


from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
######################## RAPPORT DKLOG ###############"
def export_mouvement0(request):
    if request.method == 'POST':
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        # Récupérer les données des mouvements
        mouvements = Mouvement0.objects.filter(date_entree__range=[date_debut, date_fin]).values(
            'id_mvt', 'chauffeur_id', 'camion_id', 'statut_entree', 'statut_sortie', 'zone_entree', 'zone_sortie',
            'remorque', 'date_entree', 'pointeur_entree_id', 'date_sortie', 'pointeur_sortie_id'
        )

        mouvement_list = list(mouvements)

        for mouvement in mouvement_list:
            # Traduire le statut d'entrée
            statut_entree = mouvement.get('statut_entree', None)
            if statut_entree == 0:
                mouvement['statut_entree_label'] = 'Vide'
            elif statut_entree == 1:
                mouvement['statut_entree_label'] = 'Partielle'
            elif statut_entree == 2:
                mouvement['statut_entree_label'] = 'Plein'
            else:
                mouvement['statut_entree_label'] = 'Inconnu'

            # Traiter les autres champs
            camion_id = mouvement.get('camion_id', None)
            if camion_id:
                camion = Camion.objects.get(id_cam=camion_id)
                mouvement['transporteur'] = camion.transporteur or 'Non Assigné'
                mouvement['type'] = camion.type or 'Non Assigné'
                mouvement['immatriculation'] = camion.immatriculation or 'Non Assigné'
            else:
                mouvement['transporteur'] = 'Non Assigné'
                mouvement['type'] = 'Non Assigné'
                mouvement['immatriculation'] = 'Non Assigné'

            chauffeur_id = mouvement.get('chauffeur_id', None)
            if chauffeur_id:
                chauffeur = Chaffeur.objects.get(id_chauffeur=chauffeur_id)
                mouvement['chauffeur_name'] = chauffeur.fullname or 'Non Assigné'
                mouvement['permis'] = chauffeur.permis or 'Non Assigné'
            else:
                mouvement['chauffeur_name'] = 'Non Assigné'
                mouvement['permis'] = 'Non Assigné'

            pointeur_entree_id = mouvement.get('pointeur_entree_id', None)
            if pointeur_entree_id:
                pointeur_entree = Utilisateurs.objects.get(id_user=pointeur_entree_id)
                mouvement['pointeur_entree_name'] = pointeur_entree.fullname or 'Non Assigné'
            else:
                mouvement['pointeur_entree_name'] = 'Non Assigné'

            pointeur_sortie_id = mouvement.get('pointeur_sortie_id', None)
            if pointeur_sortie_id:
                pointeur_sortie = Utilisateurs.objects.get(id_user=pointeur_sortie_id)
                mouvement['pointeur_sortie_name'] = pointeur_sortie.fullname or 'Non Assigné'
            else:
                mouvement['pointeur_sortie_name'] = 'Non Assigné'

            # Assurer que les zones sont non nulles
            mouvement['zone_entree'] = mouvement.get('zone_entree', 'Non Assigné')
            mouvement['zone_sortie'] = mouvement.get('zone_sortie', 'Non Assigné')
            mouvement['remorque'] = mouvement.get('remorque', 'Non Assigné')
            mouvement['date_entree'] = str(mouvement.get('date_entree', 'Non Assigné'))
            mouvement['date_sortie'] = str(mouvement.get('date_sortie', 'Non Assigné'))

        # Créer un nouveau workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Camions"

        # Titre du rapport en grandes lettres
        ws.merge_cells('A1:L1')
        ws['A1'] = "RAPPORT DKLOG"
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Tête du tableau
        headers = [
            "Zone Entree", "Zone Sortie", "Statut Entree", "Camion", "Transporteur", "Type", "Remorque",
            "Chauffeur", "Permis", "Date Entrée", "Pointeur Entrée", "Date Sortie", "Pointeur Sortie"
        ]
        ws.append(headers)

        # Appliquer du style aux en-têtes
        header_font = Font(bold=True, size=12)
        header_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}2"].font = header_font
            ws[f"{col_letter}2"].border = header_border
            ws[f"{col_letter}2"].alignment = Alignment(horizontal='center')

        # Ajouter les données dans les lignes suivantes
        data_font = Font(size=11)
        data_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for row_idx, mouvement in enumerate(mouvement_list, start=3):
            ws.append([
                mouvement['zone_entree'], mouvement['zone_sortie'], mouvement['statut_entree_label'],
                mouvement['immatriculation'], mouvement['transporteur'], mouvement['type'],
                mouvement['remorque'], mouvement['chauffeur_name'], mouvement['permis'],
                mouvement['date_entree'], mouvement['pointeur_entree_name'], mouvement['date_sortie'],
                mouvement['pointeur_sortie_name']
            ])

            # Appliquer les styles de bordure et de police à chaque cellule de la ligne
            for col_num in range(1, 14):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.font = data_font
                cell.border = data_border
        # Ajuster la largeur des colonnes pour qu'elles soient plus lisibles
        for col_num, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        # Préparer la réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="RAPPORT_DKLOG.xlsx"'
        wb.save(response)
        return response
    else:
        return HttpResponseBadRequest()

#################### RAPPORT ICD TOM ######################
def export_mouvement1(request):
    if request.method == 'POST':
        # Récupérer les données des mouvements
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        mouvements = Mouvement1.objects.filter(
            date_entree__range=[date_debut, date_fin]
        ).values(
            'id_mvt', 'chauffeur_id', 'camion_id', 'statut_entree', 'statut_sortie',
            'zone_entree', 'zone_sortie', 'remorque', 'date_entree', 'pointeur_entree_id',
            'date_sortie', 'pointeur_sortie_id', 'mission'
        )

        mouvement_list = list(mouvements)
        for mouvement in mouvement_list:
            # Traitement des champs
            mouvement['statut_entree'] = (
                'NON CHARGE' if mouvement['statut_entree'] == 0 else
                'Partielle' if mouvement['statut_entree'] == 1 else
                'Plein' if mouvement['statut_entree'] == 2 else
                'Vide'
            )

            mouvement['statut_sortie'] = (
                'NON CHARGE' if mouvement['statut_sortie'] == 0 else
                'Partielle' if mouvement['statut_sortie'] == 1 else
                'Plein' if mouvement['statut_sortie'] == 2 else
                'Vide'
            )

            camion_id = mouvement.get('camion_id', None)
            mouvement['transporteur'] = (
                Camion.objects.get(id_cam=camion_id).transporteur if camion_id else 'Non Assigné'
            )

            mouvement['type'] = (
                Camion.objects.get(id_cam=camion_id).type if camion_id else 'Non Assigné'
            )

            mouvement['immatriculation'] = (
                Camion.objects.get(id_cam=camion_id).immatriculation if camion_id else 'Non Assigné'
            )

            chauffeur_id = mouvement.get('chauffeur_id', None)
            mouvement['chauffeur_name'] = (
                Chaffeur.objects.get(id_chauffeur=chauffeur_id).fullname if chauffeur_id else 'Non Assigné'
            )

            mouvement['permis'] = (
                Chaffeur.objects.get(id_chauffeur=chauffeur_id).permis if chauffeur_id else 'Non Assigné'
            )

            pointeur_entree_id = mouvement.get('pointeur_entree_id', None)
            mouvement['pointeur_entree_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_entree_id).fullname if pointeur_entree_id else 'Non Assigné'
            )

            pointeur_sortie_id = mouvement.get('pointeur_sortie_id', None)
            mouvement['pointeur_sortie_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_sortie_id).fullname if pointeur_sortie_id else 'Non Assigné'
            )

        # Créer un nouveau workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Camions"

        # Titre du rapport
        ws.merge_cells('A1:N1')
        ws['A1'] = "RAPPORT ICD TOM"
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Tête du tableau
        headers = [
            "Statut Entree", "Statut Sortie", "Zone Entree", "Zone Sortie", "Camion", "Transporteur", "Type",
            "Remorque", "Chauffeur", "Permis", "Date Entrée", "Pointeur Entrée", "Date Sortie", "Pointeur Sortie", "Mission"
        ]
        ws.append(headers)

        # Appliquer du style aux en-têtes
        header_font = Font(bold=True, size=12)
        header_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}2"].font = header_font
            ws[f"{col_letter}2"].border = header_border
            ws[f"{col_letter}2"].alignment = Alignment(horizontal='center')

        # Ajouter les données dans les lignes suivantes
        data_font = Font(size=11)
        data_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for row_idx, mouvement in enumerate(mouvement_list, start=3):
            ws.append([
                mouvement['statut_entree'], mouvement['statut_sortie'], mouvement['zone_entree'] or 'Non Assigné',
                mouvement['zone_sortie'] or 'Non Assigné', mouvement['immatriculation'],
                mouvement['transporteur'], mouvement['type'], mouvement['remorque'] or 'Non Assigné',
                mouvement['chauffeur_name'], mouvement['permis'], str(mouvement['date_entree']),
                mouvement['pointeur_entree_name'], str(mouvement['date_sortie']), mouvement['pointeur_sortie_name'], mouvement['mission']
            ])

            # Appliquer les styles de bordure et de police à chaque cellule de la ligne
            for col_num in range(1, 15):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.font = data_font
                cell.border = data_border

        # Ajuster la largeur des colonnes pour qu'elles soient plus lisibles
        for col_num, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Préparer la réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rapport_ICD_TOM.xlsx"'
        wb.save(response)
        return response
    else:
        return HttpResponseBadRequest()
######################## RAPPORT SACHERIE #######################
def export_mouvement2(request):
    if request.method == 'POST':
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')

        # Récupérer les données des mouvements avec filtrage par date
        mouvements = Mouvement2.objects.filter(
            date_entree__range=[date_debut, date_fin]
        ).values(
            'id_mvt', 'chauffeur_id', 'camion_id', 'statut_entree', 'statut_sortie',
            'zone_entree', 'zone_sortie', 'remorque', 'date_entree', 'pointeur_entree_id',
            'date_sortie', 'pointeur_sortie_id', 'mission'
        )

        mouvement_list = list(mouvements)
        for mouvement in mouvement_list:
            # Traitement des champs
            mouvement['statut_entree'] = (
                'NON CHARGE' if mouvement['statut_entree'] == 0 else
                'Partielle' if mouvement['statut_entree'] == 1 else
                'Plein' if mouvement['statut_entree'] == 2 else
                'Vide'
            )

            mouvement['statut_sortie'] = (
                'NON CHARGE' if mouvement['statut_sortie'] == 0 else
                'Partielle' if mouvement['statut_sortie'] == 1 else
                'Plein' if mouvement['statut_sortie'] == 2 else
                'Vide'
            )

            camion_id = mouvement.get('camion_id', None)
            mouvement['transporteur'] = (
                Camion.objects.get(id_cam=camion_id).transporteur if camion_id else 'Non Assigné'
            )

            mouvement['type'] = (
                Camion.objects.get(id_cam=camion_id).type if camion_id else 'Non Assigné'
            )

            mouvement['immatriculation'] = (
                Camion.objects.get(id_cam=camion_id).immatriculation if camion_id else 'Non Assigné'
            )

            chauffeur_id = mouvement.get('chauffeur_id', None)
            mouvement['chauffeur_name'] = (
                Chaffeur.objects.get(id_chauffeur=chauffeur_id).fullname if chauffeur_id else 'Non Assigné'
            )

            mouvement['permis'] = (
                Chaffeur.objects.get(id_chauffeur=chauffeur_id).permis if chauffeur_id else 'Non Assigné'
            )

            pointeur_entree_id = mouvement.get('pointeur_entree_id', None)
            mouvement['pointeur_entree_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_entree_id).fullname if pointeur_entree_id else 'Non Assigné'
            )

            pointeur_sortie_id = mouvement.get('pointeur_sortie_id', None)
            mouvement['pointeur_sortie_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_sortie_id).fullname if pointeur_sortie_id else 'Non Assigné'
            )

        # Créer un nouveau workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Camions"

        # Titre du rapport
        ws.merge_cells('A1:N1')
        ws['A1'] = "RAPPORT SACHERIE"
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Tête du tableau
        headers = [
            "Statut Entree", "Statut Sortie", "Zone Entree", "Zone Sortie", "Camion", "Transporteur", "Type",
            "Remorque", "Chauffeur", "Permis", "Date Entrée", "Pointeur Entrée", "Date Sortie", "Pointeur Sortie", "Mission"
        ]
        ws.append(headers)

        # Appliquer du style aux en-têtes
        header_font = Font(bold=True, size=12)
        header_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}2"].font = header_font
            ws[f"{col_letter}2"].border = header_border
            ws[f"{col_letter}2"].alignment = Alignment(horizontal='center')

        # Ajouter les données dans les lignes suivantes
        data_font = Font(size=11)
        data_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for row_idx, mouvement in enumerate(mouvement_list, start=3):
            ws.append([
                mouvement['statut_entree'], mouvement['statut_sortie'],
                mouvement['zone_entree'] or 'Non Assigné', mouvement['zone_sortie'] or 'Non Assigné',
                mouvement['immatriculation'], mouvement['transporteur'], mouvement['type'],
                mouvement['remorque'] or 'Non Assigné', mouvement['chauffeur_name'], mouvement['permis'],
                str(mouvement['date_entree']), mouvement['pointeur_entree_name'],
                str(mouvement['date_sortie']), mouvement['pointeur_sortie_name'], mouvement['mission']
            ])

            # Appliquer les styles de bordure et de police à chaque cellule de la ligne
            for col_num in range(1, 15):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.font = data_font
                cell.border = data_border

        # Ajuster la largeur des colonnes pour qu'elles soient plus lisibles
        for col_num, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        # Préparer la réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rapport_Sacherie.xlsx"'
        wb.save(response)
        return response
    else:
        return HttpResponseBadRequest()
####################### RAPPORT ZUD ##########################
def export_mouvement3(request):
    if request.method == 'POST':
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        # Récupérer les données des mouvements avec filtrage par date
        mouvements = Mouvement3.objects.filter(
            date_entree__range=[date_debut, date_fin]
        ).values(
            'id_mvt', 'chauffeur_id', 'camion_id', 'statut_entree', 'statut_sortie',
            'zone_entree', 'zone_sortie', 'remorque', 'date_entree', 'pointeur_entree_id',
            'date_sortie', 'pointeur_sortie_id', 'mission'
        )
        mouvement_list = list(mouvements)
        for mouvement in mouvement_list:
            # Traitement des champs
            mouvement['statut_entree'] = (
                'NON CHARGE' if mouvement['statut_entree'] == 0 else
                'Partielle' if mouvement['statut_entree'] == 1 else
                'Plein' if mouvement['statut_entree'] == 2 else
                'Vide'
            )
            mouvement['statut_sortie'] = (
                'NON CHARGE' if mouvement['statut_sortie'] == 0 else
                'Partielle' if mouvement['statut_sortie'] == 1 else
                'Plein' if mouvement['statut_sortie'] == 2 else
                'Vide'
            )
            camion_id = mouvement.get('camion_id', None)
            mouvement['transporteur'] = (
                Camion.objects.get(id_cam=camion_id).transporteur if camion_id else 'Non Assigné'
            )
            mouvement['type'] = (
                Camion.objects.get(id_cam=camion_id).type if camion_id else 'Non Assigné'
            )
            mouvement['immatriculation'] = (
                Camion.objects.get(id_cam=camion_id).immatriculation if camion_id else 'Non Assigné'
            )
            chauffeur_id = mouvement.get('chauffeur_id', None)
            mouvement['chauffeur_name'] = (
                Chaffeur.objects.get(id_chauffeur=chauffeur_id).fullname if chauffeur_id else 'Non Assigné'
            )
            mouvement['permis'] = (
                Chaffeur.objects.get(id_chauffeur=chauffeur_id).permis if chauffeur_id else 'Non Assigné'
            )
            pointeur_entree_id = mouvement.get('pointeur_entree_id', None)
            mouvement['pointeur_entree_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_entree_id).fullname if pointeur_entree_id else 'Non Assigné'
            )
            pointeur_sortie_id = mouvement.get('pointeur_sortie_id', None)
            mouvement['pointeur_sortie_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_sortie_id).fullname if pointeur_sortie_id else 'Non Assigné'
            )
        # Créer un nouveau workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Camions"

        # Titre du rapport
        ws.merge_cells('A1:N1')
        ws['A1'] = "RAPPORT ZUD"
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Tête du tableau
        headers = [
            "Statut Entree", "Statut Sortie", "Zone Entree", "Zone Sortie", "Camion", "Transporteur", "Type",
            "Remorque", "Chauffeur", "Permis", "Date Entrée", "Pointeur Entrée", "Date Sortie", "Pointeur Sortie", "Mission"
        ]
        ws.append(headers)

        # Appliquer du style aux en-têtes
        header_font = Font(bold=True, size=12)
        header_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}2"].font = header_font
            ws[f"{col_letter}2"].border = header_border
            ws[f"{col_letter}2"].alignment = Alignment(horizontal='center')

        # Ajouter les données dans les lignes suivantes
        data_font = Font(size=11)
        data_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for row_idx, mouvement in enumerate(mouvement_list, start=3):
            ws.append([
                mouvement['statut_entree'], mouvement['statut_sortie'],
                mouvement['zone_entree'] or 'Non Assigné', mouvement['zone_sortie'] or 'Non Assigné',
                mouvement['immatriculation'], mouvement['transporteur'], mouvement['type'],
                mouvement['remorque'] or 'Non Assigné', mouvement['chauffeur_name'], mouvement['permis'],
                str(mouvement['date_entree']), mouvement['pointeur_entree_name'],
                str(mouvement['date_sortie']), mouvement['pointeur_sortie_name'], mouvement['mission']
            ])

            # Appliquer les styles de bordure et de police à chaque cellule de la ligne
            for col_num in range(1, 15):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.font = data_font
                cell.border = data_border

        # Ajuster la largeur des colonnes pour qu'elles soient plus lisibles
        for col_num, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Préparer la réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rapport_Zud.xlsx"'
        wb.save(response)
        return response

    else:
        return HttpResponseBadRequest()
    ############## RAPPORT PARTICULIER ########################
def export_mouvementpar(request):
    if request.method == 'POST':
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        # Récupérer les données des mouvements avec filtrage par date
        mouvements = Mouvement8.objects.filter(
            date_entree__range=[date_debut, date_fin]
        ).values(
            'id_mvt',
            'zone_entree', 'zone_sortie', 'vehicule_id', 'date_entree', 'pointeur_entree_id',
            'date_sortie', 'pointeur_sortie_id'
        )
        mouvement_list = list(mouvements)
        for mouvement in mouvement_list:
            # Traitement des champs
            vehicule_id = mouvement.get('vehicule_id', None)
            mouvement['immatriculation'] = (
                Vehicule.objects.get(id_veh=vehicule_id).immatriculation if vehicule_id else 'Non Assigné'
            )
            pointeur_entree_id = mouvement.get('pointeur_entree_id', None)
            mouvement['pointeur_entree_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_entree_id).fullname if pointeur_entree_id else 'Non Assigné'
            )
            pointeur_sortie_id = mouvement.get('pointeur_sortie_id', None)
            mouvement['pointeur_sortie_name'] = (
                Utilisateurs.objects.get(id_user=pointeur_sortie_id).fullname if pointeur_sortie_id else 'Non Assigné'
            )
        # Créer un nouveau workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicules"
        # Titre du rapport
        ws.merge_cells('A1:N1')
        ws['A1'] = "RAPPORT MOUVEMENTS PARTICULIERS"
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        # Tête du tableau
        headers = [
             "Zone Entree", "Zone Sortie", "Vehicule", "Date Entrée", "Pointeur Entrée", "Date Sortie", "Pointeur Sortie"
        ]
        ws.append(headers)
        # Appliquer du style aux en-têtes
        header_font = Font(bold=True, size=12)
        header_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}2"].font = header_font
            ws[f"{col_letter}2"].border = header_border
            ws[f"{col_letter}2"].alignment = Alignment(horizontal='center')
        # Ajouter les données dans les lignes suivantes
        data_font = Font(size=11)
        data_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        for row_idx, mouvement in enumerate(mouvement_list, start=3):
            ws.append([
                mouvement['zone_entree'] or 'Non Assigné', mouvement['zone_sortie'] or 'Non Assigné',
                mouvement['immatriculation'], str(mouvement['date_entree']), mouvement['pointeur_entree_name'],
                str(mouvement['date_sortie']), mouvement['pointeur_sortie_name']
            ])

            # Appliquer les styles de bordure et de police à chaque cellule de la ligne
            for col_num in range(1, 15):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.font = data_font
                cell.border = data_border
        # Ajuster la largeur des colonnes pour qu'elles soient plus lisibles
        for col_num, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        # Préparer la réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rapport_PARTICULIER.xlsx"'
        wb.save(response)
        return response
    else:
        return HttpResponseBadRequest()
def liste_parametrage(request, id):
    user = get_object_or_404(Utilisateurs, id_user=id)
    listeboite = ParametrageDelais.objects.all()
    for boite in listeboite:
        delais_maximal = int(boite.delais_maximal) if boite.delais_maximal else 0
        delais_urgent = int(boite.delais_urgent) if boite.delais_urgent else 0
        boite.delais_maximal_hours =delais_maximal // 60
        boite.delais_maximal_minutes =delais_maximal % 60
        boite.delais_urgent_hours = delais_urgent // 60
        boite.delais_urgent_minutes =delais_urgent % 60
    return render(request, 'pages/parametrage1.html', {'boite': listeboite, 'util': user})


######################## AJOUT PARAMETRE #######################################
def ajouter_parametre(request):
    if request.method == 'POST':
        id_user = request.POST.get('id_user')
        util = get_object_or_404(Utilisateurs, id_user=id_user)
        h_urg = int(request.POST.get('h_urg'))  # Conversion en entier avec valeur par défaut 0
        m_urg = int(request.POST.get('m_urg'))  # Conversion en entier avec valeur par défaut 0
        h_max = int(request.POST.get('h_max'))  # Conversion en entier avec valeur par défaut 0
        m_max = int(request.POST.get('m_max'))  # Conversion en entier avec valeur par défaut 0
        entite = request.POST.get('entite')

        if entite is None:
            return HttpResponse("Invalid user role", status=400)

        parametre, created = ParametrageDelais.objects.get_or_create(
            type=request.POST.get('type'),
            entite=entite
        )

        # Calcul en minutes
        parametre.delais_urgent = h_urg * 60 + m_urg
        parametre.delais_maximal = h_max * 60 + m_max
        parametre.type = request.POST.get('type')
        parametre.nbr_max = request.POST.get('nbr_max')
        parametre.entite = entite
        parametre.save()

        return redirect(f"/parametrage/{id_user}?success=true")
    else:
        return redirect(f"/doc/ajouterboite_page/{request.POST.get('id_user')}")

def modifier_parametre(request):
    if request.method == 'POST':
        boite_id = request.POST['boite_id']
        id_user = request.POST.get('id_user')
        parametre = get_object_or_404(ParametrageDelais, id_para=boite_id)
        parametre.delais_urgent = request.POST.get('delais_urgent')
        parametre.delais_maximal = request.POST.get('delais_maximal')
        parametre.type = request.POST.get('type')
        parametre.nbr_max = request.POST.get('nbr_max')
        parametre.save()
        return redirect(f"/parametrage/{id_user}")

    return render(request, 'index.html')

################################  GESTIONNAIRE DE ZUD ##########################
 ##################### ENTREE #######################
def index_entree_zud(request, id):
  util = Utilisateurs.objects.get(id_user=id)
  camions = Camion.objects.all()
  return render(request, 'pages/index_entree_zud.html', {'util': util, 'camions': camions})
 ##################### SORTIE #######################
def index_sortie_zud(request, id):
  util = Utilisateurs.objects.get(id_user=id)
  camions = Camion.objects.all()
  return render(request, 'pages/index_sortie_zud.html', {'util': util, 'camions': camions})

### LISTE DES MOUVEMENTS  D'entree############
def liste_mouvements_entree_zud(request, id_user):
    mouvements = Mouvement3.objects.filter(date_entree__isnull=True, date_sortie__isnull=True).values('id_mvt', 'mission', 'camion_id',
                                                                            'statut_entree', 'statut_sortie',
                                                                            'chauffeur_id'
                                                                            , 'remorque')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        mouvement['camion'] = camion
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)

########################## LISTE DES MOUVEMENTS DE SORTIE ###################
def liste_mouvements_sortie_zud(request, id_user):
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=True).values('id_mvt',
                                                                                                       'mission',
                                                                                                       'camion_id',
                                                                                                       'statut_entree',
                                                                                                       'statut_sortie',
                                                                                             'chauffeur_id'
                                                                                                       , 'remorque')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        mouvement['camion'] = camion
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    response_data = {
        'mouvements': mouvement_list,
        'id_usr': id_user,  # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)
####################### ENTREE CAMION ZUD #####################################
def entree_zud(request):
    if request.method == 'POST':
        id_mvt = request.POST.get('id_mvt')
        id_user = request.POST.get('id_user')
        # Récupérer les objets Utilisateurs et Mouvement en fonction des IDs fournis
        util = get_object_or_404(Utilisateurs, id_user=id_user)
        mouvement = get_object_or_404(Mouvement3, id_mvt=id_mvt)
        mouvement.pointeur_entree_id = util.id_user
        mouvement.date_entree = timezone.now()
        mouvement.mission = request.POST.get('mission')
        mouvement.save()


    # Rediriger vers une vue après sauvegarde
        return redirect(f'/index_entree_zud/{util.id_user}')
    else:
        form = SortieForm()

    # Assurez-vous que 'mouvement' et 'util' sont définis si vous les utilisez dans le template


################################ SORTIE ZUD ####################################
def sortie_zud(request):
    if request.method == 'POST':
        form = SortieForm(request.POST)
        if form.is_valid():
            id_mvt = request.POST.get('id_mvt')
            id_user = request.POST.get('id_user')

            # Récupérer les objets Utilisateurs et Mouvement en fonction des IDs fournis
            util = get_object_or_404(Utilisateurs, id_user=id_user)
            mouvement = get_object_or_404(Mouvement3, id_mvt=id_mvt)

            # Mettre à jour les informations du mouvement
            mouvement.date_sortie = timezone.now()
            mouvement.pointeur_sortie_id = util.id_user
            mouvement.statut_sortie = form.cleaned_data.get(
                'statut_sortie')  # Récupérer le statut depuis les données validées du formulaire
            mouvement.save()

            # Rediriger vers une vue après sauvegarde
            #return redirect(f'/index2_view/{util.id_user}')
            return  redirect("/index_sortie_zud/" + str(util.id_user))

    else:
        form = SortieForm()
    # Assurez-vous que 'mouvement' et 'util' sont définis si vous les utilisez dans le template
    return render(request, 'pages/ajoutsortiedp.html', {'form': form})
###############Gestion zud
def index_admin_zud(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    ###############Calcul des variable
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False,date_sortie__isnull=True) | Mouvement3.objects.filter(
        date_entree__isnull=False,date_sortie__isnull=False)
    urg = 0
    dep = 0
    lg_30 = 0
    lg_moins = 0
    total_cours = 0
    mouvements_avec_attributs = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)

        diff = (maintenant - mvt.date_entree).total_seconds() / 60
        mvt.chrono = int(diff)
        duree = float(diff)
        if timedelta(minutes=duree) >= duree_dk:
            mvt.etat = 3
        elif timedelta(minutes=duree) < duree_dk and timedelta(minutes=duree) >= delais_urg:
            mvt.etat = 2
        else:
            mvt.etat = 1
        mouvements_avec_attributs.append(mvt)
        if mvt.date_sortie is None:
            # Mouvements en cours
            total_cours += 1
            if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk:
                urg += 1
            elif (maintenant - mvt.date_entree) >= duree_dk:
                dep += 1
        else:
            # Mouvements sortis
            duree_mouvement = mvt.date_sortie - mvt.date_entree
            if duree_mouvement >= duree_dk:
                lg_30 += 1
            else:
                lg_moins += 1

    totalter = lg_30 + lg_moins
    mouvements_filtrés = [mvt for mvt in mouvements_avec_attributs if mvt.date_sortie is None][:50]
    user = Utilisateurs.objects.get(id_user=id)
    return render(request, 'pages/index_admin_zud.html',
                  {'urg': urg, 'dep': dep, 'lg_30': lg_30, 'lg_mois': lg_moins,
                   'total_cours': total_cours, 'total_ter': totalter, 'mvt': mouvements_filtrés, 'util': user})
#util = Utilisateurs.objects.get(id_user=id)
  #camions = Camion.objects.all()
  #return render(request, 'pages/index_entree_zud.html', {'util': util, 'camions': camions})
def detail_urgent_zud(request,id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_avec_attributs = []
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk:
            mouvements_filtrés.append(mvt)
        # Ajoutez l'objet mouvement enrichi à la liste
        # mouvements_avec_attributs.append(mvt)
    # Appliquez le filtre après avoir enrichi les objets mouvement
    # mouvements_filtrés = [mvt for mvt in mouvements_avec_attributs if (maintenant - mvt.date_entree) >= delais_urg and (maintenant - mvt.date_entree) < duree_dk]
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_urgent = (lg / total_encours) * 100
    else:
        pourcentage_urgent = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        'lg': lg,
        'pourcentage_urgent': int(pourcentage_urgent),
        'util': user
    }
    return render(request, "pages/detail_urgent_zud.html", context)
def detail_depassement_zud(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=True)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)

        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname

        if (maintenant - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_encours = len(mouvements)
    if total_encours >= 1:
        pourcentage_depassement = (lg / total_encours) * 100
    else:
        pourcentage_depassement = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'pourcentage_depassement': int(pourcentage_depassement),
        'util': user
    }
    return render(request, "pages/detail_depassement_zud.html", context)
def detail_plus_30_zud(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (mvt.date_sortie - mvt.date_entree) > duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    total_plus_30 = len(mouvements)
    if total_plus_30 >= 1:
        pourcentage_plus_30 = (lg / total_plus_30) * 100
    else:
        pourcentage_plus_30 = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'util': user,
        'pourcentage_plus_30': int(pourcentage_plus_30)
    }
    return render(request, "pages/detail_plus_30_zud.html", context)
def detail_moins_30_zud(request, id):
    maintenant = timezone.now()
    duree_dk = 30
    delais_urg = 15
    mouvements = Mouvement3.objects.filter(date_entree__isnull=False, date_sortie__isnull=False)
    mouvements_filtrés = []
    for mvt in mouvements:
        camion = Camion.objects.get(id_cam=mvt.camion_id)
        mvt.cam = camion.immatriculation
        type = camion.type
        try:
            para = ParametrageDelais.objects.get(entite='zud', type=type)
            duree_dk = timedelta(minutes=int(para.delais_maximal))
            delais_urg = timedelta(minutes=int(para.delais_urgent))
        except:
            duree_dk = timedelta(minutes=30)
            delais_urg = timedelta(minutes=20)
        poineur = Utilisateurs.objects.get(id_user=mvt.pointeur_entree_id)
        mvt.pointeur = poineur.fullname
        poineur_srt = Utilisateurs.objects.get(id_user=mvt.pointeur_sortie_id)
        mvt.pointeur_srt = poineur_srt.fullname
        mvt.trans = camion.transporteur
        mvt.imat = camion.immatriculation
        chauffeur_name = Chaffeur.objects.get(id_chauffeur=mvt.chauffeur_id)
        mvt.chauffeur_name = chauffeur_name.fullname
        if (mvt.date_sortie - mvt.date_entree) <= duree_dk:
            mouvements_filtrés.append(mvt)
    lg = len(mouvements_filtrés)
    lg = len(mouvements_filtrés)
    total_moins_30 = len(mouvements)
    if total_moins_30 >= 1:
        pourcentage_moins_30 = (lg / total_moins_30) * 100
    else:
        pourcentage_moins_30 = 0
    user = Utilisateurs.objects.get(id_user=id)
    context = {
        'segment': 'index',
        'mvt': mouvements_filtrés,
        # 'products' : Product.objects.all()
        'lg': lg,
        'pourcentage_moins_30': int(pourcentage_moins_30),
        'util': user
    }
    #return render(request, "pages/detail_depassement_zud.html", context)
    return render(request, "pages/detail_moins_30_zud.html", context)
def tout_mouvement_zud(request, id_mvt):
    user=Utilisateurs.objects.get(id_user=id_mvt)
    return render(request, 'pages/liste_mouvement_admin_zud.html', {'util': user})
def liste_mouvements_zud(request):
    mouvements = Mouvement3.objects.all().values('id_mvt','mission','camion_id', 'statut_entree', 'statut_sortie','chauffeur_id', 'remorque','date_entree','date_sortie','pointeur_sortie_id','pointeur_entree_id')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        camion_id = mouvement['camion_id']
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur').first()
        mouvement['camion'] = camion
        user_entre=Utilisateurs.objects.filter(id_user=mouvement['pointeur_entree_id']).values('fullname').first()
        if mouvement['pointeur_sortie_id']  :
            user_sortie= Utilisateurs.objects.filter(id_user=mouvement['pointeur_sortie_id']).values('fullname').first()
        else:
            user_sortie = Utilisateurs.objects.filter(id_user=1).values('fullname').first()
            #for user_sortie in user_sortie :
                # Mettre l'attribut fullname vide
            user_sortie = {'fullname': 'original_name'}
            user_sortie['fullname'] = 'null'
            #user_sortie='oo'
        mouvement['user_ert'] = user_entre
        mouvement['user_srt'] = user_sortie
        chauffeur_id = mouvement['chauffeur_id']
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur
    return JsonResponse(list(mouvement_list), safe=False)
def export_mouvement4(request):
    if request.method == 'POST':
        # Récupérer les données des mouvements
        mouvements = Mouvement3.objects.filter(
            date_entree__isnull=False,
            date_sortie__isnull=False
        ).values(
            'id_mvt', 'chauffeur_id', 'camion_id', 'statut_entree', 'statut_sortie',
            'remorque', 'date_entree', 'pointeur_entree_id', 'date_sortie', 'pointeur_sortie_id','mission'
        )

        mouvement_list = list(mouvements)
        for mouvement in mouvement_list:
            ################# STATUT ENTREE TEXT ###########
            if mouvement['statut_entree']== 0:
                mouvement['statut_entree'] = 'NON CHARGE'
            elif mouvement['statut_entree'] == 1 :
                mouvement['statut_entree'] = 'NON CHARGE'
            else:
                mouvement['statut_entree'] = 'Vide'
                ################# STATUT SORTIE TEXT ###########
            if mouvement['statut_sortie'] == 0:
                 mouvement['statut_sortie'] = 'NON CHARGE'
            elif mouvement['statut_sortie'] == 1:
                 mouvement['statut_sortie'] = 'NON CHARGE'
            else:
                 mouvement['statut_sortie'] = 'Vide'
            ############## CAMION #####################
            camion_id = mouvement['camion_id']
            camion = Camion.objects.get(id_cam=camion_id)
            mouvement['transporteur'] = camion.transporteur
            mouvement['type'] = camion.type
            mouvement['immatriculation'] = camion.immatriculation
            ################### CHAUFFEUR ###################
            chauffeur_id = mouvement['chauffeur_id']
            chauffeur = Chaffeur.objects.get(id_chauffeur=chauffeur_id)
            mouvement['chauffeur_name'] = chauffeur.fullname
            mouvement['permis'] = chauffeur.permis
            ################## POINTEUR ENTREE ##############
            pointeur_entree_id = mouvement['pointeur_entree_id']
            Pointeur_entree = Utilisateurs.objects.get(id_user=pointeur_entree_id)
            mouvement['pointeur_entree_name'] = Pointeur_entree.fullname
            ################## POINTEUR SORTIE ##############
            pointeur_sortie_id = mouvement['pointeur_sortie_id']
            Pointeur_sortie = Utilisateurs.objects.get(id_user=pointeur_sortie_id)
            mouvement['pointeur_sortie_name'] = Pointeur_sortie.fullname
        # Créer un nouveau workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Camions"
        ws.append([
            "Statut Entree", "Statut Sortie", "Camion", "Transporteur", "Type", "Remorque",
            "Chauffeur", "Permis", "Date Entrée", "Pointeur Entrée", "Date Sortie", "Pointeur Sortie", "mission"
        ])
        for mouvement in mouvement_list:
            ws.append([
                mouvement['statut_entree'], mouvement['statut_sortie'], mouvement['immatriculation'],
                mouvement['transporteur'], mouvement['type'], mouvement['remorque'],
                mouvement['chauffeur_name'], mouvement['permis'], str(mouvement['date_entree']),
                mouvement['pointeur_entree_name'], str(mouvement['date_sortie']), mouvement['pointeur_sortie_name'], mouvement['mission']
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="mouvements.xlsx"'

        wb.save(response)
        return response

    else:
        return HttpResponseBadRequest()


@csrf_exempt
def modifier_mouvement4(request):
    if request.method == 'POST':
      boite_id = request.POST['mouvementId']
      boite = get_object_or_404(Mouvement3, id_mvt=boite_id)
      id_cam=boite.camion_id
      id_chauff=boite.chauffeur_id
      boite.mission = request.POST['mission']
      boite.remorque = request.POST['remorque']
      boite.statut_entree = request.POST['statut_entree']
      boite.statut_sortie = request.POST['statut_sortie']

      if  request.POST['camion'] :
          try :
              cam_mof=Camion.objects.filter(immatriculation= request.POST['camion']).first()
              boite.camion_id=cam_mof.id_cam
          except :
           pass
      if  request.POST['permis'] :
          try :
              chau_mof=Chaffeur.objects.filter(permis= request.POST['permis']).first()
              boite.chauffeur_id=chau_mof.id_chauffeur
          except :
           pass



      boite.save()
      #return redirect(reverse('index'))  # Redirigez vers la page index après la modification
      return redirect("/liste_mvt_zud/" + str(request.POST['id_user']))

    return render(request, 'index.html')





@csrf_exempt

def ajouter_transitaire(request):
    if request.method == 'POST':
        form = TransitaireForm(request.POST)
        if form.is_valid():
            chauffeur = form.save()
            # Retourner une réponse JSON indiquant le succès de l'enregistrement
            #return JsonResponse({'success': True, 'chauffeur': {'id': chauffeur.id_chauffeur, 'fullname': chauffeur.fullname}})
            return JsonResponse(
                {'success': True})
        else:
            # Retourner une réponse JSON avec des erreurs de formulaire
            return JsonResponse({'success': False, 'errors': form.errors})

    return JsonResponse({'success': False, 'errors': 'Méthode non autorisée'})







########################Gestion Chauffeur

########################################AGENT#######################################################
def liste_chauffeur(request, id):
    user = Utilisateurs.objects.get(id_user=id)
    listeboite=Chaffeur.objects.all()
    return render(request, 'pages/gestion_chauffeur.html', {'boite': listeboite,'util': user})
#####Gestion User

def Crer_agent_page(request, id):
   # forms = UtilisateurForm(request.POST)
   user = Utilisateurs.objects.get(id_utilisateur=id)
   # return render(request, 'templatetra/crer_agent.html', {'forms': forms, 'util': user})
   return render(request, 'templatetra/crer_agent.html', {'util': user})


def enregistrer_agent(request):
  if request.method == 'POST':
    # form = UtilisateuwrForm(request.POST)
    form = 1
    user = Utilisateurs.objects.get(id_utilisateur=request.POST["id_user"])
    if form.is_valid() or 1:
      cleaned_data = form.cleaned_data
      instance = Utilisateurs(
        nom=request.POST['nom'],
        prenom=request.POST['prenom'],
        email=request.POST['email'],
        password='passer',
        telephone=request.POST['telephone'],
        role=request.POST['rolee']

        # form.cleaned_data['direction']
      )
      instance.save()
      # return redirect('/users/login/')
      return redirect('/liste_user/{user.id_utilisateur}')
    else:
      # forms = UtilisateurForm()
      return Crer_agent_page(request, request.POST["id_user"])
      # return render(request, 'templatetra/crer_agent.html', {'forms': forms})


def update_agent_page(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  return render(request, 'templatetra/update_agent.html', {'agt': agt, 'util': user, 'user': agt})


def upadte_agent(request):
  agt = Utilisateurs.objects.get(id_utilisateur=request.POST['id_agt'])
  user = Utilisateurs.objects.get(id_utilisateur=request.POST['id_user'])
  if request.POST['prenom']:
    agt.prenom = request.POST['prenom']
  if request.POST['nom']:
    agt.email = request.POST['nom']
  if request.POST['telephone']:
    agt.telephone = request.POST['telephone'],

  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')
def desactiver_agent(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  agt.etat = 0
  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')
def activer_agent(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  agt.etat = 1
  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')

def reinitiliaser_mdp(request, id_user, id_agt):
  user = Utilisateurs.objects.get(id_utilisateur=id_user)
  agt = Utilisateurs.objects.get(id_utilisateur=id_agt)
  agt.password = 'reinit'
  agt.save()
  return redirect('/liste_user/{user.id_utilisateur}')
##############Import Fichier Camion
@csrf_exempt
def importer_donnees_camion_excel(fichier_excel):
    """
    Importe les données d'un fichier Excel dans la table Camion.

    Args:
        fichier_excel (UploadedFile): Le fichier Excel à importer.

    Raises:
        ValueError: Si le fichier Excel n'est pas valide.
    """

    if not fichier_excel.content_type.startswith('application/vnd.openxmlformats'):
        raise ValueError('Le fichier Excel n\'est pas valide.')

    # Lire le fichier Excel dans un DataFrame Pandas
    df = pd.read_excel(fichier_excel)

    # Vérifier si les colonnes du DataFrame correspondent aux champs du modèle Camion
    champs_camion = ['immatriculation', 'transporteur','type']
    if not df.columns.isin(champs_camion).all():
        raise ValueError('Le fichier Excel ne contient pas les colonnes requises.')

    # Enregistrer les données dans la base de données
    with transaction.atomic():
        for index, row in df.iterrows():
            camion = Camion(
                immatriculation=row['immatriculation'],
                transporteur=row['transporteur'],
                type=row['type'],
            )
            camion.save()


@csrf_exempt
def importer_donnees_camion(request):
    if request.method == 'POST':
        fichier_excel = request.FILES['fichier_excel']
        try:
            importer_donnees_camion_excel(fichier_excel)
            message = 'Données importées avec succès.'
        except ValueError as e:
            message = e.args[0]
    else:
        fichier_excel = None
        message = ''
    #user=Utilisateurs.objects.get(poste='admin')
    user = Utilisateurs.objects.get(id_user=1)
    return  redirect("/liste_amion/" +str(user.id_user))


def ajouter_camion_page(request):
  return render(request, 'pages/ajouter_camion.html',)
@csrf_exempt

def ajouter_chauffeur(request):
  if request.method == 'POST':
    #form = BoiteForm(request.POST)
    #id_user = request.session.get('user_id')
    if Chaffeur.objects.filter(permis=request.POST['permis']).exists():
      return redirect()
      # cleaned_data = form.cleaned_data
    else:
      camion = Chaffeur(
        permis=request.POST['permis'],
        telephone=request.POST['telephone'],
          fullname=request.POST['fullname'],
      )
      camion.save()

      return redirect("/liste_chauffeur/" + str(request.POST['id_user']))

  else:

    # return ajouterboite_page(request, request.POST['id_user'])
    return redirect("/doc/jouterboite_page/" + str(request.POST['id_user']))

def modifier_chauffeur(request):
    if request.method == 'POST':
      boite_id = request.POST['boite_id']
      boite = get_object_or_404(Chaffeur, id_chauffeutr=boite_id)
      boite.fullname = request.POST['fullname']
      boite.permis = request.POST['permis']
      boite.telephone = request.POST['telephone']
      boite.save()
      #return redirect(reverse('index'))  # Redirigez vers la page index après la modification
      return redirect("/liste_chauffeur/" + str(request.POST['id_user']))

    return render(request, 'index.html')
#################### FETCH LISTE DES MOVUEMENTS MODIFICATIONS ###############
########## ICD ################
def liste_modifs(request):
    mouvements = Mouvement0.objects.filter(destination__icontains='icd', date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination', 'mission', 'client_id'
    )
    mouvement_list = list(mouvements)

    for mouvement in mouvement_list:
        client_id = mouvement.get('client_id')
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        client = Client.objects.filter(id_client=client_id).values('id_client', 'fullname', 'telephone').first()
        mouvement['client']=client if client else{
            'id_client': None,
            'fullname': 'Non Assigné',
            'telephone': 'Non Assigné'
        }
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }

        # Récupérer les informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}

        # Récupérer les informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        if user_sortie_id:
            user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first()
        else:
            user_sortie = {'fullname': 'null'}
        mouvement['user_srt'] = user_sortie

        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }

        # Assigner des valeurs par défaut pour les autres champs
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')
        mouvement['mission'] = mouvement.get('mission', 'Non assigné')
        mouvement['date_entree'] = mouvement.get('date_entree', 'Non assigné')
        mouvement['date_sortie'] = mouvement.get('date_sortie', 'Non assigné')
        mouvement['destination'] = mouvement.get('destination', 'Non assigné')

    return JsonResponse(mouvement_list, safe=False)

########## SACHERIE ################
def liste_modifs1(request):
    mouvements = Mouvement0.objects.filter(destination__icontains='hangar', date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination', 'transitaire_id', 'mission',
        'client_id', 'marchandise'
    )
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        client_id = mouvement.get('client_id')
        # Récupérer les informations du camion
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        client = Client.objects.filter(id_client=client_id).values('id_client', 'fullname', 'telephone').first()
        mouvement['client'] = client if client else {
            'id_client': None,
            'fullname': 'Non Assigné',
            'telephone': 'Non Assigné'
        }
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }
        # Récupérer les informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}
        # Récupérer les informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        if user_sortie_id:
            user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first()
        else:
            user_sortie = {'fullname': 'null'}
        mouvement['user_srt'] = user_sortie
        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }
        # Assigner des valeurs par défaut pour les autres champs
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')
        mouvement['mission'] = mouvement.get('mission', 'Non assigné')
        mouvement['marchandise'] = mouvement.get('marchandise', 'Non Assigné')
        mouvement['date_entree'] = mouvement.get('date_entree', 'Non assigné')
        mouvement['date_sortie'] = mouvement.get('date_sortie', 'Non assigné')
        mouvement['destination'] = mouvement.get('destination', 'Non assigné')
    return JsonResponse(mouvement_list, safe=False)
########## ZUD ################
def liste_modifs2(request):
    mouvements = Mouvement0.objects.filter(destination__icontains='zud', date_sortie__isnull=True).values(
        'id_mvt', 'camion_id', 'statut_entree', 'statut_sortie', 'chauffeur_id', 'remorque', 'date_entree',
        'date_sortie', 'pointeur_sortie_id', 'pointeur_entree_id', 'destination', 'transitaire_id','mission', 'client_id', 'marchandise', 'representant_id','transitaire_id',
    )
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        ############################# SELECTION DU CAMION #####################
        camion_id = mouvement.get('camion_id')
        camion = Camion.objects.filter(id_cam=camion_id).values('id_cam', 'immatriculation', 'transporteur',
                                                                'type').first()
        mouvement['camion'] = camion if camion else {
            'id_cam': None,
            'immatriculation': 'Non assigné',
            'transporteur': 'Non assigné',
            'type': 'Non assigné'
        }
        ###################### SELECT DU CLIENT ###################
        client_id = mouvement.get('client_id')
        client = Client.objects.filter(id_client=client_id).values('id_client', 'fullname', 'telephone').first()
        mouvement['client'] = client if client else {
            'id_client': None,
            'fullname': 'Non Assigné',
            'telephone': 'Non Assigné'
        }
        ####################### SELECTION DU TRANSITAIRE #####################
        transitaire_id = mouvement.get('transitaire_id')
        transitaire = Transitaire.objects.filter(id_transit=transitaire_id).values('id_transit', 'fullname',
                                                                                   'telephone').first()
        mouvement['transitaire'] = transitaire if transitaire else {
            'id_transit': None,
            'fullname': 'Non A§ssigné',
            'telephone': 'Non Assigné'
        }
        ###################### SELECTION DU REPRESENTANT ###########################
        representant_id = mouvement.get('representant_id')
        representant = Transitaire.objects.filter(id_transit=representant_id).values('id_transit', 'fullname', 'telephone').first()
        mouvement['representant'] = representant if representant else {
            'id_transit': None,
            'fullname': 'Non Assigné',
            'telephone': 'Non Assigné'
        }

        # Récupérer les informations du pointeur d'entrée
        user_entre_id = mouvement.get('pointeur_entree_id')
        user_entre = Utilisateurs.objects.filter(id_user=user_entre_id).values('fullname').first()
        mouvement['user_ert'] = user_entre if user_entre else {'fullname': 'Non assigné'}
        # Récupérer les informations du pointeur de sortie
        user_sortie_id = mouvement.get('pointeur_sortie_id')
        if user_sortie_id:
            user_sortie = Utilisateurs.objects.filter(id_user=user_sortie_id).values('fullname').first()
        else:
            user_sortie = {'fullname': 'null'}
        mouvement['user_srt'] = user_sortie
        # Récupérer les informations du chauffeur
        chauffeur_id = mouvement.get('chauffeur_id')
        chauffeur = Chaffeur.objects.filter(id_chauffeur=chauffeur_id).values('id_chauffeur', 'fullname',
                                                                              'permis').first()
        mouvement['chauffeur'] = chauffeur if chauffeur else {
            'id_chauffeur': None,
            'fullname': 'Non assigné',
            'permis': 'N/A'
        }
        # Assigner des valeurs par défaut pour les autres champs
        mouvement['remorque'] = mouvement.get('remorque', 'Non assigné')
        mouvement['mission'] = mouvement.get('mission', 'Non assigné')
        mouvement['marchandise'] = mouvement.get('marchandise', 'Non Assigné')
        mouvement['date_entree'] = mouvement.get('date_entree', 'Non assigné')
        mouvement['date_sortie'] = mouvement.get('date_sortie', 'Non assigné')
        mouvement['destination'] = mouvement.get('destination', 'Non assigné')
    return JsonResponse(mouvement_list, safe=False)
########## PARTICULIER ################
def liste_modifspar(request):
    mouvements = Mouvement8.objects.filter(date_sortie__isnull=True).values('id_mvt', 'destination', 'vehicule_id',
                                                                            'date_entree')
    mouvement_list = list(mouvements)
    for mouvement in mouvement_list:
        vehicule_id = mouvement['vehicule_id']
        vehicule = Vehicule.objects.filter(id_veh=vehicule_id).values('id_veh', 'immatriculation').first()
        mouvement['vehicule'] = vehicule
    response_data = {
        'mouvements': mouvement_list,
        # Inclure l'ID de l'utilisateur dans la réponse JSON
    }
    return JsonResponse(response_data)
###################### REDIRECTION DES MESSAGES ERREUR #################
def handle_errors(request, exception=None, template_name="pages/404.html", status_code=500):
    """
    Vue générique pour afficher une page d'erreur.
    """
    context = {
        'code': status_code,  # Le code d'erreur comme 404, 500, etc.
        'message': '',
    }
    if status_code == 404:
        context['message'] = "Oups ! La page que vous cherchez n'existe pas."
    elif status_code == 403:
        context['message'] = "Désolé, vous n'avez pas accès à cette page."
    elif status_code == 500:
        context['message'] = "Erreur interne du serveur."
    elif status_code == 400:
        context['message'] = "Mauvaise requête."
    return render(request, template_name, context, status=status_code)
def help_render(request, page):
   # page = request.POST.get('page')
    if page== 'acceuil':
        return render(request, 'pages/RAPPORT/acceuil.html')
    if page == 'active_user':
        return render(request, 'pages/RAPPORT/activer-utilisateurs.html')
    if page == 'entree_icd':
        return render(request, 'pages/RAPPORT/entree-icd.html')
    if page == 'entree_zud':
        return render(request, 'pages/RAPPORT/entree-zud.html')
    if page == 'entree_sacherie':
        return render(request, 'pages/RAPPORT/entree-sacherie.html')
    if page == 'entree_par':
        return render(request, 'pages/RAPPORT/entree-particulier.html')
    if page == 'liaison_zud':
        return render(request, 'pages/RAPPORT/entree-zud-liaison.html')
    if page == 'sortie_icd':
        return render(request, 'pages/RAPPORT/sortie-icd.html')
    if page == 'sortie_sacherie':
        return render(request, 'pages/RAPPORT/sortie-sacherie.html')
    if page == 'sortie_zud':
        return render(request, 'pages/RAPPORT/sortie-zud.html')
    if page == 'sortie_par':
        return render(request, 'pages/RAPPORT/sortie-particulier.html')
    if page == 'modif':
        return render(request, 'pages/RAPPORT/modification.html')
    if page == 'extraction':
        return render(request, 'pages/RAPPORT/extraction.html')
    if page == 'parametrage':
        return render(request, 'pages/RAPPORT/parametrage.html')
    if page == 'icd_entree':
        return render(request, 'pages/RAPPORT/icd-entree-icd.html')
    if page == 'icd_sortie':
        return render(request, 'pages/RAPPORT/icd-sortie-icd.html')
    if page == 'sacherie_entree':
        return render(request, 'pages/RAPPORT/sacherie-entree-sacherie.html')
    if page == 'sacherie_sortie':
        return render(request, 'pages/RAPPORT/sacherie-sortie-sacherie.html')
    if page == 'zud_entree':
        return render(request, 'pages/RAPPORT/zud-entree-zud.html')
    if page == 'zud_sortie':
        return render(request, 'pages/RAPPORT/zud-sortie-zud.html')
    if page == 'controle1':
            return render(request, 'pages/RAPPORT/controle1.html')
    if page == 'controle2':
        return render(request, 'pages/RAPPORT/controle2.html')
    if page == 'ajout_user':
        return render(request, 'pages/RAPPORT/ajout-utilisateurs.html')
    if page == 'modif_user':
        return render(request, 'pages/RAPPORT/modifier-utilisateurs.html')
    if page == 'active_user':
        return render(request, 'pages/RAPPORT/active-utilisateurs.html')
